from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, BackgroundTasks, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
import secrets
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import json
import ftplib
import asyncio
import httpx
import math
import unicodedata
import re
from contextlib import asynccontextmanager
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# Import security utilities
from security import (
    validate_password_strength,
    LoginAttemptTracker,
    AuditLogger,
    sanitize_string,
    sanitize_phone,
    sanitize_email,
    validate_jwt_secret
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Exports directory for Excel files
EXPORTS_DIR = ROOT_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

# FTP Configuration - stored in DB or environment
FTP_HOST = os.environ.get('FTP_HOST', '')
FTP_USER = os.environ.get('FTP_USER', '')
FTP_PASS = os.environ.get('FTP_PASS', '')
FTP_FOLDER = os.environ.get('FTP_FOLDER', '/dolgozok_backup')

# Global FTP config that can be updated at runtime
ftp_config = {
    "host": FTP_HOST,
    "user": FTP_USER,
    "password": FTP_PASS,
    "folder": FTP_FOLDER
}

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration - MUST be set in environment!
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError(
        "❌ KRITIKUS HIBA: JWT_SECRET nincs beállítva!\n"
        "Állítsd be a backend/.env file-ban:\n"
        "JWT_SECRET=\"ide-irj-egy-minimum-32-karakteres-random-stringet\"\n"
        "Példa generálás: python -c 'import secrets; print(secrets.token_urlsafe(32))'"
    )

if not validate_jwt_secret(JWT_SECRET):
    raise ValueError(
        "❌ KRITIKUS HIBA: A JWT_SECRET túl gyenge!\n"
        "Használj minimum 32 karakteres, véletlenszerű stringet.\n"
        "Példa generálás: python -c 'import secrets; print(secrets.token_urlsafe(32))'"
    )

JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

# Rate Limiting
limiter = Limiter(key_func=get_remote_address)

# Security utilities
login_tracker = LoginAttemptTracker(db, max_attempts=5, lockout_minutes=15)
audit_logger = AuditLogger(db)

app = FastAPI(title="Dolgozó CRM API - Biztonságos Verzió")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: Optional[str] = ""
    role: str = "user"  # "admin" or "user" (toborzó)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class ProfileUpdate(BaseModel):
    name: str

class WorkerTypeCreate(BaseModel):
    name: str

class WorkerTypeResponse(BaseModel):
    id: str
    name: str

class CategoryCreate(BaseModel):
    name: str
    color: str = "#3b82f6"  # Default blue

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    order: Optional[int] = None

class CategoryResponse(BaseModel):
    id: str
    name: str
    color: str = "#3b82f6"
    order: int = 0
    worker_count: int = 0

class PositionCreate(BaseModel):
    name: str
    worker_type_id: str  # Melyik típushoz tartozik

class PositionResponse(BaseModel):
    id: str
    name: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""

class StatusCreate(BaseModel):
    name: str
    status_type: str = "neutral"  # ÚJ: "positive", "negative", "neutral"
    color: str = "#6b7280"  # ÚJ: Szín

class StatusUpdate(BaseModel):
    name: Optional[str] = None
    status_type: Optional[str] = None  # ÚJ
    color: Optional[str] = None  # ÚJ

class StatusResponse(BaseModel):
    id: str
    name: str
    status_type: str = "neutral"  # ÚJ
    color: str = "#6b7280"  # ÚJ

class TagCreate(BaseModel):
    name: str
    color: str = "#6366f1"

class TagResponse(BaseModel):
    id: str
    name: str
    color: str

class WorkerCreate(BaseModel):
    name: str
    phone: str
    worker_type_id: str
    position: Optional[str] = ""  # Szabad szöveg pozíció
    position_experience: Optional[str] = ""  # Pozícióval kapcsolatos tapasztalat
    address: Optional[str] = ""
    email: Optional[str] = ""
    experience: Optional[str] = ""
    notes: Optional[str] = ""
    global_status: str = "Feldolgozatlan"  # Alap: még nincs kommunikáció
    properties: Optional[List[str]] = []  # Tulajdonságok: megbizhato, jo_munkaero, stb.
    project_id: Optional[str] = None  # Opcionális projekt várólistához
    trial_date: Optional[str] = None  # Próba időpont (volt: start_date)
    latitude: Optional[float] = None  # Geocoding koordináta
    longitude: Optional[float] = None  # Geocoding koordináta
    county: Optional[str] = ""  # Megye
    work_type: Optional[str] = ""  # Munkavégzés típusa: "Ingázó", "Szállásos", ""
    has_car: Optional[str] = ""  # Saját autó: "Van", "Nincs", ""
    initial_status: Optional[str] = None  # Projekt státusz létrehozáskor (ha project_id van)
    gender: Optional[str] = None  # Nem: "férfi", "nő"
    # GDPR mezők
    consent_given: Optional[bool] = False  # Beleegyezés adott
    consent_date: Optional[str] = None  # Beleegyezés dátuma
    processing_basis: Optional[str] = "legitimate_interest"  # jogos_erdek, beleegyezes, szerzodes

class WorkerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    worker_type_id: Optional[str] = None
    position: Optional[str] = None
    position_experience: Optional[str] = None
    address: Optional[str] = None
    email: Optional[str] = None
    experience: Optional[str] = None
    notes: Optional[str] = None
    global_status: Optional[str] = None
    properties: Optional[List[str]] = None  # Tulajdonságok
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    county: Optional[str] = None
    work_type: Optional[str] = None  # Munkavégzés típusa
    has_car: Optional[str] = None  # Saját autó
    gender: Optional[str] = None  # Nem
    # GDPR mezők
    consent_given: Optional[bool] = None
    consent_date: Optional[str] = None
    processing_basis: Optional[str] = None

class WorkerResponse(BaseModel):
    id: str
    name: str
    phone: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""
    position: Optional[str] = ""
    position_experience: Optional[str] = ""
    address: str = ""
    email: str = ""
    experience: str = ""
    notes: str = ""
    global_status: str = "Feldolgozatlan"
    properties: List[str] = []  # Tulajdonságok
    tags: List[dict] = []
    project_statuses: List[dict] = []
    owner_id: str
    owner_name: str
    created_at: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    county: Optional[str] = ""
    work_type: Optional[str] = ""  # Munkavégzés típusa: Ingázó/Szállásos
    has_car: Optional[str] = ""  # Saját autó: Van/Nincs
    distance_km: Optional[float] = None  # Távolság km-ben (ha szűrés aktív)
    gender: Optional[str] = None  # Nem: "férfi", "nő"
    # GDPR mezők
    consent_given: Optional[bool] = False
    consent_date: Optional[str] = None
    processing_basis: Optional[str] = "legitimate_interest"
    data_retention_until: Optional[str] = None  # Adatmegőrzés lejárata

class ProjectCreate(BaseModel):
    name: str
    client_name: Optional[str] = ""  # Ügyfél / cégnév
    date: str
    location: Optional[str] = ""  # Helyszín
    training_location: Optional[str] = ""  # Betanítás / munkavégzés helye
    notes: Optional[str] = ""  # Megjegyzések, elvárások
    recruiter_ids: List[str] = []  # Hozzárendelt toborzók
    planned_headcount: Optional[int] = 0  # Tervezett létszám

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    client_name: Optional[str] = None
    date: Optional[str] = None
    location: Optional[str] = None
    training_location: Optional[str] = None
    notes: Optional[str] = None
    is_closed: Optional[bool] = None
    recruiter_ids: Optional[List[str]] = None
    planned_headcount: Optional[int] = None  # Tervezett létszám

class ProjectResponse(BaseModel):
    id: str
    name: str
    client_name: str = ""
    date: str
    location: str
    training_location: str = ""
    notes: str
    is_closed: bool
    worker_count: int
    position_count: int = 0
    total_headcount: int = 0  # Összesített létszámigény pozíciókból
    trial_count: int = 0
    recruiter_ids: List[str] = []
    recruiters: List[dict] = []
    owner_id: str = ""
    owner_name: str = ""
    created_at: str
    planned_headcount: int = 0  # Tervezett létszám
    active_worker_count: int = 0  # Aktív dolgozók (Dolgozik státusz)

# ==================== PROJECT POSITION MODELS ====================

class ProjectPositionCreate(BaseModel):
    name: str  # Pozíció neve (pl. Operátor, Raktáros)
    headcount: int = 1  # Létszámigény
    work_schedule: Optional[str] = ""  # Munkarend (volt: shift_schedule)
    salary: Optional[str] = ""  # Bérezés (pl. 2500 Ft/óra bruttó)
    experience_required: Optional[str] = ""  # Tapasztalat
    qualifications: Optional[str] = ""  # Végzettség / jogosítvány
    physical_requirements: Optional[str] = ""  # Fizikai elvárások
    position_details: Optional[str] = ""  # Pozíció részletei
    notes: Optional[str] = ""  # Egyéb megjegyzések

class ProjectPositionUpdate(BaseModel):
    name: Optional[str] = None
    headcount: Optional[int] = None
    work_schedule: Optional[str] = None  # Munkarend (volt: shift_schedule)
    salary: Optional[str] = None  # Bérezés
    experience_required: Optional[str] = None
    qualifications: Optional[str] = None
    physical_requirements: Optional[str] = None
    position_details: Optional[str] = None  # Pozíció részletei
    notes: Optional[str] = None

class ProjectPositionResponse(BaseModel):
    id: str
    project_id: str
    name: str
    headcount: int
    work_schedule: str = ""  # Munkarend (volt: shift_schedule)
    salary: str = ""  # Bérezés
    experience_required: str = ""
    qualifications: str = ""
    physical_requirements: str = ""
    position_details: str = ""  # Pozíció részletei
    notes: str = ""
    assigned_workers: int = 0
    created_at: str

# ==================== TRIAL MODELS ====================

class TrialCreate(BaseModel):
    date: str  # Próba dátuma
    time: Optional[str] = ""  # Próba időpontja (pl. "09:00")
    notes: Optional[str] = ""

class TrialUpdate(BaseModel):
    date: Optional[str] = None
    time: Optional[str] = None
    notes: Optional[str] = None

class TrialPositionCreate(BaseModel):
    position_id: Optional[str] = None  # Meglévő projekt pozíció ID
    position_name: str  # Pozíció neve (új vagy meglévő)
    headcount: int = 1  # Létszámigény
    hourly_rate: Optional[str] = ""  # Nettó órabér
    accommodation: bool = False  # Van-e szállás
    requirements: str = ""  # Egyéb elvárások
    add_to_project: bool = False  # Új pozíció esetén hozzáadjuk a projekthez?

class TrialPositionUpdate(BaseModel):
    position_name: Optional[str] = None
    headcount: Optional[int] = None
    hourly_rate: Optional[str] = None  # Nettó órabér
    accommodation: Optional[bool] = None
    requirements: Optional[str] = None

class TrialPositionResponse(BaseModel):
    id: str
    trial_id: str
    position_id: Optional[str] = None
    position_name: str
    headcount: int = 1
    hourly_rate: str = ""
    accommodation: bool = False
    requirements: str = ""
    assigned_count: int = 0  # Hány dolgozó van hozzárendelve

class TrialResponse(BaseModel):
    id: str
    project_id: str
    date: str
    time: str = ""
    notes: str = ""
    worker_count: int = 0
    workers: List[dict] = []
    positions: List[TrialPositionResponse] = []  # Próba pozíciók
    created_at: str

class TrialWorkerAdd(BaseModel):
    worker_id: str
    position_id: Optional[str] = None  # Melyik próba pozícióra (trial_position_id)

class ProjectWorkerAdd(BaseModel):
    worker_id: str
    status_id: Optional[str] = None
    position_ids: Optional[List[str]] = []  # ÚJ: Többszörös pozíció választás
    force_add: Optional[bool] = False  # Duplikáció figyelmeztetés után mégis hozzáadás
    from_form: Optional[bool] = False  # Űrlapról érkezett
    trial_id: Optional[str] = None  # ÚJ: Próba ID (ha "Próba megbeszélve")
    trial_position_id: Optional[str] = None  # ÚJ: Próba pozíció ID

class ProjectRecruiterAdd(BaseModel):
    user_id: str

class ProjectWorkerStatusUpdate(BaseModel):
    status_id: str
    notes: Optional[str] = None
    trial_id: Optional[str] = None  # Melyik próbához van rendelve (ha "Próba megbeszélve")

class WorkerHistoryEntry(BaseModel):
    project_id: str
    project_name: str
    project_date: str
    status_id: str
    status_name: str
    notes: str
    updated_at: str

# ==================== WAITLIST MODELS ====================

class WaitlistWorkerAdd(BaseModel):
    worker_id: str
    trial_date: Optional[str] = None  # Próba időpont (volt: start_date)
    notes: Optional[str] = ""

class WaitlistWorkerUpdate(BaseModel):
    trial_date: Optional[str] = None  # Próba időpont (volt: start_date)
    notes: Optional[str] = None

class WaitlistWorkerResponse(BaseModel):
    id: str
    project_id: str
    worker_id: str
    worker_name: str
    worker_phone: str
    worker_email: str = ""
    trial_date: str = ""  # Próba időpont (volt: start_date)
    notes: str = ""
    added_at: str
    added_by: str
    added_by_name: str = ""

# ==================== FORM MODELS ====================

class FormCreate(BaseModel):
    sheet_url: str
    name: Optional[str] = ""
    column_mapping: Dict[str, str]  # {"name": "B", "phone": "C", "position": "D", ...}
    default_category: str = "Felvitt dolgozók"
    default_position_id: Optional[str] = ""  # Alapértelmezett pozíció ID
    sync_frequency: str = "hourly"  # hourly, manual

class FormUpdate(BaseModel):
    name: Optional[str] = None
    column_mapping: Optional[Dict[str, str]] = None
    default_category: Optional[str] = None
    default_position_id: Optional[str] = None
    sync_frequency: Optional[str] = None

class FormResponse(BaseModel):
    id: str
    project_id: str
    sheet_url: str
    name: str
    column_mapping: Dict[str, str]
    owner_id: str
    owner_name: str = ""
    shared_with: List[str] = []
    shared_with_names: List[str] = []
    default_category: str
    default_position_id: Optional[str] = ""
    sync_frequency: str
    last_synced_at: Optional[str] = None
    last_row_processed: int = 0
    lead_count: int = 0
    created_at: str

class FormShareRequest(BaseModel):
    shared_with: List[str]  # User IDs

class FormLeadResponse(BaseModel):
    id: str
    form_id: str
    project_id: str
    name: str
    phone: str
    address: Optional[str] = ""
    email: Optional[str] = ""
    notes: Optional[str] = ""
    submitted_at: Optional[str] = ""
    status: str  # unprocessed, duplicate, processed, ignored
    is_duplicate: bool = False
    duplicate_worker_id: Optional[str] = None
    duplicate_worker: Optional[Dict] = None
    created_at: str

class FormLeadResolve(BaseModel):
    action: str  # keep_both, keep_existing, keep_new, merge
    merge_fields: Optional[List[str]] = None

# ==================== NOTIFICATION MODELS ====================

class NotificationResponse(BaseModel):
    id: str
    user_id: str
    type: str  # "project_assigned", "trial_assigned"
    title: str
    message: str
    link: str = ""
    is_read: bool = False
    created_at: str

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def create_notification(user_id: str, notification_type: str, title: str, message: str, link: str = ""):
    """Helper function to create a notification"""
    notif_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "type": notification_type,
        "title": title,
        "message": message,
        "link": link,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notif_doc)
    return notif_doc

# ==================== GEOCODING HELPER ====================

HUNGARIAN_COUNTIES = [
    "Bács-Kiskun", "Baranya", "Békés", "Borsod-Abaúj-Zemplén", "Budapest",
    "Csongrád-Csanád", "Fejér", "Győr-Moson-Sopron", "Hajdú-Bihar", "Heves",
    "Jász-Nagykun-Szolnok", "Komárom-Esztergom", "Nógrád", "Pest", "Somogy",
    "Szabolcs-Szatmár-Bereg", "Tolna", "Vas", "Veszprém", "Zala"
]

async def geocode_address(address: str) -> dict:
    """Geocode an address using OpenStreetMap Nominatim"""
    if not address or len(address.strip()) < 3:
        return {"latitude": None, "longitude": None, "county": None}
    
    try:
        async with httpx.AsyncClient() as client:
            # Add Hungary to improve results
            search_query = f"{address}, Magyarország"
            response = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={
                    "q": search_query,
                    "format": "json",
                    "limit": 1,
                    "addressdetails": 1,
                    "countrycodes": "hu"
                },
                headers={"User-Agent": "DolgozoCRM/1.0"},
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data and len(data) > 0:
                    result = data[0]
                    lat = float(result.get("lat", 0))
                    lon = float(result.get("lon", 0))
                    
                    # Extract county from address details
                    address_details = result.get("address", {})
                    county = (
                        address_details.get("county", "") or 
                        address_details.get("state", "") or
                        address_details.get("city", "") or
                        address_details.get("town", "") or
                        address_details.get("village", "")
                    )
                    
                    # Clean up county name
                    if county:
                        county = county.replace(" megye", "").replace(" vármegye", "")
                    
                    return {
                        "latitude": lat,
                        "longitude": lon,
                        "county": county,
                        "display_name": result.get("display_name", "")
                    }
    except Exception as e:
        logging.error(f"Geocoding error for '{address}': {e}")
    
    return {"latitude": None, "longitude": None, "county": None}

def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great circle distance between two points in kilometers"""
    if not all([lat1, lon1, lat2, lon2]):
        return float('inf')
    
    R = 6371  # Earth's radius in kilometers
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

# ==================== SEARCH HELPER FUNCTIONS ====================

def remove_accents(text: str) -> str:
    """
    Remove Hungarian accents from text for accent-insensitive search
    Example: "Gépkezelő" -> "Gepkezelo"
    """
    if not text:
        return ""
    # Normalize to NFD (decomposed form) and filter out combining characters
    nfd = unicodedata.normalize('NFD', text)
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')

def create_flexible_search_regex(search_term: str) -> str:
    """
    Create MongoDB regex pattern for flexible, accent-insensitive search
    Splits search term into words and creates pattern that matches any word
    
    Example:
    Input: "operátor"
    Output: Matches "CNC Gépkezelő", "Gépkezelő operátor", "Operátor"
    """
    if not search_term:
        return ""
    
    # Remove accents from search term
    normalized_search = remove_accents(search_term.lower())
    
    # Split into words
    words = normalized_search.split()
    
    # Create regex pattern that matches any of the words
    # (?i) for case insensitive
    if len(words) == 1:
        # Single word: match if it appears anywhere
        return f".*{re.escape(words[0])}.*"
    else:
        # Multiple words: match if ANY word appears
        word_patterns = [f".*{re.escape(word)}.*" for word in words]
        return "|".join(word_patterns)

def normalize_text_for_search(text: str) -> str:
    """
    Normalize text for storage in search-optimized field
    Used for creating searchable versions of names, positions, etc.
    """
    if not text:
        return ""
    return remove_accents(text.lower())

# ==================== AI-ALAPÚ GENDER DETECTION GROQ LLAMA-VAL ====================
# VALÓBAN INGYENES API! (Groq Cloud - llama3-8b-8192)

from groq import Groq

# Groq API - TELJESEN INGYENES!
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')

def detect_gender_from_name(full_name: str) -> Optional[str]:
    """
    VALÓDI AI-alapú gender detection Groq Llama3 modellel
    100% INGYENES, KORLÁTLAN használat!
    
    Returns: "férfi", "nő", or None
    
    Groq Cloud API: Ingyenes, gyors LLM inference
    Model: llama3-8b-8192 (teljesen ingyenes!)
    """
    if not full_name or len(full_name.strip()) < 2:
        return None
    
    # Ha nincs API key, fallback
    if not GROQ_API_KEY:
        logger.warning("GROQ_API_KEY nincs beállítva - fallback heurisztikára")
        return _fallback_gender_detection(full_name)
    
    try:
        # Groq client inicializálás
        client = Groq(api_key=GROQ_API_KEY)
        
        # AI prompt - rövid és egyértelmű
        prompt = f"""Név: "{full_name}"

Magyar név neme? Válasz CSAK: férfi VAGY nő"""

        # Groq Llama3 API hívás (INGYENES!)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",  # Újabb, aktív model!
            messages=[
                {"role": "system", "content": "Te egy magyar név gender detector vagy. Csak 'férfi' vagy 'nő' szóval válaszolj."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=5,
            temperature=0.0  # Determinisztikus
        )
        
        # Válasz feldolgozása
        result = response.choices[0].message.content.strip().lower()
        
        if "férfi" in result or "ferfi" in result:
            return "férfi"
        elif "nő" in result or "no" in result:
            return "nő"
        else:
            return _fallback_gender_detection(full_name)
            
    except Exception as e:
        logger.warning(f"Groq AI gender detection hiba '{full_name}': {e}")
        return _fallback_gender_detection(full_name)

def _fallback_gender_detection(full_name: str) -> Optional[str]:
    """Fallback heurisztika ha az AI nem elérhető"""
    name_lower = full_name.lower().strip()
    
    # "né" utótag -> női
    if " né " in name_lower or name_lower.endswith("né"):
        return "nő"
    
    # Keresztnév vizsgálata (második szó magyar sorrendben)
    words = name_lower.split()
    if len(words) >= 2:
        last_word = words[-1]
        
        # Női végződések
        if last_word.endswith('a') and len(last_word) >= 3:
            if last_word not in ['béla', 'bela', 'gyula', 'attila']:
                return "nő"
        
        # Férfi végződések
        if len(last_word) >= 3 and last_word[-1] in 'bcdfghjklmnpqrstvwxyz':
            return "férfi"
    
    return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Felhasználó nem található")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token lejárt")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Érvénytelen token")

async def require_admin(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin jogosultsággal")
    return user

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=dict)
async def register(data: UserCreate, current_user: dict = Depends(require_admin), request: Request = None):
    """Admin csak hozhat létre új felhasználót - BIZTONSÁGOS"""
    # Sanitize inputs
    email = sanitize_email(data.email)
    name = sanitize_string(data.name, max_length=100)
    
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Ez az email már regisztrálva van")
    
    # Strong password validation
    is_valid, error_msg = validate_password_strength(data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password": hash_password(data.password),
        "name": name or email.split("@")[0],
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Audit log
    ip_addr = request.client.host if request else "unknown"
    await audit_logger.log(
        user_id=current_user["id"],
        user_email=current_user["email"],
        action="created",
        resource_type="user",
        resource_id=user_doc["id"],
        details={"new_user_email": email, "new_user_role": data.role},
        ip_address=ip_addr
    )
    
    return {"message": "Felhasználó létrehozva", "email": email}

@api_router.post("/auth/login", response_model=dict)
@limiter.limit("5/minute")  # Max 5 login attempts per minute per IP
async def login(data: UserLogin, request: Request):
    """Login with rate limiting and account lockout - BIZTONSÁGOS"""
    email = sanitize_email(data.email)
    ip_addr = request.client.host if request else "unknown"
    
    # Check if account is locked out
    is_locked, unlock_time = await login_tracker.is_locked_out(email)
    if is_locked:
        remaining_minutes = int((unlock_time - datetime.now(timezone.utc)).total_seconds() / 60)
        raise HTTPException(
            status_code=429,
            detail=f"Túl sok sikertelen bejelentkezési kísérlet. Próbáld újra {remaining_minutes} perc múlva."
        )
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        # Record failed attempt
        await login_tracker.record_failed_attempt(email, ip_addr)
        remaining = await login_tracker.get_remaining_attempts(email)
        
        if remaining > 0:
            raise HTTPException(
                status_code=401,
                detail=f"Hibás email vagy jelszó. Még {remaining} próbálkozásod van."
            )
        else:
            raise HTTPException(
                status_code=429,
                detail="Túl sok sikertelen kísérlet. A fiók 15 percre zárolva."
            )
    
    # Successful login
    await login_tracker.record_successful_attempt(email, ip_addr)
    
    # Audit log
    await audit_logger.log(
        user_id=user["id"],
        user_email=user["email"],
        action="login",
        resource_type="auth",
        details={"success": True},
        ip_address=ip_addr
    )
    
    token = create_token(user["id"], user["email"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"]
        }
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user.get("name", ""),
        role=user["role"],
        created_at=user.get("created_at", "")
    )

@api_router.get("/auth/audit-logs")
async def get_audit_logs(
    user: dict = Depends(require_admin),
    resource_type: Optional[str] = None,
    limit: int = 100
):
    """Admin audit log lekérése - CSAK ADMIN"""
    logs = await audit_logger.get_logs(resource_type=resource_type, limit=limit)
    return {"logs": logs, "count": len(logs)}

@api_router.get("/auth/my-activity")
async def get_my_activity(user: dict = Depends(get_current_user)):
    """Saját aktivitás lekérése"""
    logs = await audit_logger.get_recent_activity(user["id"], days=30)
    return {"logs": logs, "count": len(logs)}

@api_router.put("/auth/profile")
async def update_profile(data: ProfileUpdate, user: dict = Depends(get_current_user)):
    """Csak admin tudja módosítani a profilt"""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin módosíthatja a profilt")
    await db.users.update_one({"id": user["id"]}, {"$set": {"name": data.name}})
    return {"message": "Profil frissítve"}

@api_router.put("/auth/password")
async def change_password(data: PasswordChange, user: dict = Depends(get_current_user), request: Request = None):
    """Change password with strong validation - BIZTONSÁGOS"""
    db_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not verify_password(data.current_password, db_user["password"]):
        raise HTTPException(status_code=400, detail="Hibás jelenlegi jelszó")
    
    # Strong password validation
    is_valid, error_msg = validate_password_strength(data.new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    await db.users.update_one(
        {"id": user["id"]}, 
        {"$set": {"password": hash_password(data.new_password)}}
    )
    
    # Audit log
    ip_addr = request.client.host if request else "unknown"
    await audit_logger.log(
        user_id=user["id"],
        user_email=user["email"],
        action="password_changed",
        resource_type="auth",
        ip_address=ip_addr
    )
    
    return {"message": "Jelszó megváltoztatva"}

# ==================== WORKER TYPES ====================

@api_router.get("/worker-types", response_model=List[WorkerTypeResponse])
async def get_worker_types(user: dict = Depends(get_current_user)):
    types = await db.worker_types.find({}, {"_id": 0}).to_list(100)
    return [WorkerTypeResponse(**t) for t in types]

@api_router.post("/worker-types", response_model=WorkerTypeResponse)
async def create_worker_type(data: WorkerTypeCreate, user: dict = Depends(require_admin)):
    type_doc = {"id": str(uuid.uuid4()), "name": data.name}
    await db.worker_types.insert_one(type_doc)
    return WorkerTypeResponse(**type_doc)

@api_router.delete("/worker-types/{type_id}")
async def delete_worker_type(type_id: str, user: dict = Depends(require_admin)):
    result = await db.worker_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Típus nem található")
    # Töröljük a típushoz tartozó pozíciókat is
    await db.positions.delete_many({"worker_type_id": type_id})
    return {"message": "Típus törölve"}

# ==================== POSITIONS ====================

@api_router.get("/positions", response_model=List[PositionResponse])
async def get_positions(worker_type_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    """Pozíciók lekérése, opcionálisan típus szerint szűrve"""
    query = {}
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    
    positions = await db.positions.find(query, {"_id": 0}).to_list(100)
    
    result = []
    for p in positions:
        type_doc = await db.worker_types.find_one({"id": p.get("worker_type_id")}, {"_id": 0})
        result.append(PositionResponse(
            id=p["id"],
            name=p["name"],
            worker_type_id=p["worker_type_id"],
            worker_type_name=type_doc["name"] if type_doc else ""
        ))
    return result

@api_router.post("/positions", response_model=PositionResponse)
async def create_position(data: PositionCreate, user: dict = Depends(require_admin)):
    # Ellenőrizzük, hogy létezik-e a típus
    type_doc = await db.worker_types.find_one({"id": data.worker_type_id}, {"_id": 0})
    if not type_doc:
        raise HTTPException(status_code=404, detail="Típus nem található")
    
    position_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "worker_type_id": data.worker_type_id
    }
    await db.positions.insert_one(position_doc)
    return PositionResponse(**position_doc, worker_type_name=type_doc["name"])

@api_router.delete("/positions/{position_id}")
async def delete_position(position_id: str, user: dict = Depends(require_admin)):
    result = await db.positions.delete_one({"id": position_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    return {"message": "Pozíció törölve"}

# ==================== STATUSES ====================

@api_router.get("/statuses", response_model=List[StatusResponse])
async def get_statuses(user: dict = Depends(get_current_user)):
    statuses = await db.statuses.find({}, {"_id": 0}).to_list(100)
    return [StatusResponse(
        id=s["id"],
        name=s["name"],
        status_type=s.get("status_type", "neutral"),
        color=s.get("color", "#6b7280")
    ) for s in statuses]

@api_router.post("/statuses", response_model=StatusResponse)
async def create_status(data: StatusCreate, user: dict = Depends(require_admin)):
    """Státusz létrehozása - CSAK ADMIN"""
    status_doc = {
        "id": str(uuid.uuid4()),
        "name": sanitize_string(data.name, max_length=100),
        "status_type": data.status_type,
        "color": data.color
    }
    await db.statuses.insert_one(status_doc)
    return StatusResponse(**status_doc)

@api_router.put("/statuses/{status_id}", response_model=StatusResponse)
async def update_status(status_id: str, data: StatusUpdate, user: dict = Depends(require_admin)):
    """Státusz szerkesztése - CSAK ADMIN"""
    status = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    if not status:
        raise HTTPException(status_code=404, detail="Státusz nem található")
    
    update_data = {}
    if data.name is not None:
        update_data["name"] = sanitize_string(data.name, max_length=100)
    if data.status_type is not None:
        update_data["status_type"] = data.status_type
    if data.color is not None:
        update_data["color"] = data.color
    
    if update_data:
        await db.statuses.update_one({"id": status_id}, {"$set": update_data})
    
    updated = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    return StatusResponse(
        id=updated["id"],
        name=updated["name"],
        status_type=updated.get("status_type", "neutral"),
        color=updated.get("color", "#6b7280")
    )

@api_router.delete("/statuses/{status_id}")
async def delete_status(status_id: str, user: dict = Depends(require_admin)):
    """Státusz törlése - CSAK ADMIN"""
    result = await db.statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Státusz nem található")
    return {"message": "Státusz törölve"}

# ==================== TAGS ====================

@api_router.get("/tags", response_model=List[TagResponse])
async def get_tags(user: dict = Depends(get_current_user)):
    tags = await db.tags.find({}, {"_id": 0}).to_list(100)
    return [TagResponse(**t) for t in tags]

@api_router.post("/tags", response_model=TagResponse)
async def create_tag(data: TagCreate, user: dict = Depends(require_admin)):
    tag_doc = {"id": str(uuid.uuid4()), "name": data.name, "color": data.color}
    await db.tags.insert_one(tag_doc)
    return TagResponse(**tag_doc)

@api_router.delete("/tags/{tag_id}")
async def delete_tag(tag_id: str, user: dict = Depends(require_admin)):
    result = await db.tags.delete_one({"id": tag_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Jellemző nem található")
    return {"message": "Jellemző törölve"}

# ==================== CATEGORIES ====================

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(user: dict = Depends(get_current_user)):
    """Kategóriák lekérése worker count-tal és rendezve"""
    categories = await db.categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    # Ha nincs még kategória, visszaadjuk az alapértelmezetteket
    if not categories:
        default_categories = [
            {"id": str(uuid.uuid4()), "name": "Felvitt dolgozók", "color": "#3b82f6", "order": 0},
            {"id": str(uuid.uuid4()), "name": "Hideg jelentkező", "color": "#22c55e", "order": 1},
            {"id": str(uuid.uuid4()), "name": "Űrlapon jelentkezett", "color": "#f97316", "order": 2},
            {"id": str(uuid.uuid4()), "name": "Állásra jelentkezett", "color": "#a855f7", "order": 3},
            {"id": str(uuid.uuid4()), "name": "Ingázó", "color": "#64748b", "order": 4},
            {"id": str(uuid.uuid4()), "name": "Szállásos", "color": "#f59e0b", "order": 5},
        ]
        await db.categories.insert_many(default_categories)
        categories = default_categories
    
    # Add worker count to each category
    result = []
    for c in categories:
        worker_count = await db.workers.count_documents({"category": c["name"]})
        result.append(CategoryResponse(
            id=c["id"],
            name=c["name"],
            color=c.get("color", "#3b82f6"),
            order=c.get("order", 0),
            worker_count=worker_count
        ))
    return result

@api_router.get("/categories/stats")
async def get_category_stats(user: dict = Depends(get_current_user)):
    """Kategória statisztikák dashboard-hoz"""
    categories = await db.categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Toborzó csak saját dolgozóit látja
    worker_query = {}
    if user["role"] != "admin":
        worker_query["owner_id"] = user["id"]
    
    stats = []
    total_workers = 0
    for cat in categories:
        count_query = {**worker_query, "category": cat["name"]}
        count = await db.workers.count_documents(count_query)
        total_workers += count
        stats.append({
            "id": cat["id"],
            "name": cat["name"],
            "color": cat.get("color", "#3b82f6"),
            "count": count
        })
    
    # Recent activity - workers added in last 7 days per category
    from datetime import timedelta
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    
    recent_stats = []
    for cat in categories:
        recent_query = {**worker_query, "category": cat["name"], "created_at": {"$gte": week_ago}}
        recent_count = await db.workers.count_documents(recent_query)
        recent_stats.append({
            "name": cat["name"],
            "color": cat.get("color", "#3b82f6"),
            "count": recent_count
        })
    
    return {
        "total_workers": total_workers,
        "category_stats": stats,
        "recent_activity": recent_stats,
        "categories_count": len(categories)
    }

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(data: CategoryCreate, user: dict = Depends(require_admin)):
    """Új kategória létrehozása - csak admin"""
    # Ellenőrizzük, hogy ne legyen duplikált név
    existing = await db.categories.find_one({"name": data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Ilyen nevű kategória már létezik")
    
    # Get max order
    max_order_cat = await db.categories.find_one({}, {"_id": 0}, sort=[("order", -1)])
    next_order = (max_order_cat.get("order", 0) + 1) if max_order_cat else 0
    
    category_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "color": data.color,
        "order": next_order
    }
    await db.categories.insert_one(category_doc)
    return CategoryResponse(**category_doc, worker_count=0)

class CategoryOrderItem(BaseModel):
    id: str
    order: int

class CategoryReorderRequest(BaseModel):
    orders: List[CategoryOrderItem]

@api_router.put("/categories/reorder")
async def reorder_categories(data: CategoryReorderRequest, user: dict = Depends(require_admin)):
    """Kategóriák átrendezése - csak admin"""
    for item in data.orders:
        await db.categories.update_one(
            {"id": item.id},
            {"$set": {"order": item.order}}
        )
    return {"message": "Sorrend frissítve"}

@api_router.put("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(category_id: str, data: CategoryUpdate, user: dict = Depends(require_admin)):
    """Kategória szerkesztése - csak admin"""
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Kategória nem található")
    
    update_data = {}
    old_name = category["name"]
    
    if data.name is not None and data.name != old_name:
        # Check for duplicate name
        existing = await db.categories.find_one({"name": data.name, "id": {"$ne": category_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Ilyen nevű kategória már létezik")
        update_data["name"] = data.name
        # Update all workers with old category name
        await db.workers.update_many(
            {"category": old_name},
            {"$set": {"category": data.name}}
        )
    
    if data.color is not None:
        update_data["color"] = data.color
    
    if data.order is not None:
        update_data["order"] = data.order
    
    if update_data:
        await db.categories.update_one({"id": category_id}, {"$set": update_data})
    
    updated = await db.categories.find_one({"id": category_id}, {"_id": 0})
    worker_count = await db.workers.count_documents({"category": updated["name"]})
    return CategoryResponse(**updated, worker_count=worker_count)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user: dict = Depends(require_admin)):
    """Kategória törlése - csak admin"""
    # Ellenőrizzük, hogy nem használja-e dolgozó
    workers_using = await db.workers.count_documents({"category": {"$exists": True}})
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if category:
        workers_with_cat = await db.workers.count_documents({"category": category["name"]})
        if workers_with_cat > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Nem törölhető: {workers_with_cat} dolgozó használja ezt a kategóriát"
            )
    
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kategória nem található")
    return {"message": "Kategória törölve"}

@api_router.get("/global-statuses")
async def get_global_statuses():
    """Get global worker statuses - egységes státuszok"""
    # Egységes státuszok: Feldolgozatlan, Próbára vár, Próba megbeszélve, Dolgozik, Tiltólista
    return [
        {"name": "Feldolgozatlan", "color": "#9CA3AF", "order": 0},
        {"name": "Próbára vár", "color": "#F97316", "order": 1},
        {"name": "Próba megbeszélve", "color": "#8B5CF6", "order": 2},
        {"name": "Dolgozik", "color": "#10B981", "order": 3},
        {"name": "Tiltólista", "color": "#EF4444", "order": 4}
    ]

@api_router.post("/sync-statuses")
async def sync_statuses(user: dict = Depends(require_admin)):
    """Admin: Státuszok szinkronizálása - átnevezi 'Feldolgozás alatt' → 'Feldolgozatlan' és hozzáadja 'Tiltólista' státuszt"""
    changes = []
    
    # 1. "Feldolgozás alatt" → "Feldolgozatlan" átnevezés
    feldolgozas = await db.statuses.find_one({"name": "Feldolgozás alatt"})
    if feldolgozas:
        await db.statuses.update_one(
            {"name": "Feldolgozás alatt"},
            {"$set": {"name": "Feldolgozatlan", "color": "#9CA3AF"}}
        )
        changes.append("Feldolgozás alatt → Feldolgozatlan átnevezve")
    
    # 2. Tiltólista státusz hozzáadása ha nincs
    tiltolista = await db.statuses.find_one({"name": "Tiltólista"})
    if not tiltolista:
        await db.statuses.insert_one({
            "id": str(uuid.uuid4()),
            "name": "Tiltólista",
            "status_type": "negative",
            "color": "#EF4444",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        changes.append("Tiltólista státusz létrehozva")
    
    # 3. Frissítsük a meglévő dolgozók globális státuszát ha szükséges
    # "Feldolgozás alatt" → "Feldolgozatlan"
    result1 = await db.workers.update_many(
        {"global_status": "Feldolgozás alatt"},
        {"$set": {"global_status": "Feldolgozatlan"}}
    )
    if result1.modified_count > 0:
        changes.append(f"{result1.modified_count} dolgozó globális státusza frissítve: Feldolgozás alatt → Feldolgozatlan")
    
    # "Projektben" → próbáljuk meghatározni a pontos státuszt
    result2 = await db.workers.update_many(
        {"global_status": "Projektben"},
        {"$set": {"global_status": "Próbára vár"}}
    )
    if result2.modified_count > 0:
        changes.append(f"{result2.modified_count} dolgozó globális státusza frissítve: Projektben → Próbára vár")
    
    # "Kuka" → "Feldolgozatlan"
    result3 = await db.workers.update_many(
        {"global_status": "Kuka"},
        {"$set": {"global_status": "Feldolgozatlan"}}
    )
    if result3.modified_count > 0:
        changes.append(f"{result3.modified_count} dolgozó globális státusza frissítve: Kuka → Feldolgozatlan")
    
    # "Máshol dolgozik" → "Dolgozik"
    result4 = await db.workers.update_many(
        {"global_status": "Máshol dolgozik"},
        {"$set": {"global_status": "Dolgozik"}}
    )
    if result4.modified_count > 0:
        changes.append(f"{result4.modified_count} dolgozó globális státusza frissítve: Máshol dolgozik → Dolgozik")
    
    # "Inaktív" → "Feldolgozatlan"
    result5 = await db.workers.update_many(
        {"global_status": "Inaktív"},
        {"$set": {"global_status": "Feldolgozatlan"}}
    )
    if result5.modified_count > 0:
        changes.append(f"{result5.modified_count} dolgozó globális státusza frissítve: Inaktív → Feldolgozatlan")
    
    if not changes:
        changes.append("Nincs szükséges módosítás, minden szinkronban van")
    
    return {"message": "Státusz szinkronizáció befejezve", "changes": changes}


# ==================== USERS (Admin) ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return [UserResponse(
        id=u["id"],
        email=u["email"],
        name=u.get("name", ""),
        role=u["role"],
        created_at=u.get("created_at", "")
    ) for u in users]

@api_router.get("/users/stats")
async def get_user_stats(user: dict = Depends(require_admin)):
    """Toborzónként hány dolgozót vitt fel"""
    pipeline = [
        {"$group": {"_id": "$owner_id", "count": {"$sum": 1}}},
    ]
    stats = await db.workers.aggregate(pipeline).to_list(100)
    
    result = []
    for s in stats:
        owner = await db.users.find_one({"id": s["_id"]}, {"_id": 0, "password": 0})
        if owner:
            result.append({
                "user_id": s["_id"],
                "user_name": owner.get("name", owner["email"]),
                "user_email": owner["email"],
                "worker_count": s["count"]
            })
    return result

# ==================== WORKERS ====================

@api_router.get("/workers", response_model=List[WorkerResponse])
async def get_workers(
    search: Optional[str] = None,
    worker_type_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    owner_id: Optional[str] = None,
    county: Optional[str] = None,
    position_filter: Optional[str] = None,
    center_lat: Optional[float] = None,
    center_lon: Optional[float] = None,
    radius_km: Optional[float] = None,
    work_type: Optional[str] = None,  # Munkavégzés típusa szűrő
    has_car: Optional[str] = None,  # Saját autó szűrő
    property_filter: Optional[str] = None,  # Tulajdonság szűrő (megbizhato, jo_munkaero, stb.)
    date_from: Optional[str] = None,  # Dátum szűrés (felvétel dátuma)
    date_to: Optional[str] = None,  # Dátum szűrés vége
    global_status: Optional[str] = None,  # Globális státusz szűrő
    project_id: Optional[str] = None,  # Projekt szűrő
    project_status: Optional[str] = None,  # Projekt státusz szűrő
    gender: Optional[str] = None,  # ÚJ: Nem szűrő ("férfi", "nő")
    page: int = 1,  # Pagination
    page_size: int = 100,  # Pagination méret
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # Toborzó csak saját dolgozóit látja
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    elif owner_id:
        query["owner_id"] = owner_id
    
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    if tag_id:
        query["tag_ids"] = tag_id
    if county:
        query["county"] = {"$regex": county, "$options": "i"}
    if position_filter:
        query["position"] = {"$regex": position_filter, "$options": "i"}
    
    # Tulajdonság szűrő
    if property_filter:
        query["properties"] = property_filter
    
    # Egyéb szűrők
    if work_type:
        query["work_type"] = work_type
    if has_car:
        query["has_car"] = has_car
    if global_status:
        query["global_status"] = global_status
    if gender:  # ÚJ: Nem szűrő
        query["gender"] = gender
    
    # Dátum szűrés (felvétel dátuma)
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        if date_query:
            query["created_at"] = date_query
    
    if search:
        # ÚJ: Rugalmas, ékezet-mentes keresés
        # Eltávolítjuk az ékezeteket a keresési kifejezésből
        search_normalized = normalize_text_for_search(search)
        search_regex = create_flexible_search_regex(search)
        
        query["$or"] = [
            {"name": {"$regex": search_regex, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search_regex, "$options": "i"}},
            {"experience": {"$regex": search_regex, "$options": "i"}},
            {"position": {"$regex": search_regex, "$options": "i"}},
            {"notes": {"$regex": search_regex, "$options": "i"}}  # ÚJ: Jegyzetek keresés (skills ide kerülnek!)
        ]
    
    # Projekt státusz szűrő - előre lekérjük a megfelelő worker_id-kat
    project_worker_ids = None
    if project_id or project_status:
        pw_query = {}
        if project_id:
            pw_query["project_id"] = project_id
        
        if project_status:
            # Keressük meg a státusz ID-t név alapján
            status_doc = await db.statuses.find_one({"name": project_status}, {"_id": 0})
            if status_doc:
                pw_query["status_id"] = status_doc["id"]
        
        # Lekérjük a megfelelő project_workers rekordokat
        pw_list = await db.project_workers.find(pw_query, {"_id": 0}).to_list(10000)
        project_worker_ids = [pw["worker_id"] for pw in pw_list]
        
        # Ha nincs találat, üres listát adunk vissza
        if not project_worker_ids:
            return []
        
        # Hozzáadjuk a query-hez
        query["id"] = {"$in": project_worker_ids}
    
    # Pagination
    skip = (page - 1) * page_size
    workers = await db.workers.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Filter by distance if center coordinates and radius are provided
    if center_lat is not None and center_lon is not None and radius_km is not None:
        filtered_workers = []
        for w in workers:
            w_lat = w.get("latitude")
            w_lon = w.get("longitude")
            if w_lat and w_lon:
                distance = haversine_distance(center_lat, center_lon, w_lat, w_lon)
                if distance <= radius_km:
                    w["distance_km"] = round(distance, 1)
                    filtered_workers.append(w)
            else:
                # Ha nincs koordináta, de van cím, akkor is megmutatjuk de jelezzük
                w["distance_km"] = None
        workers = filtered_workers
        # Sorrendezés távolság szerint
        workers.sort(key=lambda x: x.get("distance_km") or 99999)
    
    # Enrich with type names, tags, project statuses
    result = []
    
    # ÚJ: Blacklist lekérése a user számára (ha nem admin)
    user_blacklist_worker_ids = []
    if user["role"] != "admin":
        blacklist_entries = await db.user_blacklist.find(
            {"user_id": user["id"]},
            {"_id": 0, "worker_id": 1}
        ).to_list(1000)
        user_blacklist_worker_ids = [entry["worker_id"] for entry in blacklist_entries]
    
    for w in workers:
        # ÚJ: Kiszűrjük a blacklist-elt dolgozókat (csak toborzóknál, admin mindent lát!)
        if user["role"] != "admin" and w["id"] in user_blacklist_worker_ids:
            continue
        
        # Get type name
        type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
        w["worker_type_name"] = type_doc["name"] if type_doc else ""
        
        # Position is now free text
        w["position"] = w.get("position", "")
        w["position_experience"] = w.get("position_experience", "")
        
        # Global status
        w["global_status"] = w.get("global_status", "Feldolgozatlan")
        
        # Geocoding fields
        w["latitude"] = w.get("latitude")
        w["longitude"] = w.get("longitude")
        w["county"] = w.get("county", "")
        
        # Új mezők
        w["work_type"] = w.get("work_type", "")
        w["has_car"] = w.get("has_car", "")
        w["gender"] = w.get("gender")  # ÚJ: Nem
        
        # Get tags
        tag_ids = w.get("tag_ids", [])
        tags = []
        for tid in tag_ids:
            tag = await db.tags.find_one({"id": tid}, {"_id": 0})
            if tag:
                tags.append(tag)
        w["tags"] = tags
        
        # Get project statuses
        project_workers = await db.project_workers.find(
            {"worker_id": w["id"]}, {"_id": 0}
        ).sort("updated_at", -1).to_list(100)
        
        project_statuses = []
        for pw in project_workers:
            project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            if project:
                project_statuses.append({
                    "project_id": project["id"],
                    "project_name": project["name"],
                    "project_date": project.get("date", ""),
                    "status_id": pw.get("status_id", ""),
                    "status_name": status["name"] if status else "Hozzárendelve",
                    "notes": pw.get("notes", ""),
                    "updated_at": pw.get("updated_at", "")
                })
        w["project_statuses"] = project_statuses
        
        # Get owner name
        owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
        w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
        
        result.append(WorkerResponse(**w))
    
    return result

@api_router.post("/workers/bulk-geocode")
async def bulk_geocode_workers(
    background_tasks: BackgroundTasks,
    user: dict = Depends(require_admin)
):
    """Start bulk geocoding for all workers without coordinates"""
    # Count workers that need geocoding
    query = {
        "address": {"$exists": True, "$ne": ""},
        "$or": [
            {"latitude": {"$exists": False}},
            {"latitude": None}
        ]
    }
    count = await db.workers.count_documents(query)
    
    if count == 0:
        return {"message": "Minden dolgozó címe már geocodolva van", "total": 0}
    
    # Start background task
    job_id = str(uuid.uuid4())
    job_doc = {
        "id": job_id,
        "type": "bulk_geocode",
        "status": "running",
        "total": count,
        "processed": 0,
        "success": 0,
        "failed": 0,
        "started_at": datetime.now(timezone.utc).isoformat()
    }
    await db.background_jobs.insert_one(job_doc)
    
    background_tasks.add_task(process_bulk_geocoding, job_id)
    
    return {
        "message": "Geocodolás elindítva",
        "job_id": job_id,
        "total": count
    }

async def process_bulk_geocoding(job_id: str):
    """Background task to geocode all workers"""
    try:
        query = {
            "address": {"$exists": True, "$ne": ""},
            "$or": [
                {"latitude": {"$exists": False}},
                {"latitude": None}
            ]
        }
        workers = await db.workers.find(query, {"_id": 0, "id": 1, "address": 1}).to_list(10000)
        
        total = len(workers)
        processed = 0
        success = 0
        failed = 0
        
        for w in workers:
            try:
                # Rate limit: 1 request per second for Nominatim
                await asyncio.sleep(1)
                
                geo_data = await geocode_address(w["address"])
                
                if geo_data.get("latitude"):
                    await db.workers.update_one(
                        {"id": w["id"]},
                        {"$set": {
                            "latitude": geo_data["latitude"],
                            "longitude": geo_data["longitude"],
                            "county": geo_data.get("county", "")
                        }}
                    )
                    success += 1
                else:
                    failed += 1
                
                processed += 1
                
                # Update job status every 10 workers
                if processed % 10 == 0:
                    await db.background_jobs.update_one(
                        {"id": job_id},
                        {"$set": {
                            "processed": processed,
                            "success": success,
                            "failed": failed
                        }}
                    )
            except Exception as e:
                failed += 1
                processed += 1
                logging.error(f"Geocoding error for worker {w['id']}: {e}")
        
        # Final update
        await db.background_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "completed",
                "processed": processed,
                "success": success,
                "failed": failed,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    except Exception as e:
        logging.error(f"Bulk geocoding error: {e}")
        await db.background_jobs.update_one(
            {"id": job_id},
            {"$set": {"status": "failed", "error": str(e)}}
        )

@api_router.get("/workers/geocode-status/{job_id}")
async def get_geocode_job_status(job_id: str, user: dict = Depends(get_current_user)):
    """Get the status of a bulk geocoding job"""
    job = await db.background_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job nem található")
    return job

@api_router.get("/workers/geocode-stats")
async def get_geocode_stats(user: dict = Depends(get_current_user)):
    """Get statistics about geocoded workers"""
    total = await db.workers.count_documents({})
    geocoded = await db.workers.count_documents({
        "latitude": {"$exists": True, "$ne": None}
    })
    not_geocoded = await db.workers.count_documents({
        "address": {"$exists": True, "$ne": ""},
        "$or": [
            {"latitude": {"$exists": False}},
            {"latitude": None}
        ]
    })
    no_address = await db.workers.count_documents({
        "$or": [
            {"address": {"$exists": False}},
            {"address": ""}
        ]
    })
    
    return {
        "total": total,
        "geocoded": geocoded,
        "not_geocoded": not_geocoded,
        "no_address": no_address
    }

@api_router.get("/workers/no-address")
async def get_workers_without_address(user: dict = Depends(get_current_user)):
    """Lakóhely nélküli dolgozók listája - manuális megadáshoz"""
    workers = await db.workers.find(
        {"$or": [
            {"address": {"$exists": False}},
            {"address": ""},
            {"address": None}
        ]},
        {"_id": 0, "id": 1, "name": 1, "phone": 1}
    ).to_list(1000)
    return {"workers": workers, "count": len(workers)}

@api_router.get("/workers/{worker_id}", response_model=WorkerResponse)
async def get_worker(worker_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    w = await db.workers.find_one(query, {"_id": 0})
    if not w:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Enrich
    type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
    w["worker_type_name"] = type_doc["name"] if type_doc else ""
    
    # Position is now free text
    w["position"] = w.get("position", "")
    w["position_experience"] = w.get("position_experience", "")
    
    # Global status
    w["global_status"] = w.get("global_status", "Feldolgozatlan")
    
    # Új mezők
    w["work_type"] = w.get("work_type", "")
    w["has_car"] = w.get("has_car", "")
    w["gender"] = w.get("gender")  # ÚJ: Nem
    w["distance_km"] = None
    
    tag_ids = w.get("tag_ids", [])
    tags = []
    for tid in tag_ids:
        tag = await db.tags.find_one({"id": tid}, {"_id": 0})
        if tag:
            tags.append(tag)
    w["tags"] = tags
    
    project_workers = await db.project_workers.find(
        {"worker_id": w["id"]}, {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    
    project_statuses = []
    for pw in project_workers:
        project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
        status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
        if project:
            project_statuses.append({
                "project_id": project["id"],
                "project_name": project["name"],
                "project_date": project.get("date", ""),
                "status_id": pw.get("status_id", ""),
                "status_name": status["name"] if status else "Hozzárendelve",
                "notes": pw.get("notes", ""),
                "updated_at": pw.get("updated_at", "")
            })
    w["project_statuses"] = project_statuses
    
    owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
    w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
    
    return WorkerResponse(**w)

@api_router.post("/workers", response_model=WorkerResponse)
async def create_worker(data: WorkerCreate, user: dict = Depends(get_current_user)):
    if len(data.name) < 2:
        raise HTTPException(status_code=400, detail="A név minimum 2 karakter legyen")
    
    # ÚJ: Automatikus nem detektálás névből (ha nincs megadva)
    detected_gender = data.gender or detect_gender_from_name(data.name)
    
    # Geocode the address if provided
    geo_data = {"latitude": None, "longitude": None, "county": ""}
    if data.address and len(data.address.strip()) > 3:
        geo_data = await geocode_address(data.address)
    
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "worker_type_id": data.worker_type_id,
        "position": data.position or "",
        "position_experience": data.position_experience or "",
        "address": data.address or "",
        "email": data.email or "",
        "experience": data.experience or "",
        "notes": data.notes or "",
        "global_status": data.global_status or "Feldolgozatlan",
        "properties": data.properties or [],
        "tag_ids": [],
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "latitude": data.latitude or geo_data.get("latitude"),
        "longitude": data.longitude or geo_data.get("longitude"),
        "county": data.county or geo_data.get("county", ""),
        "work_type": data.work_type or "",  # Új: Munkavégzés típusa
        "has_car": data.has_car or "",  # Új: Saját autó
        "gender": detected_gender  # ÚJ: Automatikusan detektált nem
    }
    await db.workers.insert_one(worker_doc)
    
    # Ha project_id meg van adva, közvetlenül a projekthez adjuk vagy várólistára
    if data.project_id:
        project = await db.projects.find_one({"id": data.project_id}, {"_id": 0})
        if project:
            # Ellenőrizzük, hogy a user hozzáfér-e a projekthez
            if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
                raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
            
            # Ha van initial_status, közvetlenül a projekthez adjuk az adott státusszal
            if data.initial_status:
                # Dolgozó közvetlenül a projekthez
                project_worker_doc = {
                    "id": str(uuid.uuid4()),
                    "project_id": data.project_id,
                    "worker_id": worker_doc["id"],
                    "status_id": data.initial_status,
                    "notes": "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.project_workers.insert_one(project_worker_doc)
            else:
                # Várólistára adjuk
                waitlist_doc = {
                    "id": str(uuid.uuid4()),
                    "project_id": data.project_id,
                    "worker_id": worker_doc["id"],
                    "trial_date": data.trial_date or "",
                    "notes": "",
                    "added_at": datetime.now(timezone.utc).isoformat(),
                    "added_by": user["id"]
                }
                await db.project_waitlist.insert_one(waitlist_doc)
    
    worker_doc["worker_type_name"] = ""
    worker_doc["tags"] = []
    worker_doc["project_statuses"] = []
    worker_doc["owner_name"] = user.get("name", user["email"])
    worker_doc["distance_km"] = None
    
    return WorkerResponse(**worker_doc)

@api_router.put("/workers/{worker_id}", response_model=WorkerResponse)
async def update_worker(worker_id: str, data: WorkerUpdate, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # ÚJ: Ha a név változik és a nem nincs explicit megadva, detektáljuk újra
    if "name" in update_data and "gender" not in update_data:
        detected_gender = detect_gender_from_name(update_data["name"])
        if detected_gender:
            update_data["gender"] = detected_gender
    
    # If address changed, re-geocode
    if "address" in update_data and update_data["address"] != worker.get("address"):
        geo_data = await geocode_address(update_data["address"])
        if geo_data.get("latitude"):
            update_data["latitude"] = geo_data["latitude"]
            update_data["longitude"] = geo_data["longitude"]
            update_data["county"] = geo_data.get("county", "")
    
    if update_data:
        await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    
    return await get_worker(worker_id, user)

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet"""
    result = await db.workers.delete_one({"id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Töröljük a projekt kapcsolatokat is
    await db.project_workers.delete_many({"worker_id": worker_id})
    
    return {"message": "Dolgozó törölve"}

@api_router.post("/workers/{worker_id}/tags/{tag_id}")
async def add_tag_to_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$addToSet": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző hozzáadva"}

@api_router.delete("/workers/{worker_id}/tags/{tag_id}")
async def remove_tag_from_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$pull": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző eltávolítva"}

# ==================== GEOCODING ENDPOINTS ====================

@api_router.get("/counties")
async def get_counties():
    """Get list of Hungarian counties"""
    return HUNGARIAN_COUNTIES

@api_router.post("/geocode")
async def geocode_single_address(data: dict, user: dict = Depends(get_current_user)):
    """Geocode a single address"""
    address = data.get("address", "")
    if not address:
        raise HTTPException(status_code=400, detail="Cím megadása kötelező")
    
    result = await geocode_address(address)
    return result

# ==================== PROJECTS ====================

@api_router.get("/projects", response_model=List[ProjectResponse])
async def get_projects(user: dict = Depends(get_current_user)):
    """Toborzó csak azokat a projekteket látja, ahol ő hozta létre VAGY hozzá van rendelve"""
    projects = await db.projects.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Lekérjük a "Dolgozik" státuszt az aktív dolgozók számlálásához
    dolgozik_status = await db.statuses.find_one({"name": "Dolgozik"}, {"_id": 0})
    
    result = []
    for p in projects:
        count = await db.project_workers.count_documents({"project_id": p["id"]})
        position_count = await db.project_positions.count_documents({"project_id": p["id"]})
        trial_count = await db.trials.count_documents({"project_id": p["id"]})
        
        # Calculate total headcount from positions
        positions = await db.project_positions.find({"project_id": p["id"]}, {"_id": 0}).to_list(100)
        total_headcount = sum(pos.get("headcount", 0) for pos in positions)
        
        # Aktív dolgozók számlálása (Dolgozik státusz)
        active_worker_count = 0
        if dolgozik_status:
            active_worker_count = await db.project_workers.count_documents({
                "project_id": p["id"],
                "status_id": dolgozik_status["id"]
            })
        
        recruiter_ids = p.get("recruiter_ids", [])
        owner_id = p.get("owner_id", "")
        
        # Ha toborzó, csak azokat mutassa ahol ő hozta létre VAGY hozzá van rendelve
        if user["role"] != "admin":
            if owner_id != user["id"] and user["id"] not in recruiter_ids:
                continue
        
        # Get recruiter names
        recruiters = []
        for rid in recruiter_ids:
            r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
            if r:
                recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
        
        # Get owner name
        owner_name = ""
        if owner_id:
            owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
            if owner:
                owner_name = owner.get("name", owner["email"])
        
        result.append(ProjectResponse(
            id=p["id"],
            name=p["name"],
            client_name=p.get("client_name", ""),
            date=p["date"],
            location=p.get("location", ""),
            training_location=p.get("training_location", ""),
            notes=p.get("notes", ""),
            is_closed=p.get("is_closed", False),
            worker_count=count,
            position_count=position_count,
            total_headcount=total_headcount,
            trial_count=trial_count,
            recruiter_ids=recruiter_ids,
            recruiters=recruiters,
            owner_id=owner_id,
            owner_name=owner_name,
            created_at=p.get("created_at", ""),
            planned_headcount=p.get("planned_headcount", 0),
            active_worker_count=active_worker_count
        ))
    
    return result

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, user: dict = Depends(get_current_user)):
    p = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Ellenőrizzük jogosultságot
    recruiter_ids = p.get("recruiter_ids", [])
    owner_id = p.get("owner_id", "")
    if user["role"] != "admin" and owner_id != user["id"] and user["id"] not in recruiter_ids:
        raise HTTPException(status_code=403, detail="Nincs hozzáférésed ehhez a projekthez")
    
    # Lekérjük a "Dolgozik" státuszt az aktív dolgozók számlálásához
    dolgozik_status = await db.statuses.find_one({"name": "Dolgozik"}, {"_id": 0})
    
    # Get workers
    pw_list = await db.project_workers.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    workers = []
    active_worker_count = 0
    for pw in pw_list:
        w = await db.workers.find_one({"id": pw["worker_id"]}, {"_id": 0})
        if w:
            # Toborzó csak saját dolgozóit látja a projektben
            if user["role"] != "admin" and w.get("owner_id") != user["id"]:
                continue
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
            owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
            
            status_name = status["name"] if status else "Hozzárendelve"
            
            # Számoljuk az aktív dolgozókat
            if dolgozik_status and pw.get("status_id") == dolgozik_status["id"]:
                active_worker_count += 1
            
            # Get trial info if assigned
            trial_id = pw.get("trial_id", "")
            trial_date = ""
            trial_time = ""
            if trial_id:
                trial_doc = await db.trials.find_one({"id": trial_id}, {"_id": 0})
                if trial_doc:
                    trial_date = trial_doc.get("date", "")
                    trial_time = trial_doc.get("time", "")
            
            # Get position names from position_ids
            position_names = []
            position_ids = pw.get("position_ids", [])
            for pos_id in position_ids:
                pos_doc = await db.project_positions.find_one({"id": pos_id}, {"_id": 0})
                if pos_doc:
                    position_names.append(pos_doc.get("name", ""))
            
            workers.append({
                "id": w["id"],
                "name": w["name"],
                "phone": w["phone"],
                "global_status": w.get("global_status", "Feldolgozatlan"),
                "worker_type_name": type_doc["name"] if type_doc else "",
                "status_id": pw.get("status_id", ""),
                "status_name": status_name,
                "notes": pw.get("notes", ""),
                "added_by": owner.get("name", owner["email"]) if owner else "",
                "added_at": pw.get("created_at", ""),
                "created_at": pw.get("created_at", ""),
                "trial_id": trial_id,
                "trial_date": trial_date,
                "trial_time": trial_time,
                "position_ids": position_ids,
                "position_names": position_names
            })
    
    # Total count - toborzó csak saját dolgozóit számolja
    if user["role"] == "admin":
        total_count = await db.project_workers.count_documents({"project_id": project_id})
    else:
        # Toborzó: csak saját dolgozók számolása
        total_count = len(workers)
    
    # Get positions
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    total_headcount = sum(pos.get("headcount", 0) for pos in positions)
    
    # Get trials with positions
    trials = await db.trials.find({"project_id": project_id}, {"_id": 0}).sort("date", 1).to_list(100)
    for trial in trials:
        trial_workers = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
        trial["worker_count"] = len(trial_workers)
        
        # Get trial positions
        trial_positions = await db.trial_positions.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(50)
        trial_pos_response = []
        for tp in trial_positions:
            assigned = await db.trial_workers.count_documents({
                "trial_id": trial["id"],
                "trial_position_id": tp["id"]
            })
            trial_pos_response.append({
                **tp,
                "assigned_count": assigned
            })
        trial["positions"] = trial_pos_response
        
        # Get worker details
        workers_list = []
        for tw in trial_workers:
            w = await db.workers.find_one({"id": tw["worker_id"]}, {"_id": 0})
            if w:
                trial_pos = await db.trial_positions.find_one({"id": tw.get("trial_position_id")}, {"_id": 0}) if tw.get("trial_position_id") else None
                # Ha van trial pozíció, azt használjuk, különben a dolgozó eredeti pozícióját
                position_name = trial_pos["position_name"] if trial_pos else w.get("position", "")
                workers_list.append({
                    "id": w["id"],
                    "name": w["name"],
                    "phone": w.get("phone", ""),
                    "email": w.get("email", ""),
                    "position": w.get("position", ""),  # Dolgozó eredeti pozíciója
                    "trial_position_id": tw.get("trial_position_id", ""),
                    "position_name": position_name,
                    "added_at": tw.get("added_at", "")
                })
        trial["workers"] = workers_list
    
    # Get recruiter names
    recruiters = []
    for rid in recruiter_ids:
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return {
        "id": p["id"],
        "name": p["name"],
        "client_name": p.get("client_name", ""),
        "date": p["date"],
        "location": p.get("location", ""),
        "training_location": p.get("training_location", ""),
        "notes": p.get("notes", ""),
        "is_closed": p.get("is_closed", False),
        "total_count": total_count,  # Toborzó csak saját dolgozóit látja
        "worker_count": total_count,  # Backward compatibility
        "total_headcount": total_headcount,
        "recruiter_ids": recruiter_ids,
        "recruiters": recruiters,
        "owner_id": owner_id,
        "owner_name": owner_name,
        "workers": workers,
        "positions": positions,
        "trials": trials,
        "created_at": p.get("created_at", ""),
        "planned_headcount": p.get("planned_headcount", 0),
        "active_worker_count": active_worker_count
    }

@api_router.post("/projects", response_model=ProjectResponse)
async def create_project(data: ProjectCreate, user: dict = Depends(require_admin)):
    """Csak admin hozhat létre projektet"""
    project_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "client_name": data.client_name or "",
        "date": data.date,
        "location": data.location or "",
        "training_location": data.training_location or "",
        "notes": data.notes or "",
        "recruiter_ids": data.recruiter_ids,
        "is_closed": False,
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "planned_headcount": data.planned_headcount or 0
    }
    await db.projects.insert_one(project_doc)
    
    owner_name = user.get("name", user["email"])
    return ProjectResponse(**project_doc, worker_count=0, position_count=0, total_headcount=0, trial_count=0, recruiters=[], owner_name=owner_name, active_worker_count=0)

@api_router.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, data: ProjectUpdate, user: dict = Depends(require_admin)):
    """Csak admin módosíthatja a projektet"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.projects.update_one({"id": project_id}, {"$set": update_data})
    
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    count = await db.project_workers.count_documents({"project_id": project_id})
    position_count = await db.project_positions.count_documents({"project_id": project_id})
    trial_count = await db.trials.count_documents({"project_id": project_id})
    
    # Calculate total headcount
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    total_headcount = sum(pos.get("headcount", 0) for pos in positions)
    
    # Aktív dolgozók számlálása
    dolgozik_status = await db.statuses.find_one({"name": "Dolgozik"}, {"_id": 0})
    active_worker_count = 0
    if dolgozik_status:
        active_worker_count = await db.project_workers.count_documents({
            "project_id": project_id,
            "status_id": dolgozik_status["id"]
        })
    
    # Get recruiters
    recruiters = []
    for rid in updated.get("recruiter_ids", []):
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    owner_id = updated.get("owner_id", "")
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return ProjectResponse(**updated, worker_count=count, position_count=position_count, total_headcount=total_headcount, trial_count=trial_count, recruiters=recruiters, owner_name=owner_name, active_worker_count=active_worker_count)

@api_router.post("/projects/{project_id}/recruiters")
async def add_recruiter_to_project(project_id: str, data: ProjectRecruiterAdd, user: dict = Depends(require_admin)):
    """Admin hozzárendel egy toborzót a projekthez"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    target_user = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    # Check if already assigned
    if data.user_id in project.get("recruiter_ids", []):
        return {"message": "Toborzó már hozzá van rendelve"}
    
    await db.projects.update_one(
        {"id": project_id},
        {"$addToSet": {"recruiter_ids": data.user_id}}
    )
    
    # Create notification for the recruiter
    await create_notification(
        user_id=data.user_id,
        notification_type="project_assigned",
        title="Új projekt hozzárendelés",
        message=f"Hozzárendeltek a(z) '{project['name']}' projekthez",
        link=f"/projects/{project_id}"
    )
    
    return {"message": "Toborzó hozzárendelve a projekthez"}

@api_router.delete("/projects/{project_id}/recruiters/{user_id}")
async def remove_recruiter_from_project(project_id: str, user_id: str, user: dict = Depends(require_admin)):
    """Admin eltávolít egy toborzót a projektből"""
    result = await db.projects.update_one(
        {"id": project_id},
        {"$pull": {"recruiter_ids": user_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    return {"message": "Toborzó eltávolítva a projektről"}

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet projektet"""
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    await db.project_workers.delete_many({"project_id": project_id})
    await db.project_positions.delete_many({"project_id": project_id})
    await db.trials.delete_many({"project_id": project_id})
    return {"message": "Projekt törölve"}

# ==================== PROJECT POSITIONS ====================

@api_router.get("/projects/{project_id}/positions", response_model=List[ProjectPositionResponse])
async def get_project_positions(project_id: str, user: dict = Depends(get_current_user)):
    """Projekt pozícióinak lekérése"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    
    result = []
    for pos in positions:
        # Count assigned workers for this position
        assigned_workers = await db.project_workers.count_documents({
            "project_id": project_id,
            "position_id": pos["id"]
        })
        # Backward compatibility: map shift_schedule to work_schedule
        result.append(ProjectPositionResponse(
            id=pos["id"],
            project_id=pos["project_id"],
            name=pos["name"],
            headcount=pos["headcount"],
            work_schedule=pos.get("work_schedule", pos.get("shift_schedule", "")),  # Backward compatible
            experience_required=pos.get("experience_required", ""),
            qualifications=pos.get("qualifications", ""),
            physical_requirements=pos.get("physical_requirements", ""),
            position_details=pos.get("position_details", ""),
            notes=pos.get("notes", ""),
            created_at=pos.get("created_at", ""),
            assigned_workers=assigned_workers
        ))
    
    return result

@api_router.post("/projects/{project_id}/positions", response_model=ProjectPositionResponse)
async def create_project_position(project_id: str, data: ProjectPositionCreate, user: dict = Depends(require_admin)):
    """Pozíció létrehozása projekthez - csak admin"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    position_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "name": sanitize_string(data.name, max_length=100),
        "headcount": data.headcount,
        "work_schedule": data.work_schedule or "",
        "salary": data.salary or "",  # Bérezés mező
        "experience_required": data.experience_required or "",
        "qualifications": data.qualifications or "",
        "physical_requirements": data.physical_requirements or "",
        "position_details": data.position_details or "",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_positions.insert_one(position_doc)
    
    return ProjectPositionResponse(**position_doc, assigned_workers=0)

@api_router.put("/projects/{project_id}/positions/{position_id}", response_model=ProjectPositionResponse)
async def update_project_position(project_id: str, position_id: str, data: ProjectPositionUpdate, user: dict = Depends(require_admin)):
    """Pozíció szerkesztése - csak admin"""
    position = await db.project_positions.find_one({"id": position_id, "project_id": project_id})
    if not position:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.project_positions.update_one({"id": position_id}, {"$set": update_data})
    
    updated = await db.project_positions.find_one({"id": position_id}, {"_id": 0})
    assigned_workers = await db.project_workers.count_documents({
        "project_id": project_id,
        "position_id": position_id
    })
    
    return ProjectPositionResponse(**updated, assigned_workers=assigned_workers)

@api_router.delete("/projects/{project_id}/positions/{position_id}")
async def delete_project_position(project_id: str, position_id: str, user: dict = Depends(require_admin)):
    """Pozíció törlése - csak admin"""
    result = await db.project_positions.delete_one({"id": position_id, "project_id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    # Remove position_id from project_workers
    await db.project_workers.update_many(
        {"project_id": project_id, "position_id": position_id},
        {"$unset": {"position_id": ""}}
    )
    
    return {"message": "Pozíció törölve"}

# ==================== WAITLIST (VÁRÓLISTA) ====================

@api_router.get("/projects/{project_id}/waitlist", response_model=List[WaitlistWorkerResponse])
async def get_project_waitlist(project_id: str, user: dict = Depends(get_current_user)):
    """Get all workers in project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Check permission
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
    
    waitlist_entries = await db.project_waitlist.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    result = []
    for entry in waitlist_entries:
        worker = await db.workers.find_one({"id": entry["worker_id"]}, {"_id": 0})
        if worker:
            added_by_user = await db.users.find_one({"id": entry["added_by"]}, {"_id": 0})
            result.append(WaitlistWorkerResponse(
                id=entry["id"],
                project_id=entry["project_id"],
                worker_id=entry["worker_id"],
                worker_name=worker["name"],
                worker_phone=worker["phone"],
                worker_email=worker.get("email", ""),
                trial_date=entry.get("trial_date", entry.get("start_date", "")),  # Backward compatible
                notes=entry.get("notes", ""),
                added_at=entry["added_at"],
                added_by=entry["added_by"],
                added_by_name=added_by_user.get("name", added_by_user["email"]) if added_by_user else ""
            ))
    
    return result

@api_router.post("/projects/{project_id}/waitlist")
async def add_worker_to_waitlist(project_id: str, data: WaitlistWorkerAdd, user: dict = Depends(get_current_user)):
    """Add a worker to project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Check permission
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
    
    # Check if worker exists
    worker = await db.workers.find_one({"id": data.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Check if worker already in waitlist
    existing = await db.project_waitlist.find_one({"project_id": project_id, "worker_id": data.worker_id})
    if existing:
        raise HTTPException(status_code=400, detail="A dolgozó már a várólistán van")
    
    waitlist_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "worker_id": data.worker_id,
        "trial_date": data.trial_date or "",  # ÚJ: trial_date
        "notes": data.notes or "",
        "added_at": datetime.now(timezone.utc).isoformat(),
        "added_by": user["id"]
    }
    await db.project_waitlist.insert_one(waitlist_doc)
    
    return {"message": "Dolgozó hozzáadva a várólistához", "id": waitlist_doc["id"]}

@api_router.put("/projects/{project_id}/waitlist/{worker_id}")
async def update_waitlist_entry(project_id: str, worker_id: str, data: WaitlistWorkerUpdate, user: dict = Depends(get_current_user)):
    """Update waitlist entry"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod")
    
    entry = await db.project_waitlist.find_one({"project_id": project_id, "worker_id": worker_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Dolgozó nincs a várólistán")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.project_waitlist.update_one(
            {"project_id": project_id, "worker_id": worker_id},
            {"$set": update_data}
        )
    
    return {"message": "Várólista frissítve"}

@api_router.delete("/projects/{project_id}/waitlist/{worker_id}")
async def remove_worker_from_waitlist(project_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    """Remove worker from project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod")
    
    result = await db.project_waitlist.delete_one({"project_id": project_id, "worker_id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nincs a várólistán")
    
    return {"message": "Dolgozó eltávolítva a várólistáról"}

# ==================== TRIALS (PRÓBÁK) ====================

@api_router.get("/projects/{project_id}/trials", response_model=List[TrialResponse])
async def get_project_trials(project_id: str, user: dict = Depends(get_current_user)):
    """Projekt próbáinak lekérése"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    trials = await db.trials.find({"project_id": project_id}, {"_id": 0}).sort("date", 1).to_list(100)
    
    result = []
    for trial in trials:
        # Get workers for this trial
        trial_workers_list = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
        
        workers = []
        for tw in trial_workers_list:
            w = await db.workers.find_one({"id": tw["worker_id"]}, {"_id": 0})
            if w:
                # Toborzó csak saját dolgozóit látja
                if user["role"] != "admin" and w.get("owner_id") != user["id"]:
                    continue
                # Get trial position name if assigned
                trial_pos = await db.trial_positions.find_one({"id": tw.get("trial_position_id")}, {"_id": 0}) if tw.get("trial_position_id") else None
                # Ha van trial pozíció, azt használjuk, különben a dolgozó eredeti pozícióját
                position_name = trial_pos["position_name"] if trial_pos else w.get("position", "")
                workers.append({
                    "id": w["id"],
                    "name": w["name"],
                    "phone": w["phone"],
                    "email": w.get("email", ""),
                    "position": w.get("position", ""),  # Dolgozó eredeti pozíciója
                    "trial_position_id": tw.get("trial_position_id", ""),
                    "position_name": position_name,
                    "added_at": tw.get("created_at", "")
                })
        
        # Get trial positions
        trial_positions = await db.trial_positions.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(50)
        positions_response = []
        for tp in trial_positions:
            # Count assigned workers for this position
            assigned = await db.trial_workers.count_documents({
                "trial_id": trial["id"],
                "trial_position_id": tp["id"]
            })
            positions_response.append(TrialPositionResponse(
                **tp,
                assigned_count=assigned
            ))
        
        result.append(TrialResponse(
            **trial,
            worker_count=len(trial_workers_list),
            workers=workers,
            positions=positions_response
        ))
    
    return result

@api_router.post("/projects/{project_id}/trials", response_model=TrialResponse)
async def create_trial(project_id: str, data: TrialCreate, user: dict = Depends(require_admin)):
    """Próba létrehozása - csak admin"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    trial_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "date": data.date,
        "time": data.time or "",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trials.insert_one(trial_doc)
    
    # Értesítés küldése a projekt toborzóinak
    recruiter_ids = project.get("recruiter_ids", [])
    if project.get("owner_id"):
        recruiter_ids.append(project["owner_id"])
    
    for recruiter_id in set(recruiter_ids):  # Unique IDs
        await create_notification(
            user_id=recruiter_id,
            notification_type="trial_created",
            title="Új próba hozzáadva",
            message=f"Új próba lett hozzáadva a(z) '{project['name']}' projekthez. Dátum: {data.date}" + (f" {data.time}" if data.time else ""),
            link=f"/projects/{project_id}"
        )
    
    return TrialResponse(**trial_doc, worker_count=0, workers=[], positions=[])

@api_router.put("/projects/{project_id}/trials/{trial_id}", response_model=TrialResponse)
async def update_trial(project_id: str, trial_id: str, data: TrialUpdate, user: dict = Depends(require_admin)):
    """Próba szerkesztése - csak admin"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.trials.update_one({"id": trial_id}, {"$set": update_data})
    
    updated = await db.trials.find_one({"id": trial_id}, {"_id": 0})
    worker_count = await db.trial_workers.count_documents({"trial_id": trial_id})
    
    return TrialResponse(**updated, worker_count=worker_count, workers=[])

@api_router.delete("/projects/{project_id}/trials/{trial_id}")
async def delete_trial(project_id: str, trial_id: str, user: dict = Depends(require_admin)):
    """Próba törlése - csak admin"""
    result = await db.trials.delete_one({"id": trial_id, "project_id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    # Delete trial workers and positions
    await db.trial_workers.delete_many({"trial_id": trial_id})
    await db.trial_positions.delete_many({"trial_id": trial_id})
    
    return {"message": "Próba törölve"}

# ==================== TRIAL POSITIONS ====================

@api_router.get("/projects/{project_id}/trials/{trial_id}/positions")
async def get_trial_positions(project_id: str, trial_id: str, user: dict = Depends(get_current_user)):
    """Próba pozícióinak lekérése"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    positions = await db.trial_positions.find({"trial_id": trial_id}, {"_id": 0}).to_list(50)
    result = []
    for tp in positions:
        assigned = await db.trial_workers.count_documents({
            "trial_id": trial_id,
            "trial_position_id": tp["id"]
        })
        result.append(TrialPositionResponse(**tp, assigned_count=assigned))
    return result

@api_router.post("/projects/{project_id}/trials/{trial_id}/positions", response_model=TrialPositionResponse)
async def add_trial_position(project_id: str, trial_id: str, data: TrialPositionCreate, user: dict = Depends(require_admin)):
    """Pozíció hozzáadása próbához - csak admin"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    # Check if position name already exists in trial
    existing = await db.trial_positions.find_one({
        "trial_id": trial_id,
        "position_name": data.position_name
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ez a pozíció már hozzá van adva ehhez a próbához")
    
    # If add_to_project is True, add to project positions too
    if data.add_to_project:
        existing_project_pos = await db.project_positions.find_one({
            "project_id": project_id,
            "name": data.position_name
        })
        if not existing_project_pos:
            project_pos_doc = {
                "id": str(uuid.uuid4()),
                "project_id": project_id,
                "name": data.position_name,
                "headcount": data.headcount,
                "shift_schedule": "",
                "experience_required": data.requirements,
                "qualifications": "",
                "physical_requirements": "",
                "notes": "",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.project_positions.insert_one(project_pos_doc)
            data.position_id = project_pos_doc["id"]
    
    trial_pos_doc = {
        "id": str(uuid.uuid4()),
        "trial_id": trial_id,
        "position_id": data.position_id or "",
        "position_name": data.position_name,
        "headcount": data.headcount,
        "hourly_rate": data.hourly_rate or "",
        "accommodation": data.accommodation,
        "requirements": data.requirements,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trial_positions.insert_one(trial_pos_doc)
    
    return TrialPositionResponse(**trial_pos_doc, assigned_count=0)

@api_router.put("/projects/{project_id}/trials/{trial_id}/positions/{position_id}")
async def update_trial_position(project_id: str, trial_id: str, position_id: str, data: TrialPositionUpdate, user: dict = Depends(require_admin)):
    """Próba pozíció szerkesztése - csak admin"""
    trial_pos = await db.trial_positions.find_one({"id": position_id, "trial_id": trial_id})
    if not trial_pos:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    update_data = {}
    if data.position_name is not None:
        update_data["position_name"] = data.position_name
    if data.headcount is not None:
        update_data["headcount"] = data.headcount
    if data.hourly_rate is not None:
        update_data["hourly_rate"] = data.hourly_rate
    if data.accommodation is not None:
        update_data["accommodation"] = data.accommodation
    if data.requirements is not None:
        update_data["requirements"] = data.requirements
    
    if update_data:
        await db.trial_positions.update_one({"id": position_id}, {"$set": update_data})
    
    updated = await db.trial_positions.find_one({"id": position_id}, {"_id": 0})
    assigned = await db.trial_workers.count_documents({"trial_id": trial_id, "trial_position_id": position_id})
    return TrialPositionResponse(**updated, assigned_count=assigned)

@api_router.delete("/projects/{project_id}/trials/{trial_id}/positions/{position_id}")
async def delete_trial_position(project_id: str, trial_id: str, position_id: str, user: dict = Depends(require_admin)):
    """Próba pozíció törlése - csak admin"""
    result = await db.trial_positions.delete_one({"id": position_id, "trial_id": trial_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    # Remove position assignment from workers
    await db.trial_workers.update_many(
        {"trial_id": trial_id, "trial_position_id": position_id},
        {"$set": {"trial_position_id": ""}}
    )
    
    return {"message": "Pozíció törölve"}

# ==================== TRIAL WORKERS ====================

@api_router.post("/projects/{project_id}/trials/{trial_id}/workers")
async def add_worker_to_trial(project_id: str, trial_id: str, data: TrialWorkerAdd, user: dict = Depends(require_admin)):
    """Dolgozó hozzáadása próbához - csak admin"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id}, {"_id": 0})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    worker = await db.workers.find_one({"id": data.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    existing = await db.trial_workers.find_one({
        "trial_id": trial_id,
        "worker_id": data.worker_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Dolgozó már hozzá van rendelve ehhez a próbához")
    
    tw_doc = {
        "id": str(uuid.uuid4()),
        "trial_id": trial_id,
        "worker_id": data.worker_id,
        "trial_position_id": data.position_id or "",  # This is now trial_position_id
        "added_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trial_workers.insert_one(tw_doc)
    
    # Send notification to worker's owner (recruiter) if different from current user
    if worker.get("owner_id") and worker["owner_id"] != user["id"]:
        await create_notification(
            user_id=worker["owner_id"],
            notification_type="trial_assigned",
            title="Dolgozó próbára beosztva",
            message=f"A(z) '{worker['name']}' dolgozódat beosztották a '{project['name'] if project else 'Projekt'}' próbájára ({trial['date']})",
            link=f"/projects/{project_id}"
        )
    
    return {"message": "Dolgozó hozzáadva a próbához"}

@api_router.put("/projects/{project_id}/trials/{trial_id}/workers/{worker_id}/position")
async def assign_worker_to_trial_position(project_id: str, trial_id: str, worker_id: str, trial_position_id: str = "", user: dict = Depends(require_admin)):
    """Dolgozó pozícióhoz rendelése a próbán belül - csak admin"""
    result = await db.trial_workers.update_one(
        {"trial_id": trial_id, "worker_id": worker_id},
        {"$set": {"trial_position_id": trial_position_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nem található ezen a próbán")
    return {"message": "Pozíció frissítve"}

@api_router.delete("/projects/{project_id}/trials/{trial_id}/workers/{worker_id}")
async def remove_worker_from_trial(project_id: str, trial_id: str, worker_id: str, user: dict = Depends(require_admin)):
    """Dolgozó eltávolítása próbáról - csak admin"""
    result = await db.trial_workers.delete_one({
        "trial_id": trial_id,
        "worker_id": worker_id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Dolgozó eltávolítva a próbáról"}

@api_router.post("/projects/{project_id}/workers")
async def add_worker_to_project(project_id: str, data: ProjectWorkerAdd, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    worker = await db.workers.find_one({"id": data.worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # JOGOSULTSÁG ELLENŐRZÉS: Toborzó csak saját dolgozóját adhatja hozzá
    if user["role"] != "admin" and worker.get("owner_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Csak saját dolgozódat adhatod hozzá a projekthez")
    
    # Ellenőrizzük, hogy már van-e kapcsolat
    existing = await db.project_workers.find_one({
        "project_id": project_id,
        "worker_id": data.worker_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Dolgozó már hozzá van rendelve")
    
    # DUPLIKÁCIÓ ELLENŐRZÉS: Volt-e már kukában ebben a projektben?
    kuka_status = await db.statuses.find_one({"name": "Kuka"}, {"_id": 0})
    if kuka_status and not data.force_add:
        # Ellenőrizzük az aktív kapcsolatot ÉS a history-t is
        was_in_kuka = await db.project_workers.find_one({
            "project_id": project_id,
            "worker_id": data.worker_id,
            "status_id": kuka_status["id"]
        })
        
        # Ha nincs aktív kukás rekord, nézzük a history-t
        if not was_in_kuka:
            was_in_kuka = await db.project_worker_history.find_one({
                "project_id": project_id,
                "worker_id": data.worker_id,
                "status_id": kuka_status["id"]
            })
        
        # Ha volt kukában, warning
        if was_in_kuka:
            raise HTTPException(
                status_code=409,  # Conflict
                detail={
                    "type": "kuka_warning",
                    "message": "Ez a dolgozó már korábban ebben a projektben kukába került",
                    "reason": was_in_kuka.get("notes", ""),
                    "date": was_in_kuka.get("updated_at") or was_in_kuka.get("archived_at", "")
                }
            )
    
    # Alapértelmezett státusz projekten belül: "Próbára vár"
    default_status = await db.statuses.find_one({"name": "Próbára vár"}, {"_id": 0})
    status_id = data.status_id or (default_status["id"] if default_status else "")
    
    # VALIDÁCIÓ: Kötelező pozíció választás projekten belül (minden státusznál KIVÉVE Kuka és Tiltólista)
    status_doc = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    status_name = status_doc["name"] if status_doc else "Próbára vár"
    
    # Projekten belül kötelező pozíció választani (kivéve Kuka és Tiltólista)
    exempt_statuses = ["Kuka", "Tiltólista"]
    if status_name not in exempt_statuses and (not data.position_ids or len(data.position_ids) == 0):
        raise HTTPException(
            status_code=400, 
            detail=f"Projekten belül kötelező legalább egy pozíciót választani!"
        )
    
    pw_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "worker_id": data.worker_id,
        "status_id": status_id,
        "position_ids": data.position_ids or [],  # ÚJ: Többszörös pozíció
        "added_by": user["id"],
        "from_form": data.from_form if hasattr(data, 'from_form') else False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # ÚJ: Ha trial_id van megadva, adjuk hozzá
    trial_id = getattr(data, 'trial_id', None)
    trial_position_id = getattr(data, 'trial_position_id', None)
    
    if trial_id:
        pw_doc["trial_id"] = trial_id
        
        # Adjuk hozzá a trial_workers táblához is
        existing_trial_worker = await db.trial_workers.find_one({
            "trial_id": trial_id,
            "worker_id": data.worker_id
        })
        
        if not existing_trial_worker:
            trial_worker_doc = {
                "id": str(uuid.uuid4()),
                "trial_id": trial_id,
                "worker_id": data.worker_id,
                "trial_position_id": trial_position_id or "",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.trial_workers.insert_one(trial_worker_doc)
            
            # Update position assigned count if position was selected
            if trial_position_id:
                await db.trial_positions.update_one(
                    {"id": trial_position_id},
                    {"$inc": {"assigned_count": 1}}
                )
    
    await db.project_workers.insert_one(pw_doc)
    
    # ÉRTESÍTÉS: Ha ADMIN adta hozzá a dolgozót, értesítsük a dolgozó tulajdonosát (toborzót)
    if user["role"] == "admin" and worker.get("owner_id") != user["id"]:
        owner_id = worker.get("owner_id")
        if owner_id:
            # Get status name for notification
            status_doc = await db.statuses.find_one({"id": status_id}, {"_id": 0})
            status_name = status_doc["name"] if status_doc else "Feldolgozatlan"
            
            await create_notification(
                user_id=owner_id,
                notification_type="project_assigned",
                title="Dolgozó projekthez adva",
                message=f"{worker['name']} hozzá lett adva a(z) {project['name']} projekthez ({status_name})",
                link=f"/projects/{project_id}"
            )
    
    # Get status name for response
    status_doc = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    status_name = status_doc["name"] if status_doc else "Feldolgozatlan"
    
    return {"message": f"Dolgozó hozzáadva: {status_name}", "status": status_name}

@api_router.delete("/projects/{project_id}/workers/{worker_id}")
async def remove_worker_from_project(project_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    # Keresük meg a kapcsolatot
    pw = await db.project_workers.find_one({
        "project_id": project_id,
        "worker_id": worker_id
    }, {"_id": 0})
    
    if not pw:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    
    # Ha Kuka státuszú volt, mentsük el a history-ba
    kuka_status = await db.statuses.find_one({"name": "Kuka"}, {"_id": 0})
    if kuka_status and pw.get("status_id") == kuka_status["id"]:
        # Mentsük a kukázási történetet
        history_doc = {
            "id": str(uuid.uuid4()),
            "project_id": project_id,
            "worker_id": worker_id,
            "status_id": pw.get("status_id"),
            "notes": pw.get("notes", ""),
            "original_created_at": pw.get("created_at"),
            "original_updated_at": pw.get("updated_at"),
            "archived_at": datetime.now(timezone.utc).isoformat(),
            "archived_by": user["id"]
        }
        await db.project_worker_history.insert_one(history_doc)
    
    result = await db.project_workers.delete_one({
        "project_id": project_id,
        "worker_id": worker_id
    })
    return {"message": "Dolgozó eltávolítva a projektről"}

@api_router.put("/projects/{project_id}/workers/{worker_id}/status")
async def update_worker_status_in_project(
    project_id: str, 
    worker_id: str, 
    data: ProjectWorkerStatusUpdate,
    user: dict = Depends(get_current_user)
):
    update_fields = {
        "status_id": data.status_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if data.notes is not None:
        update_fields["notes"] = data.notes
    
    # Ha "Próba megbeszélve" státusz, és van trial_id, mentsük
    if data.trial_id is not None:
        update_fields["trial_id"] = data.trial_id
        # Ha trial_id van, akkor automatikusan hozzá is adjuk a próbához
        if data.trial_id:
            existing_trial_worker = await db.trial_workers.find_one({
                "trial_id": data.trial_id, 
                "worker_id": worker_id
            })
            if not existing_trial_worker:
                trial_worker_doc = {
                    "id": str(uuid.uuid4()),
                    "trial_id": data.trial_id,
                    "worker_id": worker_id,
                    "trial_position_id": "",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.trial_workers.insert_one(trial_worker_doc)
    
    result = await db.project_workers.update_one(
        {"project_id": project_id, "worker_id": worker_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    
    # Státusz változás naplózása a dolgozó notes mezőjébe
    status = await db.statuses.find_one({"id": data.status_id}, {"_id": 0})
    if status:
        timestamp = datetime.now(timezone.utc).strftime("%Y.%m.%d %H:%M")
        
        # Projekt neve és helyszín lekérése
        project = await db.projects.find_one({"id": project_id}, {"_id": 0, "name": 1, "location": 1})
        project_name = project["name"] if project else "Ismeretlen projekt"
        if project and project.get("location"):
            project_name += f" | {project['location']}"
        
        # Meglévő notes lekérése
        worker_doc = await db.workers.find_one({"id": worker_id}, {"_id": 0})
        existing_notes = worker_doc.get("notes", "") if worker_doc else ""
        
        # Új történet bejegyzés - formátum a specifikáció szerint
        new_entry = f"[Projekt: {project_name}]\nStátusz: {status['name']}"
        if data.notes:
            new_entry += f"\nMegjegyzés: {data.notes}"
        new_entry += f"\nDátum: {timestamp}"
        new_entry += "\n---"
        
        # Összefűzés - új bejegyzés az elejére
        updated_notes = f"{new_entry}\n{existing_notes}" if existing_notes else new_entry
        
        # GLOBÁLIS STÁTUSZ AUTOMATIKUS FRISSÍTÉS
        # Egységes státuszok: Feldolgozatlan, Próbára vár, Próba megbeszélve, Dolgozik, Tiltólista
        global_status_update = {}
        status_name = status["name"]
        
        if status_name == "Dolgozik":
            global_status_update["global_status"] = "Dolgozik"
        elif status_name == "Próba megbeszélve":
            global_status_update["global_status"] = "Próba megbeszélve"
        elif status_name == "Próbára vár":
            global_status_update["global_status"] = "Próbára vár"
        elif status_name == "Kuka":
            # Ha kukába kerül, automatikusan "Feldolgozatlan" lesz a globális státusz
            global_status_update["global_status"] = "Feldolgozatlan"
        elif status_name == "Tiltólista":
            global_status_update["global_status"] = "Tiltólista"
        
        # Mentés
        await db.workers.update_one(
            {"id": worker_id},
            {"$set": {"notes": updated_notes, **global_status_update}}
        )
    
    return {"message": "Státusz frissítve"}

# Update project worker positions
class ProjectWorkerPositionsUpdate(BaseModel):
    position_ids: List[str] = []

@api_router.put("/projects/{project_id}/workers/{worker_id}/positions")
async def update_worker_positions_in_project(
    project_id: str, 
    worker_id: str, 
    data: ProjectWorkerPositionsUpdate,
    user: dict = Depends(get_current_user)
):
    """Update worker's positions in a project"""
    # Check if project-worker relation exists
    pw = await db.project_workers.find_one({
        "project_id": project_id,
        "worker_id": worker_id
    })
    
    if not pw:
        raise HTTPException(status_code=404, detail="Dolgozó nem található a projektben")
    
    # Update position_ids
    await db.project_workers.update_one(
        {"project_id": project_id, "worker_id": worker_id},
        {"$set": {
            "position_ids": data.position_ids,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Pozíciók frissítve"}

@api_router.get("/projects/{project_id}/archive")
async def get_project_archive(project_id: str, user: dict = Depends(get_current_user)):
    """
    KUKA - Negatív státuszú dolgozók (akik már voltak próbán, de nem váltak be)
    """
    # Get all statuses marked as "negative"
    negative_statuses = await db.statuses.find(
        {"status_type": "negative"},
        {"_id": 0}
    ).to_list(100)
    
    negative_status_ids = [s["id"] for s in negative_statuses]
    
    if not negative_status_ids:
        return {"workers": [], "count": 0}
    
    # Get workers with negative status in this project (from trials or project_workers)
    archived_workers = []
    
    # 1. Check project_workers for negative statuses
    project_worker_relations = await db.project_workers.find(
        {
            "project_id": project_id,
            "status_id": {"$in": negative_status_ids}
        },
        {"_id": 0}
    ).to_list(1000)
    
    for rel in project_worker_relations:
        worker = await db.workers.find_one({"id": rel["worker_id"]}, {"_id": 0})
        if worker:
            status = await db.statuses.find_one({"id": rel["status_id"]}, {"_id": 0})
            archived_workers.append({
                "id": worker["id"],
                "name": worker["name"],
                "phone": worker["phone"],
                "email": worker.get("email", ""),
                "status_id": rel["status_id"],
                "status_name": status["name"] if status else "Ismeretlen",
                "status_color": status.get("color", "#6b7280") if status else "#6b7280",
                "notes": rel.get("notes", ""),
                "updated_at": rel.get("updated_at", ""),
                "source": "project"
            })
    
    # 2. Check trial workers for negative statuses
    trials = await db.trials.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    for trial in trials:
        trial_workers = await db.trial_workers.find(
            {
                "trial_id": trial["id"],
                "status_id": {"$in": negative_status_ids}
            },
            {"_id": 0}
        ).to_list(1000)
        
        for tw in trial_workers:
            # Check if already added from project_workers
            if not any(w["id"] == tw["worker_id"] for w in archived_workers):
                worker = await db.workers.find_one({"id": tw["worker_id"]}, {"_id": 0})
                if worker:
                    status = await db.statuses.find_one({"id": tw["status_id"]}, {"_id": 0})
                    archived_workers.append({
                        "id": worker["id"],
                        "name": worker["name"],
                        "phone": worker["phone"],
                        "email": worker.get("email", ""),
                        "status_id": tw["status_id"],
                        "status_name": status["name"] if status else "Ismeretlen",
                        "status_color": status.get("color", "#6b7280") if status else "#6b7280",
                        "notes": tw.get("notes", ""),
                        "updated_at": tw.get("added_at", ""),
                        "trial_date": trial.get("date", ""),
                        "source": "trial"
                    })
    
    # Sort by updated_at desc
    archived_workers.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    
    return {
        "workers": archived_workers,
        "count": len(archived_workers)
    }

# ==================== WORKER LOG HELPER ====================

async def add_worker_log(worker_id: str, project_name: str, status_name: str, notes: str = ""):
    """
    Naplóbejegyzés hozzáadása a dolgozó megjegyzéseihez
    Formátum: [Projekt: {név}] Státusz: {státusz} - {megjegyzés} | {dátum}
    """
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        return
    
    date_str = datetime.now(timezone.utc).strftime("%Y.%m.%d %H:%M")
    log_entry = f"[Projekt: {project_name}] Státusz: {status_name}"
    if notes:
        log_entry += f" - {notes}"
    log_entry += f" | {date_str}"
    
    current_notes = worker.get("notes", "") or ""
    # Új bejegyzés a megjegyzések elejére
    new_notes = log_entry + "\n" + current_notes if current_notes else log_entry
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"notes": new_notes}}
    )

@api_router.post("/projects/{project_id}/archive/{worker_id}")
async def archive_worker_to_kuka(
    project_id: str, 
    worker_id: str, 
    data: dict,
    user: dict = Depends(get_current_user)
):
    """
    Dolgozó kukába helyezése a projekten belül - indokkal és naplózással
    ÚJ: Globális státusz visszaállítása "Feldolgozatlan"-ra
    ÚJ: Automatikus megjegyzés hozzáadása a dolgozóhoz
    """
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    reason = data.get("reason", "")
    if not reason:
        raise HTTPException(status_code=400, detail="Indok megadása kötelező")
    
    # Keressük meg a "Kuka" vagy negatív státuszt
    kuka_status = await db.statuses.find_one(
        {"$or": [{"name": "Kuka"}, {"name": "Nem vették fel"}, {"status_type": "negative"}]},
        {"_id": 0}
    )
    
    if kuka_status:
        # Frissítjük a projekt-dolgozó kapcsolat státuszát
        await db.project_workers.update_one(
            {"project_id": project_id, "worker_id": worker_id},
            {"$set": {
                "status_id": kuka_status["id"],
                "notes": f"Kuka indok: {reason}",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # ÚJ: Globális státusz visszaállítása "Feldolgozatlan"-ra
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"global_status": "Feldolgozatlan"}}
    )
    
    # ÚJ: Automatikus megjegyzés hozzáadása a dolgozó notes mezőjéhez
    company_name = project.get("client_name") or project["name"]
    auto_note = f"{company_name} ({project['name']}) - Kuka: {reason}"
    
    # Hozzáfűzzük a meglévő megjegyzéshez
    existing_notes = worker.get("notes", "")
    if existing_notes:
        new_notes = f"{existing_notes}\n\n{auto_note}"
    else:
        new_notes = auto_note
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"notes": new_notes}}
    )
    
    # Naplóbejegyzés a dolgozóhoz
    project_display = f"{project['name']}"
    if project.get("location"):
        project_display += f" | {project['location']}"
    
    await add_worker_log(
        worker_id=worker_id,
        project_name=project_display,
        status_name="Kuka",
        notes=reason
    )
    
    return {"message": "Dolgozó kukába helyezve", "reason": reason}



@api_router.post("/workers/{worker_id}/blacklist")
async def add_worker_to_blacklist(
    worker_id: str,
    data: dict,
    user: dict = Depends(get_current_user)
):
    """
    Dolgozó hozzáadása a Tiltólistához (USER-SPECIFIKUS!)
    
    Különbség Kuka vs Tiltólista:
    - Kuka: Nem felel meg valamiért (projekten belül)
    - Tiltólista: Nem akarom foglalkoztatni (csak TE nem látod, más igen!)
    
    FONTOS: 
    - Toborzó: Csak saját viewjában lesz tiltólistán
    - Admin: Látja MINDEN dolgozót, tiltólista státusz nélkül is
    """
    # Jogosultság ellenőrzés
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található vagy nincs jogosultságod")
    
    reason = data.get("reason", "")
    if not reason:
        raise HTTPException(status_code=400, detail="Indok megadása kötelező")
    
    # USER-SPECIFIKUS blacklist tárolása külön collection-ben
    blacklist_entry = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],  # KI tiltotta le
        "worker_id": worker_id,
        "reason": reason,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("name") or user["email"]
    }
    
    # Ellenőrizzük van-e már blacklist bejegyzés
    existing = await db.user_blacklist.find_one({
        "user_id": user["id"],
        "worker_id": worker_id
    })
    
    if existing:
        # Frissítjük a meglévőt
        await db.user_blacklist.update_one(
            {"user_id": user["id"], "worker_id": worker_id},
            {"$set": {
                "reason": reason,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Új bejegyzés
        await db.user_blacklist.insert_one(blacklist_entry)
    
    # Megjegyzés hozzáadása (csak info, nem globális státusz!)
    auto_note = f"🚫 TILTÓLISTA ({user.get('name') or user['email']}): {reason} (Hozzáadva: {datetime.now().strftime('%Y.%m.%d')})"
    
    existing_notes = worker.get("notes", "")
    if existing_notes:
        new_notes = f"{existing_notes}\n\n{auto_note}"
    else:
        new_notes = auto_note
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"notes": new_notes}}
    )
    
    # Audit log
    await audit_logger.log(
        user_id=user["id"],
        action="blacklist_add",
        resource_type="worker",
        resource_id=worker_id,
        details={
            "reason": reason,
            "worker_name": worker["name"],
            "added_by": user.get("name") or user["email"],
            "user_specific": True
        }
    )
    
    return {
        "success": True,
        "message": f"Dolgozó hozzáadva a te Tiltólistádhoz (csak te nem látod többé)",
        "worker_id": worker_id,
        "blacklisted_by": user.get("name") or user["email"]
    }


@api_router.delete("/workers/{worker_id}/blacklist")
async def remove_worker_from_blacklist(
    worker_id: str,
    user: dict = Depends(get_current_user)
):
    """
    Dolgozó eltávolítása a Tiltólistáról (USER-SPECIFIKUS)
    """
    # Töröljük a user-specifikus blacklist bejegyzést
    result = await db.user_blacklist.delete_one({
        "user_id": user["id"],
        "worker_id": worker_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nincs a te Tiltólistádon")
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    
    # Megjegyzés hozzáadása
    auto_note = f"✅ Tiltólistáról eltávolítva ({user.get('name') or user['email']}) - {datetime.now().strftime('%Y.%m.%d')}"
    
    if worker:
        existing_notes = worker.get("notes", "")
        if existing_notes:
            new_notes = f"{existing_notes}\n\n{auto_note}"
        else:
            new_notes = auto_note
        
        await db.workers.update_one(
            {"id": worker_id},
            {"$set": {"notes": new_notes}}
        )
    
    # Audit log
    await audit_logger.log(
        user_id=user["id"],
        action="blacklist_remove",
        resource_type="worker",
        resource_id=worker_id,
        details={
            "worker_name": worker.get("name") if worker else "Unknown",
            "removed_by": user.get("name") or user["email"]
        }
    )
    
    return {
        "success": True,
        "message": "Dolgozó eltávolítva a Tiltólistáról",
        "worker_id": worker_id
    }


@api_router.get("/workers/my-blacklist")
async def get_my_blacklist(user: dict = Depends(get_current_user)):
    """
    Saját Tiltólista lekérése
    """
    blacklist_entries = await db.user_blacklist.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).to_list(1000)
    
    # Worker adatok hozzáadása
    result = []
    for entry in blacklist_entries:
        worker = await db.workers.find_one({"id": entry["worker_id"]}, {"_id": 0})
        if worker:
            result.append({
                "blacklist_id": entry["id"],
                "worker_id": worker["id"],
                "worker_name": worker["name"],
                "worker_phone": worker["phone"],
                "worker_email": worker.get("email", ""),
                "reason": entry["reason"],
                "blacklisted_at": entry["created_at"],
                "blacklisted_by": entry["created_by"]
            })
    
    return {
        "blacklist": result,
        "count": len(result)
    }


@api_router.post("/projects/{project_id}/workers/{worker_id}/blacklist")
async def add_project_worker_to_blacklist(
    project_id: str,
    worker_id: str,
    data: dict,
    user: dict = Depends(get_current_user)
):
    """
    Dolgozó Tiltólista státuszba helyezése projekten belül
    """
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    reason = data.get("reason", "")
    if not reason:
        raise HTTPException(status_code=400, detail="Indok megadása kötelező")
    
    # Keressük meg a "Tiltólista" státuszt
    tiltolista_status = await db.statuses.find_one({"name": "Tiltólista"}, {"_id": 0})
    
    if not tiltolista_status:
        raise HTTPException(status_code=404, detail="Tiltólista státusz nem található")
    
    # Frissítjük a projekt-dolgozó kapcsolat státuszát
    await db.project_workers.update_one(
        {"project_id": project_id, "worker_id": worker_id},
        {"$set": {
            "status_id": tiltolista_status["id"],
            "notes": f"Tiltólista indok: {reason}",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Globális státusz is "Tiltólista"-ra (ha ez komolyabb)
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"global_status": "Tiltólista"}}
    )
    
    # Megjegyzés hozzáadása
    company_name = project.get("client_name") or project["name"]
    auto_note = f"🚫 {company_name} ({project['name']}) - TILTÓLISTA: {reason}"
    
    existing_notes = worker.get("notes", "")
    if existing_notes:
        new_notes = f"{existing_notes}\n\n{auto_note}"
    else:
        new_notes = auto_note
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"notes": new_notes}}
    )
    
    # Worker log
    project_display = f"{project['name']}"
    if project.get("location"):
        project_display += f" | {project['location']}"
    
    await add_worker_log(
        worker_id=worker_id,
        project_name=project_display,
        status_name="Tiltólista",
        notes=reason
    )
    
    return {
        "success": True,
        "message": "Dolgozó Tiltólistára helyezve",
        "project_id": project_id,
        "worker_id": worker_id
    }

@api_router.get("/projects/{project_id}/summary")
async def get_project_summary(project_id: str, user: dict = Depends(get_current_user)):
    """
    ÖSSZESÍTÉS - Státusz alapú statisztika
    """
    # Check user permission
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Admin sees all, toborzó only if assigned or owner
    if user["role"] != "admin":
        if project.get("owner_id") != user["id"] and user["id"] not in project.get("recruiter_ids", []):
            raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
    
    # Get all statuses
    all_statuses = await db.statuses.find({}, {"_id": 0}).to_list(100)
    
    # Count workers by status
    status_counts = []
    total_workers = 0
    
    for status in all_statuses:
        count = await db.project_workers.count_documents({
            "project_id": project_id,
            "status_id": status["id"]
        })
        
        if count > 0:
            status_counts.append({
                "status_id": status["id"],
                "status_name": status["name"],
                "status_type": status.get("status_type", "neutral"),
                "color": status.get("color", "#6b7280"),
                "count": count
            })
            total_workers += count
    
    # Sort by count desc
    status_counts.sort(key=lambda x: x["count"], reverse=True)
    
    # Position fill rate
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    total_headcount = sum(p.get("headcount", 0) for p in positions)
    
    # Workers in positive status (dolgozik, megfelelt, stb.)
    positive_statuses = [s for s in all_statuses if s.get("status_type") == "positive"]
    positive_status_ids = [s["id"] for s in positive_statuses]
    active_workers = await db.project_workers.count_documents({
        "project_id": project_id,
        "status_id": {"$in": positive_status_ids}
    })
    
    fill_rate = (active_workers / total_headcount * 100) if total_headcount > 0 else 0
    
    return {
        "total_workers": total_workers,
        "active_workers": active_workers,
        "total_positions": len(positions),
        "total_headcount": total_headcount,
        "fill_rate": round(fill_rate, 1),
        "status_breakdown": status_counts
    }

# ==================== EXCEL EXPORT ====================

async def generate_excel_for_user(user_id: str, user_name: str):
    """Generate Excel file for a specific recruiter with workers grouped by global status"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Use global statuses instead of categories
    statuses = ["Feldolgozatlan", "Máshol dolgozik", "Dolgozik", "Inaktív", "Kuka"]
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for status in statuses:
        workers = await db.workers.find(
            {"owner_id": user_id, "global_status": status}, {"_id": 0}
        ).sort("name", 1).to_list(1000)
        
        if not workers:
            continue
            
        # Create sheet for status
        ws = wb.create_sheet(title=status[:31])  # Excel max 31 chars
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Lakcím", "Típus", "Pozíció", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("position", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # If no sheets were created, add summary
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó")
    
    # Save file
    safe_name = "".join(c for c in user_name if c.isalnum() or c in " -_").strip() or "export"
    filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return filepath, filename

@api_router.get("/export/workers")
async def export_workers_excel(user: dict = Depends(get_current_user)):
    """Export current user's workers to Excel"""
    user_name = user.get("name") or user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user["id"], user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/workers/{user_id}")
async def export_user_workers_excel(user_id: str, admin: dict = Depends(require_admin)):
    """Admin can export any user's workers to Excel"""
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    user_name = target_user.get("name") or target_user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user_id, user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/all")
async def export_all_workers_excel(admin: dict = Depends(require_admin)):
    """Admin exports all workers grouped by recruiter and category"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    recruiter_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for u in users:
        workers = await db.workers.find({"owner_id": u["id"]}, {"_id": 0}).sort("category", 1).to_list(1000)
        
        if not workers:
            continue
        
        user_name = u.get("name") or u["email"].split("@")[0]
        sheet_name = user_name[:31]  # Excel max 31 chars
        
        # Handle duplicate sheet names
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{len(wb.sheetnames)}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Kategória", "Típus", "Lakcím", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker["category"]).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó a rendszerben")
    
    filename = f"osszes_dolgozo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    """Initialize default data"""
    # Check if already seeded
    admin = await db.users.find_one({"email": "admin@dolgozocrm.hu"})
    if admin:
        return {"message": "Adatok már léteznek"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "email": "admin@dolgozocrm.hu",
        "password": hash_password("admin123"),
        "name": "Admin",
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create test recruiter
    recruiter_doc = {
        "id": str(uuid.uuid4()),
        "email": "toborzo@dolgozocrm.hu",
        "password": hash_password("toborzo123"),
        "name": "Teszt Toborzó",
        "role": "user",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(recruiter_doc)
    
    # Worker types with positions
    type_positions = {
        "Betanított munkás": ["Csomagoló", "Komissiózó", "Összeszerelő", "Gyártósori munkás"],
        "Szakmunkás": ["Hegesztő", "Villanyszerelő", "Lakatos", "Esztergályos", "CNC gépkezelő", "Szerszámkészítő"],
        "Targoncás": ["Homlok targoncás", "Oldal targoncás", "Reach truck kezelő", "Magasraktári targoncás"],
        "Gépkezelő": ["Présgép kezelő", "Fröccsöntő gép kezelő", "Hajlítógép kezelő", "Varrógép kezelő"],
        "Raktáros": ["Áruátvevő", "Kiadó", "Leltáros", "Raktári adminisztrátor"],
        "Segédmunkás": ["Takarító", "Anyagmozgató", "Betanított segéd", "Kézi rakodó"]
    }
    
    for type_name, positions in type_positions.items():
        type_id = str(uuid.uuid4())
        await db.worker_types.insert_one({"id": type_id, "name": type_name})
        for pos in positions:
            await db.positions.insert_one({
                "id": str(uuid.uuid4()),
                "name": pos,
                "worker_type_id": type_id
            })
    
    # Statuses
    statuses = ["Jelentkezett", "Megerősítve", "Dolgozik", "Megfelelt", "Nem felelt meg", "Lemondta", "Nem jelent meg"]
    for s in statuses:
        await db.statuses.insert_one({"id": str(uuid.uuid4()), "name": s})
    
    # Tags
    tags = [
        {"name": "Megbízható", "color": "#22c55e"},
        {"name": "Tapasztalt", "color": "#3b82f6"},
        {"name": "Ajánlott", "color": "#f97316"},
        {"name": "Saját autó", "color": "#8b5cf6"},
        {"name": "Éjszakás", "color": "#6366f1"}
    ]
    for t in tags:
        await db.tags.insert_one({"id": str(uuid.uuid4()), **t})
    
    return {"message": "Seed adatok létrehozva", "admin_email": "admin@dolgozocrm.hu", "admin_password": "admin123"}

# ==================== EXCEL IMPORT ====================

class ExcelColumnMapping(BaseModel):
    name: Optional[int] = None  # Column index for name
    phone: Optional[int] = None
    email: Optional[int] = None
    address: Optional[int] = None
    position: Optional[int] = None
    experience: Optional[int] = None
    notes: Optional[int] = None

class ExcelImportSettings(BaseModel):
    column_mapping: Dict[str, Optional[int]]  # field_name -> column_index
    worker_type_id: str  # Required worker type
    category: str = "Felvitt dolgozók"  # Default category for all
    global_status: str = "Feldolgozatlan"
    start_row: int = 2  # Skip header row
    apply_same_to_all: bool = True  # Apply same category/type to all

@api_router.post("/workers/import/preview")
async def preview_excel_import(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Preview Excel file - returns first 10 rows and column headers"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Csak .xlsx vagy .xls fájl tölthető fel")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Handle Google Sheets exported files where max_row/max_column can be None
        max_row = ws.max_row or 1
        max_column = ws.max_column or 1
        
        # If max_row is still 1, try to find actual data by iterating
        if max_row <= 1:
            for row in ws.iter_rows(max_row=1000):
                for cell in row:
                    if cell.value is not None:
                        max_row = max(max_row, cell.row)
                        max_column = max(max_column, cell.column)
        
        # Get headers (first row)
        headers = []
        for col in range(1, max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Oszlop {col}")
        
        # Get preview rows (first 10 data rows)
        preview_rows = []
        for row in range(2, min(12, max_row + 1)):
            row_data = []
            for col in range(1, max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            preview_rows.append(row_data)
        
        wb.close()
        
        total_rows = max(0, max_row - 1)  # Exclude header
        
        return {
            "filename": file.filename,
            "total_rows": total_rows,
            "columns": headers,
            "preview_rows": preview_rows,
            "column_count": len(headers)
        }
    except Exception as e:
        logging.error(f"Excel preview error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Hiba a fájl olvasásakor: {str(e)}")

# ==================== CV/EXCEL IMPORT AI-VAL ====================

@api_router.post("/workers/import/cv-parse")
async def parse_cv_with_ai(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """
    CV/Excel fájl AI-alapú elemzése és adatok kinyerése
    Támogatott formátumok: PDF, DOCX, TXT, XLSX, XLS
    
    AI kivonja:
    - Név
    - Telefonszám
    - Email
    - Lakcím
    - Tapasztalat (évek)
    - Pozíció/szakma
    - Készségek
    """
    filename = file.filename.lower()
    
    # Támogatott fájl típusok
    supported = filename.endswith(('.pdf', '.docx', '.txt', '.xlsx', '.xls'))
    if not supported:
        raise HTTPException(
            status_code=400, 
            detail="Csak PDF, DOCX, TXT, XLSX, XLS fájlok támogatottak"
        )
    
    try:
        contents = await file.read()
        
        # Szöveg kinyerése fájl típus szerint
        if filename.endswith('.pdf'):
            text = extract_text_from_pdf(contents)
        elif filename.endswith('.docx'):
            text = extract_text_from_docx(contents)
        elif filename.endswith('.txt'):
            text = contents.decode('utf-8', errors='ignore')
        elif filename.endswith(('.xlsx', '.xls')):
            text = extract_text_from_excel(contents)
        else:
            raise HTTPException(status_code=400, detail="Nem támogatott fájl típus")
        
        if not text or len(text.strip()) < 20:
            raise HTTPException(status_code=400, detail="Nem sikerült szöveget kinyerni a fájlból")
        
        # AI-alapú adatkinyerés Groq Llama-val
        extracted_data = await extract_worker_data_with_ai(text)
        
        return {
            "success": True,
            "extracted_data": extracted_data,
            "original_text_preview": text[:500]  # Első 500 karakter preview
        }
        
    except Exception as e:
        logger.error(f"CV parse error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Hiba a CV feldolgozása során: {str(e)}")


def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """PDF szöveg kinyerése - több módszerrel próbálkozik"""
    import io
    
    text = ""
    
    # 1. Először pdfplumber-rel próbáljuk (jobb minőség)
    try:
        import pdfplumber
        pdf_file = io.BytesIO(pdf_bytes)
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        if text.strip():
            return text.strip()
    except Exception as e:
        logger.warning(f"pdfplumber failed: {e}")
    
    # 2. Ha nem sikerült, PyPDF2-vel próbáljuk
    try:
        from PyPDF2 import PdfReader
        pdf_file = io.BytesIO(pdf_bytes)
        reader = PdfReader(pdf_file, strict=False)
        
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        if text.strip():
            return text.strip()
    except Exception as e:
        logger.warning(f"PyPDF2 failed: {e}")
    
    # 3. Ha minden más nem sikerült
    raise ValueError("Nem sikerült szöveget kinyerni a PDF-ből. Próbálj meg egy másik formátumot (DOCX, TXT) vagy mentsd újra a PDF-et.")


def extract_text_from_docx(docx_bytes: bytes) -> str:
    """DOCX szöveg kinyerése"""
    # Egyszerű XML parsing DOCX-hez
    import zipfile
    import io
    import xml.etree.ElementTree as ET
    
    docx_file = io.BytesIO(docx_bytes)
    zf = zipfile.ZipFile(docx_file)
    
    # document.xml tartalmazza a szöveget
    xml_content = zf.read('word/document.xml')
    tree = ET.XML(xml_content)
    
    # Namespace
    namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    
    # Szöveg kinyerése
    paragraphs = tree.findall('.//w:t', namespaces)
    text = ' '.join([node.text for node in paragraphs if node.text])
    
    return text.strip()


def extract_text_from_excel(excel_bytes: bytes) -> str:
    """Excel szöveg kinyerése (minden cella összefűzése)"""
    import io
    from openpyxl import load_workbook
    
    wb = load_workbook(filename=io.BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb.active
    
    text_parts = []
    for row in ws.iter_rows(max_row=100):  # Max 100 sor
        for cell in row:
            if cell.value:
                text_parts.append(str(cell.value))
    
    wb.close()
    return ' '.join(text_parts)


async def extract_worker_data_with_ai(text: str) -> dict:
    """
    AI-alapú adatkinyerés Groq Llama 3.3 70B-vel
    Kivonja: név, telefon, email, cím, tapasztalat, pozíció, készségek
    """
    from groq import Groq
    import re
    
    if not GROQ_API_KEY:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY nincs beállítva")
    
    try:
        client = Groq(api_key=GROQ_API_KEY)
        
        # AI prompt
        prompt = f"""
Elemezd az alábbi CV/önéletrajz szöveget és vond ki a következő adatokat JSON formátumban.
Ha valamit nem találsz, használj null értéket.

Szöveg:
{text[:2000]}

Válasz CSAK ez a JSON legyen, semmi más szöveg előtte vagy utána:
{{"name": "Teljes név", "phone": "Telefonszám", "email": "email@example.com", "address": "Lakcím", "position": "Pozíció", "experience": "Tapasztalat", "skills": ["készség1", "készség2"], "notes": "Egyéb infók"}}
"""

        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Te egy CV elemző vagy. CSAK tiszta JSON-t adj vissza, semmilyen magyarázatot vagy extra szöveget ne írj!"},
                {"role": "user", "content": prompt}
            ],
            max_tokens=500,
            temperature=0.1
        )
        
        # Válasz parsing
        result_text = response.choices[0].message.content.strip()
        logger.info(f"AI raw response: {result_text[:500]}")
        
        # JSON kinyerése többféle módon
        extracted = None
        
        # 1. Próba: markdown kód blokk
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0].strip()
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0].strip()
        
        # 2. Próba: találjuk meg a JSON objektumot regex-szel
        json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', result_text, re.DOTALL)
        if json_match:
            result_text = json_match.group(0)
        
        # 3. Próba: parse
        try:
            extracted = json.loads(result_text)
        except json.JSONDecodeError:
            # Ha még mindig nem megy, próbáljuk a sorvégeket javítani
            result_text = result_text.replace('\n', ' ').replace('\r', '')
            # Keressük meg az első { és utolsó } között
            start = result_text.find('{')
            end = result_text.rfind('}')
            if start != -1 and end != -1:
                result_text = result_text[start:end+1]
                extracted = json.loads(result_text)
        
        if not extracted:
            raise ValueError("Nem sikerült JSON-t kinyerni")
        
        # AI Gender Detection a névre
        if extracted.get("name"):
            extracted["gender"] = detect_gender_from_name(extracted["name"])
        
        return extracted
        
    except Exception as e:
        logger.error(f"AI extraction error: {str(e)}")
        # Fallback: alapértelmezett üres adatok
        return {
            "name": None,
            "phone": None,
            "email": None,
            "address": None,
            "position": None,
            "experience": None,
            "skills": [],
            "notes": f"AI extraction hiba: {str(e)}"
        }


@api_router.post("/workers/import")
async def import_workers_from_excel(
    file: UploadFile = File(...),
    settings: str = Form(...),  # JSON string of ExcelImportSettings
    user: dict = Depends(get_current_user)
):
    """Import workers from Excel file with column mapping - HÁTTÉR FELDOLGOZÁS"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Csak .xlsx vagy .xls fájl tölthető fel")
    
    try:
        import_settings = json.loads(settings)
    except:
        raise HTTPException(status_code=400, detail="Érvénytelen beállítások")
    
    column_mapping = import_settings.get("column_mapping", {})
    worker_type_id = import_settings.get("worker_type_id", "")
    category = import_settings.get("category", "Felvitt dolgozók")
    global_status = import_settings.get("global_status", "Feldolgozatlan")
    start_row = import_settings.get("start_row", 2)
    
    # Validate worker_type_id
    worker_type = await db.worker_types.find_one({"id": worker_type_id})
    if not worker_type:
        raise HTTPException(status_code=400, detail="Érvénytelen dolgozó típus")
    
    # Name column is required
    name_col = column_mapping.get("name")
    if name_col is None:
        raise HTTPException(status_code=400, detail="Név oszlop megadása kötelező")
    
    # Save file temporarily
    contents = await file.read()
    
    # Create background job ID
    job_id = str(uuid.uuid4())
    
    # Save job info for tracking
    job_doc = {
        "id": job_id,
        "user_id": user["id"],
        "status": "processing",
        "total": 0,
        "processed": 0,
        "imported": 0,
        "skipped": 0,
        "errors": [],
        "started_at": datetime.now(timezone.utc).isoformat()
    }
    await db.import_jobs.insert_one(job_doc)
    
    # Schedule background processing
    scheduler.add_job(
        process_excel_import_background,
        args=[job_id, contents, column_mapping, worker_type_id, category, global_status, start_row, user["id"]],
        id=f"import_job_{job_id}",
        replace_existing=True
    )
    
    return {
        "message": "Import háttérben elindítva",
        "job_id": job_id,
        "status": "processing"
    }


async def process_excel_import_background(
    job_id: str,
    file_contents: bytes,
    column_mapping: dict,
    worker_type_id: str,
    category: str,
    global_status: str,
    start_row: int,
    user_id: str
):
    """Background job for processing Excel import with AI gender detection"""
    try:
        wb = load_workbook(filename=io.BytesIO(file_contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Handle Google Sheets exported files where max_row can be None
        max_row = ws.max_row or 1
        if max_row <= 1:
            # Try to find actual data
            for row in ws.iter_rows(max_row=2000):
                for cell in row:
                    if cell.value is not None:
                        max_row = max(max_row, cell.row)
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Process rows (up to 1000)
        max_rows = min(max_row, start_row + 1000)
        total_rows = max_rows - start_row + 1
        
        # Update total count
        await db.import_jobs.update_one(
            {"id": job_id},
            {"$set": {"total": total_rows}}
        )
        
        for row_idx in range(start_row, max_rows + 1):
            try:
                # Get name (required)
                name_value = ws.cell(row=row_idx, column=column_mapping.get("name") + 1).value
                if not name_value or str(name_value).strip() == "":
                    skipped_count += 1
                    continue
                
                name = str(name_value).strip()
                if len(name) < 2:
                    skipped_count += 1
                    continue
                
                # Get optional fields
                phone_col = column_mapping.get("phone")
                phone = str(ws.cell(row=row_idx, column=phone_col + 1).value or "").strip() if phone_col is not None else ""
                
                email_col = column_mapping.get("email")
                email = str(ws.cell(row=row_idx, column=email_col + 1).value or "").strip() if email_col is not None else ""
                
                address_col = column_mapping.get("address")
                address = str(ws.cell(row=row_idx, column=address_col + 1).value or "").strip() if address_col is not None else ""
                
                position_col = column_mapping.get("position")
                position = str(ws.cell(row=row_idx, column=position_col + 1).value or "").strip() if position_col is not None else ""
                
                experience_col = column_mapping.get("experience")
                experience = str(ws.cell(row=row_idx, column=experience_col + 1).value or "").strip() if experience_col is not None else ""
                
                notes_col = column_mapping.get("notes")
                notes = str(ws.cell(row=row_idx, column=notes_col + 1).value or "").strip() if notes_col is not None else ""
                
                # 🤖 AI-ALAPÚ GENDER DETECTION
                detected_gender = detect_gender_from_name(name)
                
                # Rate limit kezelés - max 30 név/perc
                if imported_count > 0 and imported_count % 25 == 0:
                    await asyncio.sleep(2)  # Kis szünet minden 25 név után
                
                # Create worker document
                worker_doc = {
                    "id": str(uuid.uuid4()),
                    "name": name,
                    "phone": phone,
                    "worker_type_id": worker_type_id,
                    "position": position,
                    "position_experience": "",
                    "category": category,
                    "address": address,
                    "email": email,
                    "experience": experience,
                    "notes": notes,
                    "gender": detected_gender,  # AI-alapú gender!
                    "global_status": global_status,
                    "tag_ids": [],
                    "owner_id": user_id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.workers.insert_one(worker_doc)
                imported_count += 1
                
                # Update progress
                processed = (row_idx - start_row + 1)
                await db.import_jobs.update_one(
                    {"id": job_id},
                    {"$set": {
                        "processed": processed,
                        "imported": imported_count,
                        "skipped": skipped_count
                    }}
                )
                
            except Exception as e:
                errors.append(f"Sor {row_idx}: {str(e)}")
                skipped_count += 1
        
        wb.close()
        
        # Mark job as completed
        await db.import_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "completed",
                "imported": imported_count,
                "skipped": skipped_count,
                "errors": errors[:10],
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # 🔔 ÉRTESÍTÉS KÜLDÉSE A FELHASZNÁLÓNAK
        await create_notification(
            user_id=user_id,
            notification_type="import_completed",
            title="✅ Excel import befejezve",
            message=f"{imported_count} dolgozó sikeresen importálva és nemük AI-val beazonosítva! ({skipped_count} átugorva)",
            link="/workers"
        )
        
        logger.info(f"Excel import completed: {imported_count} imported, {skipped_count} skipped")
        
    except Exception as e:
        # Mark job as failed
        await db.import_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "failed",
                "error": str(e),
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Send error notification
        await create_notification(
            user_id=user_id,
            notification_type="import_failed",
            title="❌ Excel import hiba",
            message=f"Hiba történt az import során: {str(e)}",
            link="/workers"
        )
        
        logger.error(f"Excel import failed: {str(e)}")


@api_router.get("/import-jobs/{job_id}")
async def get_import_job_status(job_id: str, user: dict = Depends(get_current_user)):
    """Get status of background import job"""
    job = await db.import_jobs.find_one({"id": job_id, "user_id": user["id"]}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    return job

# ==================== FTP SYNC / BACKUP ====================

def generate_recruiter_excel(recruiter_id: str, recruiter_name: str, workers: list) -> bytes:
    """Generate Excel file with all worker data for a recruiter"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dolgozók"
    
    # Headers
    headers = [
        "Teljes név", "Telefon", "Email", "Lakcím", "Pozíció", 
        "Tapasztalat", "Kategória", "Típus", "Globális státusz",
        "Projektek", "Megjegyzés", "Létrehozva"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, w in enumerate(workers, 2):
        # Get project statuses as string
        project_info = ""
        if w.get("project_statuses"):
            project_info = "; ".join([
                f"{ps.get('project_name', '')}: {ps.get('status_name', '')}" 
                for ps in w.get("project_statuses", [])
            ])
        
        row_data = [
            w.get("name", ""),
            w.get("phone", ""),
            w.get("email", ""),
            w.get("address", ""),
            w.get("position", ""),
            w.get("experience", ""),
            w.get("category", ""),
            w.get("worker_type_name", ""),
            w.get("global_status", "Feldolgozatlan"),
            project_info,
            w.get("notes", ""),
            w.get("created_at", "")[:10] if w.get("created_at") else ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [20, 15, 25, 30, 15, 20, 15, 15, 15, 40, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_admin_summary_excel(workers: list):
    """Generate Excel file with ALL workers for admin - includes recruiter info"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Összes dolgozó"
    
    # Headers - includes Toborzó column
    headers = [
        "Név", "Telefon", "Email", "Lakóhely", "Pozíció", "Tapasztalat", 
        "Kategória", "Típus", "Státusz", "Munka típus", "Saját autó",
        "Projektek", "Megjegyzések", "Toborzó", "Felvétel dátuma"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # Green for admin
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, w in enumerate(workers, 2):
        # Get project statuses as string
        project_info = ""
        if w.get("project_statuses"):
            project_info = "; ".join([
                f"{ps.get('project_name', '')}: {ps.get('status_name', '')}" 
                for ps in w.get("project_statuses", [])
            ])
        
        row_data = [
            w.get("name", ""),
            w.get("phone", ""),
            w.get("email", ""),
            w.get("address", ""),
            w.get("position", ""),
            w.get("experience", ""),
            w.get("category", ""),
            w.get("worker_type_name", ""),
            w.get("global_status", "Feldolgozatlan"),
            w.get("work_type", ""),
            w.get("has_car", ""),
            project_info,
            w.get("notes", ""),
            w.get("recruiter_name", ""),  # Toborzó neve
            w.get("created_at", "")[:10] if w.get("created_at") else ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [20, 15, 25, 30, 15, 20, 15, 15, 15, 12, 10, 40, 30, 20, 12]
    for i, width in enumerate(column_widths, 1):
        if i <= 15:  # Excel columns A-O
            col_letter = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
            ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

async def sync_to_ftp():
    """Sync all recruiter Excel files to FTP server"""
    # Use runtime config or env vars
    host = ftp_config.get("host") or FTP_HOST
    user = ftp_config.get("user") or FTP_USER
    password = ftp_config.get("password") or FTP_PASS
    folder = ftp_config.get("folder") or FTP_FOLDER
    
    if not host or not user or not password:
        logger.warning("FTP credentials not configured, skipping sync")
        return {"status": "skipped", "reason": "FTP not configured"}
    
    try:
        # Get all users (recruiters)
        users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
        
        # Connect to FTP
        ftp = ftplib.FTP(host)
        ftp.login(user, password)
        
        # Try to create/change to backup folder
        try:
            ftp.cwd(folder)
        except:
            try:
                ftp.mkd(folder)
                ftp.cwd(folder)
            except:
                pass  # Folder might already exist
        
        synced_files = []
        today = datetime.now().strftime("%Y-%m-%d")
        all_workers_for_admin = []  # Collect all workers for admin summary
        
        for u in users:
            user_id = u["id"]
            user_name = u.get("name", u.get("email", "unknown")).replace(" ", "_").replace("@", "_at_")
            is_admin = u.get("role") == "admin"
            
            # Get all workers for this user
            workers_cursor = db.workers.find({"owner_id": user_id}, {"_id": 0})
            workers = await workers_cursor.to_list(10000)
            
            # Add to admin summary
            for w in workers:
                w["recruiter_name"] = u.get("name", u.get("email", ""))
            all_workers_for_admin.extend(workers)
            
            if not workers:
                continue
            
            # Enrich workers with type names and project statuses
            for w in workers:
                # Get type name
                type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
                w["worker_type_name"] = type_doc["name"] if type_doc else ""
                
                # Get project statuses
                pw_list = await db.project_workers.find({"worker_id": w["id"]}, {"_id": 0}).to_list(100)
                project_statuses = []
                for pw in pw_list:
                    project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
                    status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
                    if project:
                        project_statuses.append({
                            "project_name": project["name"],
                            "status_name": status["name"] if status else "Hozzárendelve"
                        })
                w["project_statuses"] = project_statuses
            
            # Generate Excel for this recruiter
            excel_data = generate_recruiter_excel(user_id, user_name, workers)
            
            # Upload to FTP
            filename = f"{user_name}_dolgozok_{today}.xlsx"
            ftp.storbinary(f"STOR {filename}", io.BytesIO(excel_data))
            synced_files.append(filename)
            logger.info(f"FTP: Uploaded {filename} with {len(workers)} workers")
        
        # Generate admin summary with ALL workers
        if all_workers_for_admin:
            # Enrich all workers for admin file
            for w in all_workers_for_admin:
                type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
                w["worker_type_name"] = type_doc["name"] if type_doc else ""
                
                pw_list = await db.project_workers.find({"worker_id": w["id"]}, {"_id": 0}).to_list(100)
                project_statuses = []
                for pw in pw_list:
                    project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
                    status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
                    if project:
                        project_statuses.append({
                            "project_name": project["name"],
                            "status_name": status["name"] if status else "Hozzárendelve"
                        })
                w["project_statuses"] = project_statuses
            
            admin_excel = generate_admin_summary_excel(all_workers_for_admin)
            admin_filename = f"ADMIN_osszes_dolgozo_{today}.xlsx"
            ftp.storbinary(f"STOR {admin_filename}", io.BytesIO(admin_excel))
            synced_files.append(admin_filename)
            logger.info(f"FTP: Uploaded {admin_filename} with {len(all_workers_for_admin)} total workers")
        
        ftp.quit()
        
        return {
            "status": "success",
            "synced_files": synced_files,
            "date": today
        }
        
    except Exception as e:
        logger.error(f"FTP sync error: {str(e)}")
        return {"status": "error", "message": str(e)}

@api_router.post("/sync/ftp")
async def trigger_ftp_sync(user: dict = Depends(require_admin)):
    """Manually trigger FTP sync (admin only)"""
    result = await sync_to_ftp()
    return result

class FtpConfigUpdate(BaseModel):
    host: str
    user: str
    password: str
    folder: str = "/dolgozok_backup"

@api_router.post("/sync/config")
async def save_ftp_config(config: FtpConfigUpdate, user: dict = Depends(require_admin)):
    """Save FTP configuration (admin only)"""
    global ftp_config
    ftp_config = {
        "host": config.host,
        "user": config.user,
        "password": config.password,
        "folder": config.folder
    }
    # Also store in database for persistence
    await db.settings.update_one(
        {"key": "ftp_config"},
        {"$set": {"key": "ftp_config", "value": ftp_config}},
        upsert=True
    )
    return {"message": "FTP konfiguráció mentve"}

@api_router.get("/sync/status")
async def get_sync_status(user: dict = Depends(require_admin)):
    """Get FTP sync configuration status"""
    # Try to load from DB first
    db_config = await db.settings.find_one({"key": "ftp_config"}, {"_id": 0})
    if db_config and db_config.get("value"):
        global ftp_config
        ftp_config = db_config["value"]
    
    host = ftp_config.get("host") or FTP_HOST
    folder = ftp_config.get("folder") or FTP_FOLDER
    is_configured = bool(host and ftp_config.get("user") and ftp_config.get("password"))
    
    # Get last backup info
    last_backup_cursor = db.backup_logs.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).limit(1)
    last_backup_list = await last_backup_cursor.to_list(1)
    last_backup = last_backup_list[0] if last_backup_list else None
    
    last_backup_info = None
    if last_backup:
        last_backup_info = {
            "status": last_backup.get("status"),
            "date": last_backup.get("created_at", "")[:10] if last_backup.get("created_at") else "",
            "files_count": last_backup.get("files_count", 0)
        }
    
    return {
        "ftp_configured": is_configured,
        "ftp_host": host if host else "Not configured",
        "ftp_folder": folder,
        "last_backup": last_backup_info,
        "next_backup": "02:00 (naponta)"
    }

@api_router.get("/sync/logs")
async def get_backup_logs(user: dict = Depends(require_admin)):
    """Get backup logs (admin only)"""
    logs = await db.backup_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(20)
    return logs

# ==================== HEALTH ====================

# ==================== NOTIFICATIONS ====================

@api_router.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(user: dict = Depends(get_current_user)):
    """Felhasználó értesítéseinek lekérése"""
    notifications = await db.notifications.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return [NotificationResponse(**n) for n in notifications]

@api_router.get("/notifications/unread-count")
async def get_unread_notification_count(user: dict = Depends(get_current_user)):
    """Olvasatlan értesítések száma"""
    count = await db.notifications.count_documents({
        "user_id": user["id"],
        "is_read": False
    })
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(get_current_user)):
    """Értesítés olvasottként jelölése"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": user["id"]},
        {"$set": {"is_read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Értesítés nem található")
    return {"message": "Megjelölve olvasottként"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    """Összes értesítés olvasottként jelölése"""
    await db.notifications.update_many(
        {"user_id": user["id"], "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "Összes értesítés olvasottként jelölve"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user: dict = Depends(get_current_user)):
    """Értesítés törlése"""
    result = await db.notifications.delete_one({
        "id": notification_id,
        "user_id": user["id"]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Értesítés nem található")
    return {"message": "Értesítés törölve"}

# ==================== CALENDAR ====================

@api_router.get("/calendar/trials")
async def get_calendar_trials(user: dict = Depends(get_current_user)):
    """Naptár események - próbák. Admin minden próbát lát, toborzó a sajátját."""
    trials = await db.trials.find({}, {"_id": 0}).to_list(1000)
    
    events = []
    for trial in trials:
        project = await db.projects.find_one({"id": trial["project_id"]}, {"_id": 0})
        if not project:
            continue
        
        # Check visibility for recruiters
        if user["role"] != "admin":
            # Toborzó csak azokat a próbákat látja, ahol van dolgozója
            trial_workers = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
            worker_ids = [tw["worker_id"] for tw in trial_workers]
            
            # Check if any of these workers belong to the recruiter
            own_workers = await db.workers.count_documents({
                "id": {"$in": worker_ids},
                "owner_id": user["id"]
            })
            
            # Also check if recruiter is assigned to the project
            if own_workers == 0 and user["id"] not in project.get("recruiter_ids", []):
                continue
        
        # Get worker count for the trial
        trial_worker_count = await db.trial_workers.count_documents({"trial_id": trial["id"]})
        
        # Build event
        event_date = trial["date"]
        event_time = trial.get("time", "")
        
        # Create datetime string
        if event_time:
            start_datetime = f"{event_date}T{event_time}:00"
            # Assume 2 hour duration for trials
            end_hour = int(event_time.split(":")[0]) + 2
            end_time = f"{end_hour:02d}:{event_time.split(':')[1] if ':' in event_time else '00'}"
            end_datetime = f"{event_date}T{end_time}:00"
        else:
            start_datetime = f"{event_date}T09:00:00"
            end_datetime = f"{event_date}T11:00:00"
        
        events.append({
            "id": trial["id"],
            "title": f"{project['name']} - Próba",
            "start": start_datetime,
            "end": end_datetime,
            "project_id": project["id"],
            "project_name": project["name"],
            "project_client": project.get("client_name", ""),
            "location": project.get("location", ""),
            "training_location": project.get("training_location", ""),
            "notes": trial.get("notes", ""),
            "worker_count": trial_worker_count,
            "color": "#6366f1",  # Indigo
            "type": "trial"
        })
    
    return events

@api_router.get("/calendar/projects")
async def get_calendar_projects(user: dict = Depends(get_current_user)):
    """Projekt dátumok a naptárban"""
    if user["role"] == "admin":
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    else:
        # Toborzó csak hozzárendelt projekteket látja
        projects = await db.projects.find({
            "$or": [
                {"owner_id": user["id"]},
                {"recruiter_ids": user["id"]}
            ]
        }, {"_id": 0}).to_list(1000)
    
    events = []
    for p in projects:
        if not p.get("date"):
            continue
        
        worker_count = await db.project_workers.count_documents({"project_id": p["id"]})
        
        events.append({
            "id": p["id"],
            "title": p["name"],
            "start": f"{p['date']}T00:00:00",
            "end": f"{p['date']}T23:59:59",
            "allDay": True,
            "client_name": p.get("client_name", ""),
            "location": p.get("location", ""),
            "worker_count": worker_count,
            "is_closed": p.get("is_closed", False),
            "color": "#22c55e" if not p.get("is_closed") else "#94a3b8",  # Green or gray
            "type": "project"
        })
    
    return events

@api_router.get("/")
async def root():
    return {"message": "Dolgozó CRM API", "status": "running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ==================== FORMS API ====================

from google_sheets_helper import (
    fetch_public_sheet_data,
    auto_detect_columns,
    extract_row_data,
    get_preview_data,
    validate_column_mapping
)

@api_router.post("/forms/test-connection")
async def test_form_connection(data: dict, user: dict = Depends(get_current_user)):
    """Test Google Sheets connection and auto-detect columns"""
    sheet_url = data.get("sheet_url")
    if not sheet_url:
        raise HTTPException(status_code=400, detail="sheet_url kötelező")
    
    success, sheet_data, error = fetch_public_sheet_data(sheet_url)
    
    if not success:
        raise HTTPException(status_code=400, detail=error)
    
    # Auto-detect columns
    detected_mapping = auto_detect_columns(sheet_data)
    
    # Get preview - raw rows (not mapped, just the actual data)
    preview_rows = sheet_data[1:6] if len(sheet_data) > 1 else []  # First 5 data rows
    
    # Get preview with mapping
    preview = get_preview_data(sheet_data, detected_mapping, max_rows=5)
    
    return {
        "success": True,
        "row_count": len(sheet_data) - 1,  # Exclude header
        "headers": sheet_data[0] if sheet_data else [],
        "preview_rows": preview_rows,  # Raw data rows for frontend mapping
        "detected_mapping": detected_mapping,
        "preview": preview
    }

@api_router.get("/projects/{project_id}/forms", response_model=List[FormResponse])
async def get_project_forms(project_id: str, user: dict = Depends(get_current_user)):
    """Get forms for a project (filtered by permissions)"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Admin sees all forms
    if user["role"] == "admin":
        forms_cursor = db.project_forms.find(
            {"project_id": project_id},
            {"_id": 0}
        )
    else:
        # User sees only own + shared forms
        forms_cursor = db.project_forms.find(
            {
                "project_id": project_id,
                "$or": [
                    {"owner_id": user["id"]},
                    {"shared_with": user["id"]}
                ]
            },
            {"_id": 0}
        )
    
    forms = await forms_cursor.to_list(100)
    
    # Enrich with user names and lead counts
    result = []
    for form in forms:
        # Owner name
        owner = await db.users.find_one({"id": form["owner_id"]}, {"_id": 0})
        form["owner_name"] = owner.get("name", owner["email"]) if owner else "Ismeretlen"
        
        # Shared with names
        shared_names = []
        for user_id in form.get("shared_with", []):
            u = await db.users.find_one({"id": user_id}, {"_id": 0})
            if u:
                shared_names.append(u.get("name", u["email"]))
        form["shared_with_names"] = shared_names
        
        # Lead count
        lead_count = await db.form_leads.count_documents({
            "form_id": form["id"],
            "status": {"$in": ["unprocessed", "duplicate"]}
        })
        form["lead_count"] = lead_count
        
        result.append(FormResponse(**form))
    
    return result

@api_router.post("/projects/{project_id}/forms", response_model=FormResponse)
async def create_project_form(
    project_id: str,
    data: FormCreate,
    user: dict = Depends(get_current_user)
):
    """Create a new form for a project"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Validate column mapping
    is_valid, error_msg = validate_column_mapping(data.column_mapping)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Test connection
    success, sheet_data, error = fetch_public_sheet_data(data.sheet_url)
    if not success:
        raise HTTPException(status_code=400, detail=f"Kapcsolódási hiba: {error}")
    
    # Create form
    form_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "sheet_url": data.sheet_url,
        "name": data.name or "Google Űrlap",
        "column_mapping": data.column_mapping,
        "owner_id": user["id"],
        "shared_with": [],
        "default_category": data.default_category,
        "default_position_id": data.default_position_id or "",
        "sync_frequency": data.sync_frequency,
        "last_synced_at": None,
        "last_row_processed": 1,  # Skip header
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.project_forms.insert_one(form_doc)
    
    # Initial sync
    await sync_form(form_doc["id"])
    
    # Return with enriched data
    owner = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    form_doc["owner_name"] = owner.get("name", owner["email"])
    form_doc["shared_with_names"] = []
    form_doc["lead_count"] = await db.form_leads.count_documents({
        "form_id": form_doc["id"],
        "status": {"$in": ["unprocessed", "duplicate"]}
    })
    
    return FormResponse(**form_doc)

async def sync_form(form_id: str):
    """Sync a form - fetch new rows from Google Sheets"""
    form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    if not form:
        return
    
    # Fetch sheet data
    success, sheet_data, error = fetch_public_sheet_data(form["sheet_url"])
    if not success:
        logger.error(f"Form sync failed for {form_id}: {error}")
        return
    
    if len(sheet_data) <= form["last_row_processed"]:
        # No new rows
        await db.project_forms.update_one(
            {"id": form_id},
            {"$set": {"last_synced_at": datetime.now(timezone.utc).isoformat()}}
        )
        return
    
    # Process new rows
    new_rows = sheet_data[form["last_row_processed"]:]
    
    for row in new_rows:
        extracted = extract_row_data(row, form["column_mapping"])
        
        if not extracted.get("name") and not extracted.get("phone"):
            continue  # Skip empty rows
        
        phone = extracted.get("phone", "").strip()
        if not phone:
            continue
        
        # Check duplicate (only within owner's workers)
        existing_worker = await db.workers.find_one({
            "phone": phone,
            "$or": [
                {"created_by": form["owner_id"]},
                {"created_by": {"$in": form.get("shared_with", [])}}
            ]
        }, {"_id": 0})
        
        # Create lead
        lead_doc = {
            "id": str(uuid.uuid4()),
            "form_id": form_id,
            "project_id": form["project_id"],
            "name": extracted.get("name", ""),
            "phone": phone,
            "address": extracted.get("address", ""),
            "email": extracted.get("email", ""),
            "notes": extracted.get("notes", ""),
            "position": extracted.get("position", ""),  # Pozíció az űrlapból
            "default_position_id": form.get("default_position_id", ""),  # Alapértelmezett pozíció ID
            "submitted_at": extracted.get("date", ""),
            "status": "duplicate" if existing_worker else "unprocessed",
            "duplicate_worker_id": existing_worker["id"] if existing_worker else None,
            "processed_by": None,
            "processed_at": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.form_leads.insert_one(lead_doc)
    
    # Update form
    await db.project_forms.update_one(
        {"id": form_id},
        {
            "$set": {
                "last_synced_at": datetime.now(timezone.utc).isoformat(),
                "last_row_processed": len(sheet_data)
            }
        }
    )

@api_router.post("/projects/{project_id}/forms/{form_id}/sync")
async def manual_sync_form(
    project_id: str,
    form_id: str,
    user: dict = Depends(get_current_user)
):
    """Manually sync a form"""
    form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    if not form or form["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="Űrlap nem található")
    
    # Check permission
    if user["role"] != "admin" and form["owner_id"] != user["id"] and user["id"] not in form.get("shared_with", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod")
    
    await sync_form(form_id)
    
    return {"message": "Szinkronizálás befejezve"}

@api_router.put("/projects/{project_id}/forms/{form_id}", response_model=FormResponse)
async def update_project_form(
    project_id: str,
    form_id: str,
    data: FormUpdate,
    user: dict = Depends(get_current_user)
):
    """Update form settings (only owner or admin)"""
    form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    if not form or form["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="Űrlap nem található")
    
    # Check permission
    if user["role"] != "admin" and form["owner_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Csak a tulajdonos vagy admin módosíthatja")
    
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.column_mapping is not None:
        is_valid, error_msg = validate_column_mapping(data.column_mapping)
        if not is_valid:
            raise HTTPException(status_code=400, detail=error_msg)
        update_data["column_mapping"] = data.column_mapping
    if data.default_category is not None:
        update_data["default_category"] = data.default_category
    if data.default_position_id is not None:
        update_data["default_position_id"] = data.default_position_id
    if data.sync_frequency is not None:
        update_data["sync_frequency"] = data.sync_frequency
    
    if update_data:
        await db.project_forms.update_one({"id": form_id}, {"$set": update_data})
    
    updated_form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    owner = await db.users.find_one({"id": updated_form["owner_id"]}, {"_id": 0})
    updated_form["owner_name"] = owner.get("name", owner["email"])
    updated_form["shared_with_names"] = []
    updated_form["lead_count"] = 0
    
    return FormResponse(**updated_form)

@api_router.delete("/projects/{project_id}/forms/{form_id}")
async def delete_project_form(
    project_id: str,
    form_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete a form (only owner or admin)"""
    form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    if not form or form["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="Űrlap nem található")
    
    # Check permission
    if user["role"] != "admin" and form["owner_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Csak a tulajdonos vagy admin törölheti")
    
    # Delete form and leads
    await db.project_forms.delete_one({"id": form_id})
    await db.form_leads.delete_many({"form_id": form_id})
    
    return {"message": "Űrlap törölve"}

@api_router.post("/projects/{project_id}/forms/{form_id}/share")
async def share_form(
    project_id: str,
    form_id: str,
    data: FormShareRequest,
    user: dict = Depends(require_admin)  # Only admin can share
):
    """Share form with other recruiters (admin only)"""
    form = await db.project_forms.find_one({"id": form_id}, {"_id": 0})
    if not form or form["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="Űrlap nem található")
    
    # Update shared_with
    await db.project_forms.update_one(
        {"id": form_id},
        {"$set": {"shared_with": data.shared_with}}
    )
    
    return {"message": "Megosztás mentve", "shared_with": data.shared_with}

@api_router.get("/projects/{project_id}/form-leads", response_model=List[FormLeadResponse])
async def get_form_leads(project_id: str, user: dict = Depends(get_current_user)):
    """Get all unprocessed/duplicate leads for a project"""
    # Get forms user has access to
    if user["role"] == "admin":
        forms = await db.project_forms.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    else:
        forms = await db.project_forms.find(
            {
                "project_id": project_id,
                "$or": [
                    {"owner_id": user["id"]},
                    {"shared_with": user["id"]}
                ]
            },
            {"_id": 0}
        ).to_list(100)
    
    form_ids = [f["id"] for f in forms]
    
    # Get leads
    leads = await db.form_leads.find(
        {
            "form_id": {"$in": form_ids},
            "status": {"$in": ["unprocessed", "duplicate"]}
        },
        {"_id": 0}
    ).to_list(1000)
    
    # Enrich duplicates
    result = []
    for lead in leads:
        if lead["status"] == "duplicate" and lead.get("duplicate_worker_id"):
            worker = await db.workers.find_one({"id": lead["duplicate_worker_id"]}, {"_id": 0})
            if worker:
                lead["duplicate_worker"] = worker
        
        lead["is_duplicate"] = (lead["status"] == "duplicate")
        result.append(FormLeadResponse(**lead))
    
    return result

@api_router.post("/form-leads/{lead_id}/resolve")
async def resolve_duplicate_lead(
    lead_id: str,
    data: FormLeadResolve,
    user: dict = Depends(get_current_user)
):
    """Resolve a duplicate lead"""
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    if data.action == "keep_both":
        # Create new worker from lead
        form = await db.project_forms.find_one({"id": lead["form_id"]}, {"_id": 0})
        worker_type = await db.worker_types.find_one({}, {"_id": 0})
        
        worker_doc = {
            "id": str(uuid.uuid4()),
            "name": lead["name"],
            "phone": lead["phone"],
            "address": lead["address"],
            "email": lead["email"] or "",
            "notes": lead["notes"] or "",
            "worker_type_id": worker_type["id"] if worker_type else "",
            "position": "",
            "position_experience": "",
            "category": form["default_category"],
            "experience": "",
            "global_status": "Feldolgozatlan (űrlap)",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"]
        }
        await db.workers.insert_one(worker_doc)
        
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": "Új dolgozó létrehozva", "worker_id": worker_doc["id"]}
    
    elif data.action == "keep_existing":
        # Just mark as processed (ignore)
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "ignored", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Meglévő dolgozó megtartva"}
    
    elif data.action == "keep_new":
        # Update existing worker with new data
        if not lead.get("duplicate_worker_id"):
            raise HTTPException(status_code=400, detail="Nincs meglévő dolgozó")
        
        await db.workers.update_one(
            {"id": lead["duplicate_worker_id"]},
            {"$set": {
                "name": lead["name"],
                "address": lead["address"],
                "email": lead["email"] or "",
                "notes": lead["notes"] or "",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": "Meglévő dolgozó frissítve"}
    
    elif data.action == "merge":
        # Merge: update only empty fields
        if not lead.get("duplicate_worker_id"):
            raise HTTPException(status_code=400, detail="Nincs meglévő dolgozó")
        
        worker = await db.workers.find_one({"id": lead["duplicate_worker_id"]}, {"_id": 0})
        if not worker:
            raise HTTPException(status_code=404, detail="Meglévő dolgozó nem található")
        
        updates = {}
        if not worker.get("address") and lead.get("address"):
            updates["address"] = lead["address"]
        if not worker.get("email") and lead.get("email"):
            updates["email"] = lead["email"]
        if not worker.get("notes") and lead.get("notes"):
            updates["notes"] = lead["notes"]
        
        if updates:
            updates["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.workers.update_one({"id": lead["duplicate_worker_id"]}, {"$set": updates})
        
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": "Adatok egyesítve", "updated_fields": list(updates.keys())}
    
    raise HTTPException(status_code=400, detail="Érvénytelen művelet")

@api_router.post("/form-leads/{lead_id}/add-to-project")
async def add_lead_to_project(lead_id: str, data: dict = None, user: dict = Depends(get_current_user)):
    """Add lead as worker and to project with optional status and trial"""
    if data is None:
        data = {}
    
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    # Get form and worker type
    form = await db.project_forms.find_one({"id": lead["form_id"]}, {"_id": 0})
    worker_type = await db.worker_types.find_one({}, {"_id": 0})
    
    # Get status - either from request or default to "Feldolgozatlan"
    status_id = data.get("status_id")
    if not status_id:
        default_status = await db.statuses.find_one({"name": "Feldolgozatlan"}, {"_id": 0})
        status_id = default_status["id"] if default_status else ""
    
    # Get status name for global_status update
    status_doc = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    status_name = status_doc["name"] if status_doc else "Feldolgozatlan"
    
    # Set global_status based on project status
    global_status = "Feldolgozatlan"
    if status_name == "Dolgozik":
        global_status = "Dolgozik"
    elif status_name == "Próba megbeszélve":
        global_status = "Próba megbeszélve"
    elif status_name == "Próbára vár":
        global_status = "Próbára vár"
    elif status_name == "Tiltólista":
        global_status = "Tiltólista"
    
    # Create worker
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": lead["name"],
        "phone": lead["phone"],
        "address": lead.get("address", ""),
        "email": lead.get("email", ""),
        "notes": lead.get("notes", ""),
        "worker_type_id": worker_type["id"] if worker_type else "",
        "position": "",
        "position_experience": "",
        "category": form.get("default_category", "Ingázós") if form else "Ingázós",
        "experience": "",
        "global_status": global_status,
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.workers.insert_one(worker_doc)
    
    # Add to project workers (not waitlist)
    project_worker_doc = {
        "id": str(uuid.uuid4()),
        "project_id": lead["project_id"],
        "worker_id": worker_doc["id"],
        "status_id": status_id,
        "notes": f"Űrlapról: {form.get('name', 'Google Űrlap')}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If trial_id is provided (for "Próba megbeszélve"), add it
    trial_id = data.get("trial_id")
    trial_position_id = data.get("trial_position_id", "")
    
    if trial_id:
        project_worker_doc["trial_id"] = trial_id
        # Also add to trial_workers
        existing_trial_worker = await db.trial_workers.find_one({
            "trial_id": trial_id, 
            "worker_id": worker_doc["id"]
        })
        if not existing_trial_worker:
            trial_worker_doc = {
                "id": str(uuid.uuid4()),
                "trial_id": trial_id,
                "worker_id": worker_doc["id"],
                "trial_position_id": trial_position_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.trial_workers.insert_one(trial_worker_doc)
            
            # Update position assigned count if position was selected
            if trial_position_id:
                await db.trial_positions.update_one(
                    {"id": trial_position_id},
                    {"$inc": {"assigned_count": 1}}
                )
    
    await db.project_workers.insert_one(project_worker_doc)
    
    # Handle position_ids for project positions
    position_ids = data.get("position_ids", [])
    if position_ids:
        await db.project_worker_positions.delete_many({"project_worker_id": project_worker_doc["id"]})
        for pos_id in position_ids:
            await db.project_worker_positions.insert_one({
                "id": str(uuid.uuid4()),
                "project_worker_id": project_worker_doc["id"],
                "position_id": pos_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            # Update position current_count
            await db.project_positions.update_one(
                {"id": pos_id},
                {"$inc": {"current_count": 1}}
            )
    
    # Mark lead as processed
    await db.form_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Dolgozó hozzáadva: {status_name}", "worker_id": worker_doc["id"], "status": status_name}


@api_router.post("/form-leads/{lead_id}/mark-processed")
async def mark_lead_processed(lead_id: str, user: dict = Depends(get_current_user)):
    """Mark lead as processed - adds to project with 'Feldolgozatlan' status"""
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    # Get form and worker type
    form = await db.project_forms.find_one({"id": lead.get("form_id")}, {"_id": 0})
    worker_type = await db.worker_types.find_one({}, {"_id": 0})
    
    # Create worker
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": lead.get("name", ""),
        "phone": lead.get("phone", ""),
        "address": lead.get("address", ""),
        "email": lead.get("email", ""),
        "notes": lead.get("notes", ""),
        "worker_type_id": worker_type["id"] if worker_type else "",
        "position": "",
        "position_experience": "",
        "category": form.get("default_category", "Ingázós") if form else "Ingázós",
        "experience": "",
        "global_status": "Feldolgozatlan",
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.workers.insert_one(worker_doc)
    
    # Get "Feldolgozatlan" status
    feldolgozatlan_status = await db.statuses.find_one({"name": "Feldolgozatlan"}, {"_id": 0})
    status_id = feldolgozatlan_status["id"] if feldolgozatlan_status else ""
    
    # Add to project workers with "Feldolgozatlan" status
    project_worker_doc = {
        "id": str(uuid.uuid4()),
        "project_id": lead["project_id"],
        "worker_id": worker_doc["id"],
        "status_id": status_id,
        "notes": f"Űrlapról: {form.get('name', 'Google Űrlap')}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_workers.insert_one(project_worker_doc)
    
    # Mark lead as processed
    await db.form_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Dolgozó hozzáadva: Feldolgozatlan", "worker_id": worker_doc["id"]}


@api_router.delete("/form-leads/{lead_id}")
async def delete_form_lead(lead_id: str, user: dict = Depends(get_current_user)):
    """Permanently delete a form lead"""
    result = await db.form_leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    return {"message": "Jelentkező véglegesen törölve"}


@api_router.post("/form-leads/{lead_id}/add-to-database")
async def add_lead_to_main_database(lead_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Add lead to main worker database with full data"""
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    # Check for duplicate by name (only within user's workers if not admin)
    duplicate_query = {"name": {"$regex": f"^{data.get('name', '')}$", "$options": "i"}}
    if user["role"] != "admin":
        duplicate_query["owner_id"] = user["id"]
    
    existing_worker = await db.workers.find_one(duplicate_query, {"_id": 0})
    
    # Get worker type
    worker_type = await db.worker_types.find_one({}, {"_id": 0})
    
    # 🤖 AI-alapú gender detection
    worker_name = data.get("name", lead.get("name", ""))
    detected_gender = detect_gender_from_name(worker_name)
    
    # Create new worker
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": worker_name,
        "phone": data.get("phone", lead.get("phone", "")),
        "address": data.get("address", lead.get("address", "")),
        "email": data.get("email", lead.get("email", "")),
        "notes": data.get("notes", ""),
        "worker_type_id": worker_type["id"] if worker_type else "",
        "position": data.get("position", ""),
        "position_experience": "",
        "category": data.get("category", "Ingázós"),
        "work_type": data.get("work_type", "Ingázó"),
        "has_car": data.get("has_car", ""),
        "experience": "",
        "gender": detected_gender,  # AI-alapú gender
        "global_status": "Feldolgozatlan",
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    
    if existing_worker:
        # Return duplicate info for frontend to handle
        return {
            "duplicate": True,
            "new_worker": worker_doc,
            "existing_worker": existing_worker
        }
    
    # No duplicate, insert worker
    await db.workers.insert_one(worker_doc)
    
    # Mark lead as processed
    await db.form_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "processed", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"duplicate": False, "worker_id": worker_doc["id"], "message": "Dolgozó hozzáadva a fő adatbázishoz"}


@api_router.post("/form-leads/{lead_id}/add-to-kuka")
async def add_lead_to_kuka(lead_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Add lead to main database with 'Kuka' status and rejection reason"""
    reason = data.get("reason", "")
    notes = data.get("notes", "")
    
    if not reason:
        raise HTTPException(status_code=400, detail="Indok megadása kötelező!")
    
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    # Get worker type
    worker_type = await db.worker_types.find_one({}, {"_id": 0})
    
    # Get form and project for context
    form = await db.project_forms.find_one({"id": lead.get("form_id")}, {"_id": 0})
    project = await db.projects.find_one({"id": lead.get("project_id")}, {"_id": 0})
    
    # Create worker with Kuka status
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": lead.get("name", ""),
        "phone": lead.get("phone", ""),
        "address": lead.get("address", ""),
        "email": lead.get("email", ""),
        "notes": notes,
        "worker_type_id": worker_type["id"] if worker_type else "",
        "position": "",
        "position_experience": "",
        "category": "Kuka",  # Special category for rejected
        "experience": "",
        "global_status": "Kuka",  # Kuka status
        "kuka_reason": reason,  # Store the rejection reason
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "source_project_id": lead.get("project_id"),
        "source_project_name": project.get("name", "") if project else "",
        "source_form_id": lead.get("form_id")
    }
    await db.workers.insert_one(worker_doc)
    
    # Mark lead as processed
    await db.form_leads.update_one(
        {"id": lead_id},
        {"$set": {
            "status": "kuka",
            "processed_by": user["id"],
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "kuka_reason": reason
        }}
    )
    
    return {"message": "Dolgozó hozzáadva a Kukához", "worker_id": worker_doc["id"]}

@api_router.post("/form-leads/{lead_id}/resolve-duplicate")
async def resolve_lead_duplicate(lead_id: str, data: dict, user: dict = Depends(get_current_user)):
    """
    Duplikáció feloldása - 4 opció:
    - keep_old: Csak régit megtartom (új ignored, régi kap badge + opcionális mezők)
    - keep_new: Csak újat megtartom (régi archiválva, új dolgozó lesz)
    - both: Mindkettőt megtartom
    - merge: Összevonás (új mezők kitöltik üreseket, konfliktusoknál választás)
    """
    action = data.get("action")  # "keep_old", "keep_new", "both", "merge"
    existing_worker_id = data.get("existing_worker_id")
    update_fields = data.get("update_fields", {})  # keep_old esetén
    merge_conflicts = data.get("merge_conflicts", {})  # merge esetén
    
    lead = await db.form_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Jelentkező nem található")
    
    old_worker = None
    if existing_worker_id:
        old_worker = await db.workers.find_one({"id": existing_worker_id}, {"_id": 0})
    
    if action == "keep_old":
        # 1. Új lead ignored
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "ignored", "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # 2. Régi dolgozó badge
        badge_text = f"Űrlap jelentkezés: {datetime.now().strftime('%Y-%m-%d')}"
        await db.workers.update_one(
            {"id": existing_worker_id},
            {"$push": {"badges": badge_text}}
        )
        
        # 3. Opcionális új infók átvétele
        if update_fields:
            update_data = {}
            if update_fields.get("position") and lead.get("position"):
                update_data["position"] = lead["position"]
            if update_fields.get("email") and lead.get("email"):
                update_data["email"] = lead["email"]
            if update_fields.get("notes") and lead.get("notes"):
                existing_notes = old_worker.get("notes", "") if old_worker else ""
                update_data["notes"] = f"{existing_notes}\n\n[Űrlap - {datetime.now().strftime('%Y-%m-%d')}]\n{lead['notes']}"
            
            if update_data:
                await db.workers.update_one({"id": existing_worker_id}, {"$set": update_data})
        
        return {"message": "Régi dolgozó megtartva, új lead figyelmen kívül hagyva"}
    
    elif action == "keep_new":
        # 1. Régi archiválás (soft delete)
        if existing_worker_id:
            await db.workers.update_one(
                {"id": existing_worker_id},
                {"$set": {"archived": True, "archived_at": datetime.now(timezone.utc).isoformat()}}
            )
        
        # 2. Új dolgozó létrehozása
        worker_type = await db.worker_types.find_one({}, {"_id": 0})
        form = await db.project_forms.find_one({"id": lead.get("form_id")}, {"_id": 0})
        
        new_worker_id = str(uuid.uuid4())
        worker_doc = {
            "id": new_worker_id,
            "name": lead.get("name", ""),
            "phone": lead.get("phone", ""),
            "address": lead.get("address", ""),
            "email": lead.get("email", ""),
            "notes": lead.get("notes", ""),
            "worker_type_id": worker_type["id"] if worker_type else "",
            "position": lead.get("position", ""),
            "position_experience": "",
            "category": form.get("default_category", "Ingázós") if form else "Ingázós",
            "experience": "",
            "global_status": "Feldolgozatlan",
            "owner_id": user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"]
        }
        await db.workers.insert_one(worker_doc)
        
        # 3. Lead processed
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "processed", "worker_id": new_worker_id, "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": "Új dolgozó létrehozva, régi archiválva", "worker_id": new_worker_id}
    
    elif action == "both":
        # Mindkettő megtartása - lead unprocessed marad
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "unprocessed"}}
        )
        return {"message": "Mindkét rekord megtartva"}
    
    elif action == "merge":
        # Összevonás
        if not existing_worker_id or not old_worker:
            raise HTTPException(status_code=400, detail="Meglévő dolgozó nem található")
        
        merge_data = {}
        
        # Konfliktusfeldolgozás
        if merge_conflicts:
            if merge_conflicts.get("name") == "new":
                merge_data["name"] = lead.get("name", "")
            
            if merge_conflicts.get("position") == "new":
                merge_data["position"] = lead.get("position", "")
            elif not old_worker.get("position") and lead.get("position"):
                merge_data["position"] = lead.get("position", "")
            
            if merge_conflicts.get("notes") == "merge":
                existing_notes = old_worker.get("notes", "")
                new_notes = lead.get("notes", "")
                merge_data["notes"] = f"{existing_notes}\n\n[Új jelentkezés - {datetime.now().strftime('%Y-%m-%d')}]\n{new_notes}"
            elif merge_conflicts.get("notes") == "new_only":
                merge_data["notes"] = lead.get("notes", "")
        
        # Üres mezők automatikus kitöltése
        if not old_worker.get("email") and lead.get("email"):
            merge_data["email"] = lead.get("email", "")
        
        # Utolsó jelentkezés dátuma
        merge_data["last_application_date"] = datetime.now(timezone.utc).isoformat()
        
        # Frissítés
        if merge_data:
            await db.workers.update_one({"id": existing_worker_id}, {"$set": merge_data})
        
        # Lead processed
        await db.form_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "processed", "worker_id": existing_worker_id, "processed_by": user["id"], "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": "Rekordok összevonva", "worker_id": existing_worker_id}
    
    raise HTTPException(status_code=400, detail="Érvénytelen művelet")



# ==================== DASHBOARD ENDPOINTS ====================

@api_router.get("/dashboard/recruiter-stats")
async def get_recruiter_stats(user: dict = Depends(get_current_user)):
    """Toborzó saját statisztikái"""
    
    # Saját dolgozók lekérése
    workers = await db.workers.find({"owner_id": user["id"]}).to_list(10000)
    
    # Státusz szerinti csoportosítás
    status_counts = {
        "Feldolgozatlan": 0,
        "Próbára vár": 0,
        "Próba megbeszélve": 0,
        "Dolgozik": 0,
        "Tiltólista": 0
    }
    
    for w in workers:
        status = w.get("global_status", "Feldolgozatlan")
        if status in status_counts:
            status_counts[status] += 1
    
    # Havi placements (Dolgozik státuszba kerültek ebben a hónapban)
    this_month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0)
    
    # Dolgozik státusz ID-k
    dolgozik_statuses = await db.statuses.find({"name": "Dolgozik"}).to_list(100)
    dolgozik_status_ids = [s["id"] for s in dolgozik_statuses]
    
    worker_ids = [w["id"] for w in workers]
    
    monthly_placements = 0
    if dolgozik_status_ids and worker_ids:
        monthly_placements = await db.project_workers.count_documents({
            "status_id": {"$in": dolgozik_status_ids},
            "updated_at": {"$gte": this_month_start.isoformat()},
            "worker_id": {"$in": worker_ids}
        })
    
    # Hozzárendelt projektek száma
    assigned_projects = await db.projects.find({
        "$or": [
            {"owner_id": user["id"]},
            {"recruiter_ids": user["id"]}
        ]
    }).to_list(1000)
    
    return {
        "total_workers": len(workers),
        "status_counts": status_counts,
        "monthly_placements": monthly_placements,
        "assigned_projects_count": len(assigned_projects)
    }

@api_router.get("/dashboard/recruiter-monthly-performance")
async def get_recruiter_monthly_performance(user: dict = Depends(get_current_user)):
    """Toborzó havi placements az elmúlt 6 hónapra"""
    
    # Saját dolgozók ID-i
    workers = await db.workers.find({"owner_id": user["id"]}, {"_id": 0, "id": 1}).to_list(10000)
    worker_ids = [w["id"] for w in workers]
    
    # Dolgozik státusz ID-k
    dolgozik_statuses = await db.statuses.find({"name": "Dolgozik"}).to_list(100)
    dolgozik_status_ids = [s["id"] for s in dolgozik_statuses]
    
    # Elmúlt 6 hónap
    monthly_data = []
    for i in range(5, -1, -1):
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0) - timedelta(days=30*i)
        # Következő hónap első napja
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(seconds=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(seconds=1)
        
        # Placements ebben a hónapban
        placements = 0
        if dolgozik_status_ids and worker_ids:
            placements = await db.project_workers.count_documents({
                "worker_id": {"$in": worker_ids},
                "status_id": {"$in": dolgozik_status_ids},
                "updated_at": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}
            })
        
        monthly_data.append({
            "month": month_start.strftime("%Y-%m"),
            "month_name": month_start.strftime("%B"),
            "placements": placements
        })
    
    return monthly_data

@api_router.get("/dashboard/recruiter-todos")
async def get_recruiter_todos(user: dict = Depends(get_current_user)):
    """Toborzó teendői"""
    
    # Saját projektek
    my_projects = await db.projects.find({
        "$or": [
            {"recruiter_ids": user["id"]},
            {"owner_id": user["id"]}
        ]
    }).to_list(100)
    my_project_ids = [p["id"] for p in my_projects]
    
    # Új feldolgozatlan leadek
    unprocessed_leads = 0
    if my_project_ids:
        unprocessed_leads = await db.form_leads.count_documents({
            "project_id": {"$in": my_project_ids},
            "status": "unprocessed"
        })
    
    # Közelgő próbák (7 napon belül)
    week_from_now = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
    upcoming_trials = []
    if my_project_ids:
        upcoming_trials = await db.trials.find({
            "project_id": {"$in": my_project_ids},
            "date": {"$gte": datetime.now(timezone.utc).isoformat(), "$lte": week_from_now}
        }).to_list(100)
    
    # Régi feldolgozatlan dolgozók (60+ napja)
    days_60_ago = (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()
    stale_workers_60plus = await db.workers.count_documents({
        "owner_id": user["id"],
        "global_status": "Feldolgozatlan",
        "created_at": {"$lte": days_60_ago}
    })
    
    return {
        "unprocessed_leads": unprocessed_leads,
        "upcoming_trials": upcoming_trials,
        "stale_workers_60plus": stale_workers_60plus
    }

@api_router.get("/dashboard/admin-stats")
async def get_admin_stats(user: dict = Depends(require_admin)):
    """Admin - teljes cég statisztikái"""
    
    # Összes dolgozó
    all_workers = await db.workers.find({}).to_list(10000)
    
    # Státusz szerinti csoportosítás
    status_counts = {
        "Feldolgozatlan": 0,
        "Próbára vár": 0,
        "Próba megbeszélve": 0,
        "Dolgozik": 0,
        "Tiltólista": 0
    }
    
    for w in all_workers:
        status = w.get("global_status", "Feldolgozatlan")
        if status in status_counts:
            status_counts[status] += 1
    
    return {
        "total_workers": len(all_workers),
        "status_counts": status_counts
    }

@api_router.get("/dashboard/admin-recruiter-performance")
async def get_admin_recruiter_performance(user: dict = Depends(require_admin)):
    """Admin - toborzók teljesítménye"""
    
    # Összes toborzó
    recruiters = await db.users.find({"role": "user"}).to_list(100)
    
    # Dolgozik státusz ID-k
    dolgozik_statuses = await db.statuses.find({"name": "Dolgozik"}).to_list(100)
    dolgozik_status_ids = [s["id"] for s in dolgozik_statuses]
    
    performance = []
    for recruiter in recruiters:
        # Dolgozók száma
        workers = await db.workers.find({"owner_id": recruiter["id"]}).to_list(10000)
        
        # Státusz szerinti bontás
        status_counts = {
            "Feldolgozatlan": 0,
            "Próbára vár": 0,
            "Próba megbeszélve": 0,
            "Dolgozik": 0,
            "Tiltólista": 0
        }
        
        for w in workers:
            status = w.get("global_status", "Feldolgozatlan")
            if status in status_counts:
                status_counts[status] += 1
        
        # Havi placements (ebben a hónapban)
        this_month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0)
        worker_ids = [w["id"] for w in workers]
        
        monthly_placements = 0
        if dolgozik_status_ids and worker_ids:
            monthly_placements = await db.project_workers.count_documents({
                "worker_id": {"$in": worker_ids},
                "status_id": {"$in": dolgozik_status_ids},
                "updated_at": {"$gte": this_month_start.isoformat()}
            })
        
        # Átlagos átfutás (created_at → Dolgozik státusz updated_at)
        avg_days = 0
        dolgozik_workers = [w for w in workers if w.get("global_status") == "Dolgozik"]
        if dolgozik_workers:
            total_days = 0
            count = 0
            for w in dolgozik_workers:
                pw = await db.project_workers.find_one({
                    "worker_id": w["id"],
                    "status_id": {"$in": dolgozik_status_ids}
                })
                if pw and w.get("created_at") and pw.get("updated_at"):
                    try:
                        created = datetime.fromisoformat(w["created_at"].replace('Z', '+00:00'))
                        updated = datetime.fromisoformat(pw["updated_at"].replace('Z', '+00:00'))
                        days = (updated - created).days
                        if days >= 0:
                            total_days += days
                            count += 1
                    except:
                        pass
            
            if count > 0:
                avg_days = int(total_days / count)
        
        performance.append({
            "recruiter_name": recruiter.get("name") or recruiter["email"],
            "recruiter_id": recruiter["id"],
            "total_workers": len(workers),
            "feldolgozatlan": status_counts["Feldolgozatlan"],
            "probara_var": status_counts["Próbára vár"],
            "proba_megbeszeve": status_counts["Próba megbeszélve"],
            "dolgozik": status_counts["Dolgozik"],
            "tiltolista": status_counts["Tiltólista"],
            "monthly_placements": monthly_placements,
            "avg_conversion_days": avg_days
        })
    
    return performance

@api_router.get("/dashboard/admin-monthly-trend")
async def get_admin_monthly_trend(user: dict = Depends(require_admin)):
    """Admin - cég teljes havi trendje (elmúlt 6 hónap)"""
    
    dolgozik_statuses = await db.statuses.find({"name": "Dolgozik"}).to_list(100)
    dolgozik_status_ids = [s["id"] for s in dolgozik_statuses]
    
    monthly_data = []
    for i in range(5, -1, -1):
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0) - timedelta(days=30*i)
        # Következő hónap első napja
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(seconds=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(seconds=1)
        
        # Felvitt dolgozók ebben a hónapban
        added_workers = await db.workers.count_documents({
            "created_at": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}
        })
        
        # Placements ebben a hónapban
        placements = 0
        if dolgozik_status_ids:
            placements = await db.project_workers.count_documents({
                "status_id": {"$in": dolgozik_status_ids},
                "updated_at": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}
            })
        
        monthly_data.append({
            "month": month_start.strftime("%Y-%m"),
            "month_name": month_start.strftime("%B"),
            "added_workers": added_workers,
            "placements": placements
        })
    
    return monthly_data

@api_router.get("/dashboard/admin-alerts")
async def get_admin_alerts(user: dict = Depends(require_admin)):
    """Admin - kritikus figyelmeztetések"""
    
    # 90+ napja feldolgozatlan dolgozók
    days_90_ago = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
    stale_workers_90plus = await db.workers.count_documents({
        "global_status": "Feldolgozatlan",
        "created_at": {"$lte": days_90_ago}
    })
    
    # Próba megbeszélve státusz ID
    proba_statuses = await db.statuses.find({"name": "Próba megbeszélve"}).to_list(100)
    proba_status_ids = [s["id"] for s in proba_statuses]
    
    # Próba megbeszélve, de 48+ órája változatlan
    two_days_ago = (datetime.now(timezone.utc) - timedelta(days=2)).isoformat()
    stale_trials = 0
    if proba_status_ids:
        stale_trials = await db.project_workers.count_documents({
            "status_id": {"$in": proba_status_ids},
            "updated_at": {"$lte": two_days_ago}
        })
    
    # Új tiltólistások (elmúlt 7 nap)
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    new_blacklist = await db.workers.count_documents({
        "global_status": "Tiltólista",
        "created_at": {"$gte": week_ago}
    })
    
    # Legtöbb feldolgozatlan toborzó
    recruiters = await db.users.find({"role": "user"}).to_list(100)
    max_feldolgozatlan = 0
    max_recruiter_name = ""
    max_recruiter_id = ""
    
    for r in recruiters:
        count = await db.workers.count_documents({
            "owner_id": r["id"],
            "global_status": "Feldolgozatlan"
        })
        if count > max_feldolgozatlan:
            max_feldolgozatlan = count
            max_recruiter_name = r.get("name") or r["email"]
            max_recruiter_id = r["id"]
    
    return {
        "stale_workers_90plus": stale_workers_90plus,
        "stale_trials": stale_trials,
        "new_blacklist": new_blacklist,
        "top_feldolgozatlan": {
            "name": max_recruiter_name,
            "count": max_feldolgozatlan,
            "recruiter_id": max_recruiter_id
        }
    }

# ==================== APP CONFIG ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize scheduler for daily backups
scheduler = AsyncIOScheduler()

async def daily_backup_job():
    """Daily backup job - runs at 2:00 AM"""
    logger.info("Starting daily FTP backup job...")
    try:
        result = await sync_to_ftp()
        if result.get("status") == "success":
            logger.info(f"Daily backup completed successfully: {len(result.get('synced_files', []))} files")
            # Store backup log
            await db.backup_logs.insert_one({
                "id": str(uuid.uuid4()),
                "type": "daily_auto",
                "status": "success",
                "files_count": len(result.get('synced_files', [])),
                "synced_files": result.get('synced_files', []),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        else:
            logger.warning(f"Daily backup skipped or failed: {result.get('reason', result.get('message', 'unknown'))}")
            await db.backup_logs.insert_one({
                "id": str(uuid.uuid4()),
                "type": "daily_auto",
                "status": result.get("status", "error"),
                "message": result.get("reason", result.get("message", "")),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    except Exception as e:
        logger.error(f"Daily backup job error: {str(e)}")
        await db.backup_logs.insert_one({
            "id": str(uuid.uuid4()),
            "type": "daily_auto",
            "status": "error",
            "message": str(e),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

@app.on_event("startup")
async def startup_event():
    """Initialize scheduler on startup"""
    # Load FTP config from database
    db_config = await db.settings.find_one({"key": "ftp_config"}, {"_id": 0})
    if db_config and db_config.get("value"):
        global ftp_config
        ftp_config = db_config["value"]
        logger.info(f"Loaded FTP config from database: {ftp_config.get('host', 'N/A')}")
    
    # Inicializáljuk az alapértelmezett státuszokat ha nincsenek
    # Egységes státuszok: Feldolgozatlan, Próbára vár, Próba megbeszélve, Dolgozik, Tiltólista, Kuka (projekt szintű)
    existing_statuses = await db.statuses.count_documents({})
    if existing_statuses == 0:
        default_statuses = [
            {"name": "Feldolgozatlan", "status_type": "neutral", "color": "#9CA3AF"},
            {"name": "Próbára vár", "status_type": "neutral", "color": "#F97316"},
            {"name": "Próba megbeszélve", "status_type": "neutral", "color": "#8B5CF6"},
            {"name": "Dolgozik", "status_type": "positive", "color": "#10B981"},
            {"name": "Tiltólista", "status_type": "negative", "color": "#EF4444"},
            {"name": "Kuka", "status_type": "negative", "color": "#6B7280"}
        ]
        for status_data in default_statuses:
            await db.statuses.insert_one({
                "id": str(uuid.uuid4()),
                "name": status_data["name"],
                "status_type": status_data["status_type"],
                "color": status_data["color"],
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        logger.info(f"✅ {len(default_statuses)} alapértelmezett státusz létrehozva")
    else:
        # Ellenőrizzük, hogy a "Tiltólista" státusz létezik-e, ha nem, hozzáadjuk
        tiltolista = await db.statuses.find_one({"name": "Tiltólista"})
        if not tiltolista:
            await db.statuses.insert_one({
                "id": str(uuid.uuid4()),
                "name": "Tiltólista",
                "status_type": "negative",
                "color": "#EF4444",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            logger.info("✅ Tiltólista státusz hozzáadva")
        
        # Frissítsük a "Feldolgozás alatt" státuszt "Feldolgozatlan"-ra ha létezik
        feldolgozas = await db.statuses.find_one({"name": "Feldolgozás alatt"})
        if feldolgozas:
            await db.statuses.update_one(
                {"name": "Feldolgozás alatt"},
                {"$set": {"name": "Feldolgozatlan", "color": "#9CA3AF"}}
            )
            logger.info("✅ Feldolgozás alatt → Feldolgozatlan átnevezve")
    
    # ==================== DATABASE INDEXEK LÉTREHOZÁSA ====================
    logger.info("🔧 MongoDB indexek létrehozása/ellenőrzése...")
    
    try:
        # WORKERS collection indexek (legtöbb lekérdezés itt van)
        await db.workers.create_index([("owner_id", 1)])  # Toborzó saját dolgozói
        await db.workers.create_index([("email", 1)], sparse=True)  # Email keresés
        await db.workers.create_index([("phone", 1)])  # Telefonszám keresés
        await db.workers.create_index([("global_status", 1)])  # Státusz szűrés
        await db.workers.create_index([("worker_type_id", 1)])  # Típus szűrés
        await db.workers.create_index([("county", 1)])  # Megye szűrés
        await db.workers.create_index([("created_at", -1)])  # Legújabbak elöl
        await db.workers.create_index([("owner_id", 1), ("global_status", 1)])  # Kombinált szűrés
        await db.workers.create_index([("position", "text"), ("name", "text")])  # Full-text search
        
        # PROJECT_WORKERS collection (projekt-dolgozó kapcsolatok)
        await db.project_workers.create_index([("project_id", 1), ("worker_id", 1)], unique=True)
        await db.project_workers.create_index([("worker_id", 1)])  # Dolgozó projektjei
        await db.project_workers.create_index([("project_id", 1), ("status_id", 1)])  # Projekt státuszok
        await db.project_workers.create_index([("project_id", 1), ("added_at", -1)])  # Legutóbb hozzáadottak
        
        # PROJECTS collection
        await db.projects.create_index([("owner_id", 1)])  # Toborzó projektjei
        await db.projects.create_index([("is_closed", 1)])  # Aktív/lezárt projektek
        await db.projects.create_index([("date", 1)])  # Dátum szerinti rendezés
        await db.projects.create_index([("recruiter_ids", 1)])  # Multi-toborzós projektek
        await db.projects.create_index([("created_at", -1)])  # Legújabbak
        
        # USERS collection
        await db.users.create_index([("email", 1)], unique=True)  # Email login
        await db.users.create_index([("role", 1)])  # Admin/User szűrés
        
        # SECURITY & AUDIT collections
        await db.login_attempts.create_index([("email", 1), ("timestamp", -1)])  # Rate limiting
        await db.login_attempts.create_index([("timestamp", 1)], expireAfterSeconds=86400)  # 24h TTL
        await db.audit_logs.create_index([("user_id", 1), ("timestamp", -1)])  # User audit log
        await db.audit_logs.create_index([("resource_type", 1), ("timestamp", -1)])  # Resource audit
        await db.audit_logs.create_index([("timestamp", -1)])  # Legutóbbi események
        
        # EMAIL collections
        await db.email_queue.create_index([("status", 1), ("scheduled_at", 1)])  # Email küldési sor
        await db.email_logs.create_index([("user_id", 1), ("sent_at", -1)])  # User email történet
        await db.email_templates.create_index([("user_id", 1)])  # User sablonok
        await db.unsubscribe_tokens.create_index([("token", 1)], unique=True)  # Leiratkozás token
        await db.unsubscribe_tokens.create_index([("worker_id", 1)])  # Dolgozó leiratkozások
        await db.gmail_tokens.create_index([("user_id", 1)], unique=True)  # Gmail OAuth token
        
        # TAGS & TYPES
        await db.tags.create_index([("owner_id", 1)])  # User tagek
        await db.worker_types.create_index([("owner_id", 1)])  # User típusok
        
        # STATUSES
        await db.statuses.create_index([("name", 1)])  # Státusz név keresés
        
        # USER_BLACKLIST (user-specific blacklist)
        await db.user_blacklist.create_index([("user_id", 1), ("worker_id", 1)], unique=True)
        await db.user_blacklist.create_index([("worker_id", 1)])
        
        logger.info("✅ MongoDB indexek sikeresen létrehozva!")
    except Exception as e:
        logger.error(f"⚠️ Index létrehozási hiba (nem kritikus): {e}")
    
    # Fő admin felhasználó létrehozása/frissítése
    main_admin = await db.users.find_one({"email": "kaszasdominik@gmail.com"})
    if not main_admin:
        admin_doc = {
            "id": str(uuid.uuid4()),
            "email": "kaszasdominik@gmail.com",
            "password": hash_password("Kokkernomokker132"),
            "name": "Kaszás Dominik",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_doc)
        logger.info("✅ Admin felhasználó létrehozva (kaszasdominik@gmail.com)")
    
    # Schedule daily backup at 2:00 AM
    scheduler.add_job(
        daily_backup_job,
        CronTrigger(hour=2, minute=0),
        id="daily_ftp_backup",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started - Daily backup scheduled at 02:00")
    
    # Schedule email queue processor every minute
    scheduler.add_job(
        process_email_queue,
        'interval',
        minutes=1,
        id="email_queue_processor",
        replace_existing=True
    )
    logger.info("Email queue processor scheduled")
    
    # Schedule weekly summary email (every Monday at 8:00 AM)
    scheduler.add_job(
        weekly_summary_email,
        CronTrigger(day_of_week='mon', hour=8, minute=0),
        id="weekly_summary_email",
        replace_existing=True
    )
    logger.info("📊 Weekly summary email scheduled: Every Monday at 08:00")
    
    # Schedule GDPR old workers notification (every 1st of month at 9:00 AM)
    scheduler.add_job(
        check_old_workers_notification,
        CronTrigger(day=1, hour=9, minute=0),
        id="gdpr_old_workers_check",
        replace_existing=True
    )
    logger.info("⚠️ GDPR old workers check scheduled: Every 1st of month at 09:00")


@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()


# ==================== BULK EMAIL ENDPOINTS ====================

from bulk_email import (
    get_authorization_url,
    exchange_code_for_tokens,
    get_gmail_credentials,
    send_email_via_gmail,
    generate_unsubscribe_token,
    replace_template_variables,
    DAILY_EMAIL_LIMIT
)
from fastapi.responses import RedirectResponse

# Pydantic models for Bulk Email
class EmailTemplateCreate(BaseModel):
    name: str
    subject: str
    body: str

class EmailTemplateUpdate(BaseModel):
    name: Optional[str] = None
    subject: Optional[str] = None
    body: Optional[str] = None

class WorkerTemplateCreate(BaseModel):
    name: str
    filters: Dict[str, Any]  # Store filter criteria

class WorkerTemplateUpdate(BaseModel):
    name: Optional[str] = None
    filters: Optional[Dict[str, Any]] = None

class BulkEmailCampaignCreate(BaseModel):
    name: str
    email_template_id: Optional[str] = None
    subject: Optional[str] = None
    body: Optional[str] = None
    worker_ids: List[str]
    send_method: str = "individual"  # "individual" or "bcc"

# Gmail OAuth endpoints
@app.get("/api/bulk-email/gmail/auth-url")
async def get_gmail_auth_url(
    current_user: dict = Depends(get_current_user)
):
    """Get Gmail OAuth authorization URL"""
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    redirect_uri = f"{frontend_url}/api/oauth/gmail/callback"
    
    logger.info(f"Gmail Auth URL request from user: {current_user['id']} ({current_user.get('email')})")
    logger.info(f"Gmail OAuth redirect_uri: {redirect_uri}")
    
    # Generate state token
    state = secrets.token_urlsafe(32)
    
    # Store state with user_id
    await db.oauth_states.insert_one({
        "state": state,
        "user_id": current_user["id"],
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=10)
    })
    
    logger.info(f"Gmail OAuth state stored: {state[:20]}... for user {current_user['id']}")
    
    auth_url = get_authorization_url(redirect_uri, state)
    if not auth_url:
        logger.error("Gmail OAuth: GOOGLE_CLIENT_ID or GOOGLE_CLIENT_SECRET not configured")
        raise HTTPException(
            status_code=500,
            detail="Gmail OAuth nincs konfigurálva. Kérjük állítsa be a GOOGLE_CLIENT_ID és GOOGLE_CLIENT_SECRET környezeti változókat."
        )
    
    logger.info(f"Gmail OAuth auth_url generated successfully")
    return {"auth_url": auth_url}


@app.get("/api/oauth/gmail/callback")
async def gmail_oauth_callback(code: str, state: str):
    """Handle Gmail OAuth callback"""
    logger.info(f"Gmail OAuth callback received - state: {state[:20]}...")
    
    # Verify state
    state_doc = await db.oauth_states.find_one({"state": state})
    if not state_doc:
        logger.error(f"Gmail OAuth: Invalid state - not found in database")
        return RedirectResponse(url="/bulk-email?error=invalid_state")
    
    if datetime.now(timezone.utc) > state_doc.get("expires_at", datetime.now(timezone.utc)):
        logger.error(f"Gmail OAuth: State expired for user {state_doc.get('user_id')}")
        await db.oauth_states.delete_one({"state": state})
        return RedirectResponse(url="/bulk-email?error=expired_state")
    
    user_id = state_doc["user_id"]
    logger.info(f"Gmail OAuth: Valid state for user_id: {user_id}")
    await db.oauth_states.delete_one({"state": state})
    
    # Exchange code for tokens - use the SAME redirect_uri as in auth URL
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    redirect_uri = f"{frontend_url}/api/oauth/gmail/callback"
    logger.info(f"Gmail OAuth: Using redirect_uri: {redirect_uri}")
    
    token_data = await exchange_code_for_tokens(code, redirect_uri)
    if not token_data:
        logger.error(f"Gmail OAuth: Token exchange failed for user {user_id}")
        return RedirectResponse(url="/bulk-email?error=token_exchange_failed")
    
    logger.info(f"Gmail OAuth: Token exchange successful for email: {token_data.get('email')}")
    
    # Store tokens for user
    result = await db.gmail_tokens.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "user_id": user_id,
                "access_token": token_data["access_token"],
                "refresh_token": token_data["refresh_token"],
                "token_uri": token_data.get("token_uri", "https://oauth2.googleapis.com/token"),
                "client_id": token_data["client_id"],
                "client_secret": token_data["client_secret"],
                "expires_at": token_data["expires_at"],
                "gmail_email": token_data["email"],
                "gmail_name": token_data.get("name"),
                "connected_at": datetime.now(timezone.utc)
            }
        },
        upsert=True
    )
    
    logger.info(f"Gmail OAuth: Token saved for user {user_id}, modified: {result.modified_count}, upserted: {result.upserted_id is not None}")
    
    return RedirectResponse(url="/bulk-email?success=gmail_connected")


@app.get("/api/bulk-email/gmail/status")
async def get_gmail_status(current_user: dict = Depends(get_current_user)):
    """Get current user's Gmail connection status"""
    logger.info(f"Gmail status check for user: {current_user['id']} ({current_user.get('email')})")
    
    token = await db.gmail_tokens.find_one(
        {"user_id": current_user["id"]},
        {"_id": 0, "access_token": 0, "refresh_token": 0, "client_secret": 0}
    )
    
    if not token:
        logger.info(f"Gmail status: No token found for user {current_user['id']}")
        return {"connected": False}
    
    logger.info(f"Gmail status: Token found for user {current_user['id']}, gmail: {token.get('gmail_email')}")
    
    # Get today's sent count
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_today = await db.email_logs.count_documents({
        "sender_user_id": current_user["id"],
        "sent_at": {"$gte": today_start},
        "status": "sent"
    })
    
    # Calculate reset time (next midnight UTC)
    next_reset = today_start + timedelta(days=1)
    
    return {
        "connected": True,
        "gmail_email": token.get("gmail_email"),
        "gmail_name": token.get("gmail_name"),
        "connected_at": token.get("connected_at"),
        "sent_today": sent_today,
        "daily_limit": DAILY_EMAIL_LIMIT,
        "remaining_today": max(0, DAILY_EMAIL_LIMIT - sent_today),
        "next_reset": next_reset.isoformat()
    }


@app.delete("/api/bulk-email/gmail/disconnect")
async def disconnect_gmail(current_user: dict = Depends(get_current_user)):
    """Disconnect Gmail account"""
    await db.gmail_tokens.delete_one({"user_id": current_user["id"]})
    return {"message": "Gmail fiók leválasztva"}


@app.get("/api/bulk-email/gmail/debug")
async def debug_gmail_oauth(current_user: dict = Depends(get_current_user)):
    """Debug endpoint for Gmail OAuth - only for admins"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin")
    
    frontend_url = os.environ.get('FRONTEND_URL', 'NOT_SET')
    google_client_id = os.environ.get('GOOGLE_CLIENT_ID', 'NOT_SET')
    
    # Check token in database
    token = await db.gmail_tokens.find_one(
        {"user_id": current_user["id"]},
        {"_id": 0, "access_token": 0, "refresh_token": 0, "client_secret": 0}
    )
    
    # Count all tokens
    all_tokens_count = await db.gmail_tokens.count_documents({})
    
    # Get all user IDs with tokens
    all_tokens = await db.gmail_tokens.find({}, {"user_id": 1, "gmail_email": 1, "_id": 0}).to_list(100)
    
    # Get pending states
    pending_states = await db.oauth_states.count_documents({})
    
    return {
        "environment": {
            "FRONTEND_URL": frontend_url,
            "GOOGLE_CLIENT_ID_SET": google_client_id != 'NOT_SET',
            "GOOGLE_CLIENT_SECRET_SET": os.environ.get('GOOGLE_CLIENT_SECRET') is not None
        },
        "current_user": {
            "id": current_user["id"],
            "email": current_user.get("email")
        },
        "gmail_token_for_user": token,
        "all_gmail_tokens_count": all_tokens_count,
        "all_gmail_tokens": all_tokens,
        "pending_oauth_states": pending_states,
        "expected_redirect_uri": f"{frontend_url}/api/oauth/gmail/callback"
    }


# Email Templates CRUD
@app.get("/api/bulk-email/templates")
async def get_email_templates(current_user: dict = Depends(get_current_user)):
    """Get user's email templates"""
    templates = await db.email_templates.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return templates


@app.post("/api/bulk-email/templates")
async def create_email_template(
    template: EmailTemplateCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new email template"""
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "name": template.name,
        "subject": template.subject,
        "body": template.body,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    await db.email_templates.insert_one(doc)
    del doc["_id"]
    return doc


@app.put("/api/bulk-email/templates/{template_id}")
async def update_email_template(
    template_id: str,
    template: EmailTemplateUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update email template"""
    update_data = {k: v for k, v in template.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.email_templates.update_one(
        {"id": template_id, "user_id": current_user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sablon nem található")
    
    return {"message": "Sablon frissítve"}


@app.delete("/api/bulk-email/templates/{template_id}")
async def delete_email_template(
    template_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete email template"""
    result = await db.email_templates.delete_one(
        {"id": template_id, "user_id": current_user["id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sablon nem található")
    
    return {"message": "Sablon törölve"}


# Worker Templates (saved filters) CRUD
@app.get("/api/bulk-email/worker-templates")
async def get_worker_templates(current_user: dict = Depends(get_current_user)):
    """Get user's worker templates (saved filters)"""
    templates = await db.worker_templates.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return templates


@app.post("/api/bulk-email/worker-templates")
async def create_worker_template(
    template: WorkerTemplateCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new worker template (saved filter)"""
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "name": template.name,
        "filters": template.filters,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    await db.worker_templates.insert_one(doc)
    del doc["_id"]
    return doc


@app.put("/api/bulk-email/worker-templates/{template_id}")
async def update_worker_template(
    template_id: str,
    template: WorkerTemplateUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update worker template"""
    update_data = {k: v for k, v in template.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.worker_templates.update_one(
        {"id": template_id, "user_id": current_user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sablon nem található")
    
    return {"message": "Sablon frissítve"}


@app.delete("/api/bulk-email/worker-templates/{template_id}")
async def delete_worker_template(
    template_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete worker template"""
    result = await db.worker_templates.delete_one(
        {"id": template_id, "user_id": current_user["id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sablon nem található")
    
    return {"message": "Sablon törölve"}


# Campaign Management
@app.get("/api/bulk-email/campaigns")
async def get_campaigns(current_user: dict = Depends(get_current_user)):
    """Get user's email campaigns"""
    campaigns = await db.email_campaigns.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return campaigns


@app.post("/api/bulk-email/campaigns")
async def create_campaign(
    campaign: BulkEmailCampaignCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create and start a bulk email campaign"""
    # Check Gmail connection
    gmail_token = await db.gmail_tokens.find_one({"user_id": current_user["id"]})
    if not gmail_token:
        raise HTTPException(status_code=400, detail="Gmail fiók nincs kapcsolva")
    
    # Get email template if provided
    subject = campaign.subject
    body = campaign.body
    
    if campaign.email_template_id:
        template = await db.email_templates.find_one({
            "id": campaign.email_template_id,
            "user_id": current_user["id"]
        })
        if template:
            subject = template.get("subject", subject)
            body = template.get("body", body)
    
    if not subject or not body:
        raise HTTPException(status_code=400, detail="Email tárgya és szövege kötelező")
    
    # Get workers - only those with email addresses
    workers = await db.workers.find({
        "id": {"$in": campaign.worker_ids},
        "email": {"$exists": True, "$ne": "", "$ne": None}
    }).to_list(None)
    
    if not workers:
        raise HTTPException(status_code=400, detail="Nincs kiválasztott dolgozó email címmel")
    
    # Create campaign
    campaign_id = str(uuid.uuid4())
    campaign_doc = {
        "id": campaign_id,
        "user_id": current_user["id"],
        "name": campaign.name,
        "subject": subject,
        "body": body,
        "send_method": campaign.send_method,
        "total_recipients": len(workers),
        "sent_count": 0,
        "failed_count": 0,
        "pending_count": len(workers),
        "status": "queued",  # queued, in_progress, completed, paused
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    await db.email_campaigns.insert_one(campaign_doc)
    
    # Create queue entries for each worker
    queue_entries = []
    for worker in workers:
        unsubscribe_token = generate_unsubscribe_token()
        
        # Store unsubscribe token
        await db.unsubscribe_tokens.update_one(
            {"worker_id": worker["id"]},
            {
                "$set": {
                    "worker_id": worker["id"],
                    "token": unsubscribe_token,
                    "created_at": datetime.now(timezone.utc)
                }
            },
            upsert=True
        )
        
        queue_entries.append({
            "id": str(uuid.uuid4()),
            "campaign_id": campaign_id,
            "user_id": current_user["id"],
            "worker_id": worker["id"],
            "worker_name": worker.get("name"),
            "worker_email": worker.get("email"),
            "unsubscribe_token": unsubscribe_token,
            "status": "pending",  # pending, sent, failed
            "created_at": datetime.now(timezone.utc),
            "attempts": 0
        })
    
    if queue_entries:
        await db.email_queue.insert_many(queue_entries)
    
    del campaign_doc["_id"]
    return campaign_doc


@app.put("/api/bulk-email/campaigns/{campaign_id}/pause")
async def pause_campaign(
    campaign_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Pause a campaign"""
    result = await db.email_campaigns.update_one(
        {"id": campaign_id, "user_id": current_user["id"], "status": {"$in": ["queued", "in_progress"]}},
        {"$set": {"status": "paused", "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kampány nem található vagy már befejezve")
    
    return {"message": "Kampány szüneteltetve"}


@app.put("/api/bulk-email/campaigns/{campaign_id}/resume")
async def resume_campaign(
    campaign_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Resume a paused campaign"""
    result = await db.email_campaigns.update_one(
        {"id": campaign_id, "user_id": current_user["id"], "status": "paused"},
        {"$set": {"status": "queued", "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kampány nem található vagy nem szünetel")
    
    return {"message": "Kampány folytatva"}


@app.delete("/api/bulk-email/campaigns/{campaign_id}")
async def delete_campaign(
    campaign_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a campaign and its queue entries"""
    # Delete queue entries
    await db.email_queue.delete_many({"campaign_id": campaign_id})
    
    # Delete campaign
    result = await db.email_campaigns.delete_one({
        "id": campaign_id,
        "user_id": current_user["id"]
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kampány nem található")
    
    return {"message": "Kampány törölve"}


# Email queue processor (runs every minute via scheduler)
async def process_email_queue():
    """Process pending emails in queue"""
    try:
        # Get all users with pending emails
        users_with_pending = await db.email_queue.distinct("user_id", {"status": "pending"})
        
        for user_id in users_with_pending:
            # Check daily limit
            today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
            sent_today = await db.email_logs.count_documents({
                "sender_user_id": user_id,
                "sent_at": {"$gte": today_start},
                "status": "sent"
            })
            
            if sent_today >= DAILY_EMAIL_LIMIT:
                continue  # Skip this user, limit reached
            
            remaining = DAILY_EMAIL_LIMIT - sent_today
            
            # Get Gmail token
            gmail_token = await db.gmail_tokens.find_one({"user_id": user_id})
            if not gmail_token:
                continue
            
            # Get credentials
            creds, new_token_data = await get_gmail_credentials(gmail_token)
            if not creds:
                continue
            
            # Update token if refreshed
            if new_token_data:
                await db.gmail_tokens.update_one(
                    {"user_id": user_id},
                    {"$set": new_token_data}
                )
            
            # Get pending emails for this user (limited by remaining quota)
            pending_emails = await db.email_queue.find({
                "user_id": user_id,
                "status": "pending"
            }).limit(remaining).to_list(remaining)
            
            for email_entry in pending_emails:
                # Get campaign
                campaign = await db.email_campaigns.find_one({"id": email_entry["campaign_id"]})
                if not campaign or campaign.get("status") == "paused":
                    continue
                
                # Update campaign status to in_progress
                await db.email_campaigns.update_one(
                    {"id": campaign["id"], "status": "queued"},
                    {"$set": {"status": "in_progress"}}
                )
                
                # Get worker data for template replacement
                worker = await db.workers.find_one({"id": email_entry["worker_id"]})
                if not worker:
                    await db.email_queue.update_one(
                        {"id": email_entry["id"]},
                        {"$set": {"status": "failed", "error": "Worker not found"}}
                    )
                    continue
                
                # Replace template variables
                subject = replace_template_variables(campaign["subject"], worker)
                body = replace_template_variables(campaign["body"], worker)
                
                # Send email
                result = await send_email_via_gmail(
                    credentials=creds,
                    to_email=email_entry["worker_email"],
                    subject=subject,
                    body=body,
                    from_email=gmail_token.get("gmail_email"),
                    unsubscribe_token=email_entry["unsubscribe_token"]
                )
                
                if result["success"]:
                    # Update queue entry
                    await db.email_queue.update_one(
                        {"id": email_entry["id"]},
                        {
                            "$set": {
                                "status": "sent",
                                "sent_at": datetime.now(timezone.utc),
                                "message_id": result.get("message_id")
                            }
                        }
                    )
                    
                    # Log email
                    await db.email_logs.insert_one({
                        "id": str(uuid.uuid4()),
                        "campaign_id": campaign["id"],
                        "sender_user_id": user_id,
                        "worker_id": email_entry["worker_id"],
                        "worker_email": email_entry["worker_email"],
                        "subject": subject,
                        "status": "sent",
                        "sent_at": datetime.now(timezone.utc)
                    })
                    
                    # Update campaign counters
                    await db.email_campaigns.update_one(
                        {"id": campaign["id"]},
                        {
                            "$inc": {"sent_count": 1, "pending_count": -1},
                            "$set": {"updated_at": datetime.now(timezone.utc)}
                        }
                    )
                else:
                    # Update queue entry with error
                    attempts = email_entry.get("attempts", 0) + 1
                    if attempts >= 3:
                        await db.email_queue.update_one(
                            {"id": email_entry["id"]},
                            {
                                "$set": {
                                    "status": "failed",
                                    "error": result.get("error"),
                                    "attempts": attempts
                                }
                            }
                        )
                        
                        await db.email_campaigns.update_one(
                            {"id": campaign["id"]},
                            {
                                "$inc": {"failed_count": 1, "pending_count": -1},
                                "$set": {"updated_at": datetime.now(timezone.utc)}
                            }
                        )
                    else:
                        await db.email_queue.update_one(
                            {"id": email_entry["id"]},
                            {"$set": {"attempts": attempts, "last_error": result.get("error")}}
                        )
            
            # Check if campaign is completed
            for email_entry in pending_emails:
                campaign = await db.email_campaigns.find_one({"id": email_entry["campaign_id"]})
                if campaign and campaign.get("pending_count", 0) <= 0:
                    await db.email_campaigns.update_one(
                        {"id": campaign["id"]},
                        {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc)}}
                    )
    
    except Exception as e:
        logger.error(f"Error processing email queue: {e}")


# Unsubscribe endpoint (public - no auth required)
@app.get("/api/leiratkozas/{token}")
async def unsubscribe_page(token: str):
    """Handle unsubscribe - redirect to frontend page"""
    return RedirectResponse(url=f"/leiratkozas/{token}")


@app.post("/api/leiratkozas/{token}")
async def process_unsubscribe(token: str):
    """Process unsubscribe request - deletes email address"""
    # Find token
    token_doc = await db.unsubscribe_tokens.find_one({"token": token})
    if not token_doc:
        raise HTTPException(status_code=404, detail="Érvénytelen leiratkozási link")
    
    worker_id = token_doc["worker_id"]
    
    # Get worker email for response
    worker = await db.workers.find_one({"id": worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Store email for response (masked)
    old_email = worker.get("email", "")
    masked_email = old_email[:3] + "***" + old_email[-3:] if old_email and len(old_email) > 6 else "***"
    
    # Delete email address from worker
    await db.workers.update_one(
        {"id": worker_id},
        {
            "$unset": {"email": ""},
            "$set": {"unsubscribed_at": datetime.now(timezone.utc)}
        }
    )
    
    # Delete the unsubscribe token
    await db.unsubscribe_tokens.delete_one({"token": token})
    
    return {
        "success": True,
        "email": masked_email,
        "message": "Sikeresen leiratkozott az email értesítésekről."
    }


# Worker email history
@app.get("/api/workers/{worker_id}/email-history")
async def get_worker_email_history(
    worker_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get email history for a worker"""
    emails = await db.email_logs.find(
        {"worker_id": worker_id},
        {"_id": 0}
    ).sort("sent_at", -1).limit(50).to_list(50)
    return emails



# ==================== GDPR COMPLIANCE ENDPOINTS ====================

@api_router.get("/workers/{worker_id}/gdpr-export")
async def gdpr_export_worker_data(
    worker_id: str,
    user: dict = Depends(get_current_user)
):
    """
    GDPR: Jog az adatokhoz (Right to Access)
    Dolgozó összes adatának exportálása JSON formátumban
    """



# ==================== 2 ÉVES DOLGOZÓK KEZELÉSE (GDPR) ====================

@api_router.get("/admin/workers/old-workers")
async def get_old_workers(
    user: dict = Depends(get_current_user)
):
    """
    2 éves vagy régebbi dolgozók listázása (GDPR adatmegőrzés)
    - Admin látja MINDEN dolgozót
    - Recruiter csak a SAJÁT dolgozóit látja
    """
    # 2 éve (730 nap)
    two_years_ago = datetime.now(timezone.utc) - timedelta(days=730)
    two_years_ago_iso = two_years_ago.isoformat()
    
    # Query építés: admin vs recruiter
    query = {"created_at": {"$lt": two_years_ago_iso}}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]  # Recruiter csak a saját dolgozóit látja
    
    # Dolgozók akik 2+ éve lettek létrehozva
    old_workers = await db.workers.find(
        query,
        {"_id": 0}
    ).sort("created_at", 1).to_list(10000)
    
    # Részletes info minden dolgozóról
    result = []
    for w in old_workers:
        # Owner info
        owner = await db.users.find_one({"id": w["owner_id"]}, {"_id": 0, "password": 0})
        owner_name = owner.get("name") or owner["email"] if owner else "Ismeretlen"
        
        # Projekt státuszok
        project_count = await db.project_workers.count_documents({"worker_id": w["id"]})
        
        # Utolsó aktivitás (email log)
        last_email = await db.email_logs.find_one(
            {"worker_id": w["id"]},
            {"_id": 0, "sent_at": 1}
        )
        last_activity = last_email["sent_at"] if last_email else w["created_at"]
        
        # Számítsuk ki mennyi ideje
        created = datetime.fromisoformat(w["created_at"].replace('Z', '+00:00'))
        days_old = (datetime.now(timezone.utc) - created).days
        years_old = round(days_old / 365, 1)
        
        result.append({
            "id": w["id"],
            "name": w["name"],
            "phone": w["phone"],
            "email": w.get("email", ""),
            "global_status": w.get("global_status", "Feldolgozatlan"),
            "created_at": w["created_at"],
            "days_old": days_old,
            "years_old": years_old,
            "owner_id": w["owner_id"],
            "owner_name": owner_name,
            "project_count": project_count,
            "last_activity": last_activity
        })
    
    return {
        "old_workers": result,
        "total_count": len(result),
        "cutoff_date": two_years_ago_iso,
        "message": f"{len(result)} dolgozó van 2+ éves" + (" (csak te saját dolgozóid)" if user["role"] != "admin" else " (minden dolgozó)")
    }


@api_router.delete("/admin/workers/delete-old-workers")
async def delete_old_workers(
    data: dict,
    user: dict = Depends(require_admin)
):
    """
    2+ éves dolgozók TÖMEGES törlése
    FIGYELEM: Ez VÉGLEGES törlés!
    
    Body:
    {
        "worker_ids": ["id1", "id2", ...],  // Opcionális: konkrét ID-k
        "delete_all_old": true  // Ha true, akkor MINDEN 2+ éves dolgozó törlődik
    }
    """
    worker_ids = data.get("worker_ids", [])
    delete_all = data.get("delete_all_old", False)
    
    if not worker_ids and not delete_all:
        raise HTTPException(
            status_code=400,
            detail="Vagy add meg a worker_ids listát, vagy állítsd be delete_all_old=true"
        )
    
    # Ha delete_all, akkor gyűjtsük össze az összes 2+ éves dolgozót
    if delete_all:
        two_years_ago = datetime.now(timezone.utc) - timedelta(days=730)
        two_years_ago_iso = two_years_ago.isoformat()
        
        old_workers = await db.workers.find(
            {"created_at": {"$lt": two_years_ago_iso}},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1}
        ).to_list(10000)
        
        worker_ids = [w["id"] for w in old_workers]
    
    if not worker_ids:
        return {
            "success": True,
            "deleted_count": 0,
            "message": "Nincs törlendő dolgozó"
        }
    
    # Audit log MINDEN dolgozóról a törlés előtt
    for wid in worker_ids:
        worker = await db.workers.find_one({"id": wid}, {"_id": 0})
        if worker:
            await audit_logger.log(
                user_id=user["id"],
                action="old_worker_deleted",
                resource_type="worker",
                resource_id=wid,
                details={
                    "reason": "2+ éves dolgozó törlése (GDPR adatmegőrzés)",
                    "worker_name": worker.get("name"),
                    "worker_email": worker.get("email"),
                    "worker_phone": worker.get("phone"),
                    "created_at": worker.get("created_at"),
                    "deleted_by": user["email"]
                }
            )
    
    # 1. Dolgozók törlése
    deleted_workers = await db.workers.delete_many({"id": {"$in": worker_ids}})
    
    # 2. Projekt kapcsolatok törlése
    await db.project_workers.delete_many({"worker_id": {"$in": worker_ids}})
    
    # 3. Email logok törlése
    await db.email_logs.delete_many({"worker_id": {"$in": worker_ids}})
    
    # 4. Email queue törlése
    await db.email_queue.delete_many({"worker_id": {"$in": worker_ids}})
    
    # 5. Unsubscribe tokenek törlése
    await db.unsubscribe_tokens.delete_many({"worker_id": {"$in": worker_ids}})
    
    # 6. Worker logs törlése
    await db.worker_logs.delete_many({"worker_id": {"$in": worker_ids}})
    
    return {
        "success": True,
        "deleted_count": deleted_workers.deleted_count,
        "message": f"{deleted_workers.deleted_count} dolgozó és kapcsolódó adataik véglegesen törölve (2+ éves dolgozók)",
        "worker_ids": worker_ids
    }


# Automatikus emlékeztető adminnak 2+ éves dolgozókról
async def check_old_workers_notification():
    """
    Automatikus értesítés adminoknak havi egyszer (rendszerben, NEM email!)
    Ha van 2+ éves dolgozó az adatbázisban
    """
    try:
        logger.info("🔔 2+ éves dolgozók ellenőrzése...")
        
        # Minden admin user keresése
        admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(100)
        if not admins:
            logger.info("⚠️ Nincs admin user")
            return
        
        # 2 éve
        two_years_ago = datetime.now(timezone.utc) - timedelta(days=730)
        two_years_ago_iso = two_years_ago.isoformat()
        
        # Dolgozók száma akik 2+ évesek
        old_worker_count = await db.workers.count_documents({
            "created_at": {"$lt": two_years_ago_iso}
        })
        
        if old_worker_count == 0:
            logger.info("✅ Nincs 2+ éves dolgozó")
            return
        
        # Értesítés küldése MINDEN adminnak a rendszerben
        for admin in admins:
            await create_notification(
                user_id=admin["id"],
                notification_type="gdpr_warning",
                title="⚠️ GDPR Figyelmeztetés",
                message=f"{old_worker_count} dolgozó van az adatbázisban, akik 2+ éve lettek felvéve. Nézd át és töröld a felesleges adatokat!",
                link="/admin/old-workers"
            )
        
        logger.info(f"✅ GDPR értesítés elküldve {len(admins)} adminnak: {old_worker_count} dolgozó")
        
    except Exception as e:
        logger.error(f"❌ GDPR értesítés hiba: {e}", exc_info=True)

    # Jogosultság ellenőrzés
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található vagy nincs jogosultságod")
    
    # Összes kapcsolódó adat gyűjtése
    export_data = {
        "worker": worker,
        "tags": [],
        "projects": [],
        "email_history": [],
        "audit_logs": [],
        "created_by": {},
        "export_date": datetime.now(timezone.utc).isoformat()
    }
    
    # Tagek
    if worker.get("tag_ids"):
        tags = await db.tags.find({"id": {"$in": worker["tag_ids"]}}, {"_id": 0}).to_list(100)
        export_data["tags"] = tags
    
    # Projektek és státuszok
    project_workers = await db.project_workers.find(
        {"worker_id": worker_id},
        {"_id": 0}
    ).to_list(1000)
    
    for pw in project_workers:
        project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
        status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
        
        export_data["projects"].append({
            "project": project,
            "status": status,
            "added_at": pw.get("added_at"),
            "notes": pw.get("notes", "")
        })
    
    # Email történet
    emails = await db.email_logs.find(
        {"worker_id": worker_id},
        {"_id": 0}
    ).to_list(1000)
    export_data["email_history"] = emails
    
    # Audit log (ha van)
    audit_logs = await db.audit_logs.find(
        {"resource_id": worker_id},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(1000)
    export_data["audit_logs"] = audit_logs
    
    # Létrehozó user
    owner = await db.users.find_one({"id": worker["owner_id"]}, {"_id": 0, "password": 0})
    export_data["created_by"] = owner
    
    # Audit log
    await audit_logger.log(
        user_id=user["id"],
        action="gdpr_export",
        resource_type="worker",
        resource_id=worker_id,
        details={"exported_by": user["email"]}
    )
    
    return export_data


@api_router.delete("/workers/{worker_id}/gdpr-delete")
async def gdpr_delete_worker_data(
    worker_id: str,
    reason: str,
    user: dict = Depends(get_current_user)
):
    """
    GDPR: Jog a törléshez (Right to Erasure)
    Dolgozó TELJES törlése minden kapcsolódó adattal
    
    FIGYELEM: Ez VÉGLEGES törlés, nem visszaállítható!
    """
    # Csak admin törölhet GDPR alapján
    if user["role"] != "admin":
        raise HTTPException(
            status_code=403, 
            detail="GDPR törlést csak admin hajthat végre biztonsági okokból"
        )
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Audit log a törlés előtt
    await audit_logger.log(
        user_id=user["id"],
        action="gdpr_delete",
        resource_type="worker",
        resource_id=worker_id,
        details={
            "reason": reason,
            "worker_name": worker.get("name"),
            "worker_email": worker.get("email"),
            "worker_phone": worker.get("phone"),
            "deleted_by": user["email"]
        }
    )
    
    # 1. Dolgozó törlése
    await db.workers.delete_one({"id": worker_id})
    
    # 2. Projekt kapcsolatok törlése
    await db.project_workers.delete_many({"worker_id": worker_id})
    
    # 3. Email log törlése
    await db.email_logs.delete_many({"worker_id": worker_id})
    
    # 4. Email queue törlése
    await db.email_queue.delete_many({"worker_id": worker_id})
    
    # 5. Unsubscribe tokenek törlése
    await db.unsubscribe_tokens.delete_many({"worker_id": worker_id})
    
    # 6. Worker logs törlése
    await db.worker_logs.delete_many({"worker_id": worker_id})
    
    return {
        "success": True,
        "message": f"Dolgozó és minden kapcsolódó adata véglegesen törölve (GDPR)",
        "deleted_records": {
            "worker": 1,
            "reason": reason
        }
    }


@api_router.post("/workers/{worker_id}/consent")
async def update_worker_consent(
    worker_id: str,
    consent_given: bool,
    user: dict = Depends(get_current_user)
):
    """
    GDPR: Beleegyezés frissítése
    """
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    update_data = {
        "consent_given": consent_given,
        "consent_date": datetime.now(timezone.utc).isoformat() if consent_given else None
    }
    
    # Adatmegőrzés lejárata: 2 év a beleegyezéstől
    if consent_given:
        retention_date = datetime.now(timezone.utc) + timedelta(days=730)  # 2 év
        update_data["data_retention_until"] = retention_date.isoformat()
    
    await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    
    await audit_logger.log(
        user_id=user["id"],
        action="consent_updated",
        resource_type="worker",
        resource_id=worker_id,
        details={"consent_given": consent_given}
    )
    
    return {
        "success": True,
        "consent_given": consent_given,
        "data_retention_until": update_data.get("data_retention_until")
    }


@api_router.get("/gdpr/retention-check")
async def gdpr_retention_check(
    user: dict = Depends(require_admin)
):
    """
    GDPR: Adatmegőrzés lejárati ellenőrzés
    Admin látja mely dolgozók adatait kell törölni
    """
    now = datetime.now(timezone.utc).isoformat()
    
    # Dolgozók akiknek lejárt az adatmegőrzési ideje
    expired_workers = await db.workers.find(
        {
            "data_retention_until": {"$lt": now},
            "consent_given": False
        },
        {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1, "data_retention_until": 1, "owner_id": 1}
    ).to_list(1000)
    
    # Gazdátlan dolgozók (owner törlődött)
    orphaned_workers = []
    all_user_ids = [u["id"] async for u in db.users.find({}, {"id": 1})]
    orphaned_workers = await db.workers.find(
        {"owner_id": {"$nin": all_user_ids}},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1, "owner_id": 1}
    ).to_list(1000)
    
    return {
        "expired_retention_workers": expired_workers,
        "orphaned_workers": orphaned_workers,
        "total_to_review": len(expired_workers) + len(orphaned_workers)
    }




# ==================== WEEKLY SUMMARY EMAIL ====================

async def weekly_summary_email():
    """
    Heti összefoglaló email küldése az adminnak
    Megmutatja ki hány dolgozót vitt be az elmúlt 7 napban
    """
    try:
        logger.info("🔔 Heti összefoglaló email generálása...")
        
        # Admin user keresése
        admin = await db.users.find_one({"email": "kaszasdominik@gmail.com"}, {"_id": 0})
        if not admin:
            logger.warning("Admin user nem található, heti összefoglaló kihagyva")
            return
        
        # Admin Gmail token ellenőrzése
        gmail_token = await db.gmail_tokens.find_one({"user_id": admin["id"]})
        if not gmail_token:
            logger.warning("Admin Gmail token nincs, heti összefoglaló kihagyva")
            return
        
        # Elmúlt 7 nap
        week_ago = datetime.now(timezone.utc) - timedelta(days=7)
        week_ago_iso = week_ago.isoformat()
        
        # Összes user lekérése
        users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1}).to_list(100)
        
        # Statisztika gyűjtése userenként
        user_stats = []
        total_workers = 0
        
        for user in users:
            # Dolgozók száma az elmúlt 7 napban
            worker_count = await db.workers.count_documents({
                "owner_id": user["id"],
                "created_at": {"$gte": week_ago_iso}
            })
            
            if worker_count > 0:
                user_stats.append({
                    "name": user.get("name") or user["email"],
                    "email": user["email"],
                    "role": user["role"],
                    "workers_added": worker_count
                })
                total_workers += worker_count
        
        # Rendezés dolgozók száma szerint (csökkenő)
        user_stats.sort(key=lambda x: x["workers_added"], reverse=True)
        
        # Email HTML generálása
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px 10px 0 0; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 28px; }}
                .content {{ background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px; }}
                .summary {{ background: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #667eea; }}
                .summary h2 {{ margin-top: 0; color: #667eea; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 15px; background: white; border-radius: 8px; overflow: hidden; }}
                th {{ background: #667eea; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 12px; border-bottom: 1px solid #e5e7eb; }}
                tr:last-child td {{ border-bottom: none; }}
                .badge {{ display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: bold; }}
                .badge-admin {{ background: #fef3c7; color: #92400e; }}
                .badge-user {{ background: #dbeafe; color: #1e40af; }}
                .total {{ background: #10b981; color: white; padding: 15px; border-radius: 8px; text-align: center; font-size: 20px; font-weight: bold; margin-top: 20px; }}
                .footer {{ text-align: center; margin-top: 30px; color: #6b7280; font-size: 14px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 Heti Összefoglaló</h1>
                    <p style="margin: 10px 0 0 0; opacity: 0.9;">{week_ago.strftime('%Y.%m.%d')} - {datetime.now().strftime('%Y.%m.%d')}</p>
                </div>
                <div class="content">
                    <div class="summary">
                        <h2>Teljesítmény Összesítő</h2>
                        <p>Az elmúlt 7 napban összesen <strong>{total_workers} dolgozó</strong> került felvitelre az adatbázisba.</p>
                    </div>
                    
                    <h3 style="color: #374151; margin-bottom: 15px;">👥 Toborzói Teljesítmény</h3>
                    <table>
                        <thead>
                            <tr>
                                <th>Név</th>
                                <th>Szerepkör</th>
                                <th style="text-align: center;">Felvitt Dolgozók</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        if user_stats:
            for stat in user_stats:
                role_badge = "badge-admin" if stat["role"] == "admin" else "badge-user"
                role_text = "Admin" if stat["role"] == "admin" else "Toborzó"
                html_body += f"""
                            <tr>
                                <td><strong>{stat["name"]}</strong><br><small style="color: #6b7280;">{stat["email"]}</small></td>
                                <td><span class="badge {role_badge}">{role_text}</span></td>
                                <td style="text-align: center; font-size: 20px; font-weight: bold; color: #667eea;">{stat["workers_added"]}</td>
                            </tr>
                """
        else:
            html_body += """
                            <tr>
                                <td colspan="3" style="text-align: center; color: #6b7280; padding: 30px;">
                                    Nincs új dolgozó az elmúlt héten
                                </td>
                            </tr>
            """
        
        html_body += f"""
                        </tbody>
                    </table>
                    
                    <div class="total">
                        Összesen: {total_workers} dolgozó
                    </div>
                    
                    <div class="footer">
                        <p>Dolgozó CRM - Automatikus Heti Riport</p>
                        <p><small>Ez az email automatikusan lett elküldve minden hétfő reggel 8:00-kor</small></p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Email küldése a bulk_email modul használatával
        from bulk_email import send_email_via_gmail
        
        await send_email_via_gmail(
            gmail_token=gmail_token,
            to_email=admin["email"],  # kaszasdominik@gmail.com
            subject=f"📊 Heti Összefoglaló - {total_workers} új dolgozó ({week_ago.strftime('%m.%d')} - {datetime.now().strftime('%m.%d')})",
            body=html_body
        )
        
        logger.info(f"✅ Heti összefoglaló email elküldve: {admin['email']}, {total_workers} dolgozó")
        
    except Exception as e:
        logger.error(f"❌ Heti összefoglaló email hiba: {e}", exc_info=True)


@api_router.post("/admin/send-weekly-summary-now")
async def send_weekly_summary_now(user: dict = Depends(require_admin)):
    """
    Manuális heti összefoglaló küldés (teszt célból)
    Csak admin hívhatja meg
    """
    await weekly_summary_email()
    return {"success": True, "message": "Heti összefoglaló email elküldve"}


# Schedule heti összefoglaló
# Hozzáadandó a startup_event() függvényhez
